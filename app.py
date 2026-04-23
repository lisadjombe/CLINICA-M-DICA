import shutil
import os
from datetime import datetime

def crear_backup(tipo='manual'):
    """Crea un backup de la base de datos"""
    db_path = 'instance/clinica.db'
    backup_dir = 'instance/backups'
    
    if not os.path.exists(db_path):
        return None, "No hay base de datos para respaldar"
    
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = f"clinica_backup_{tipo}_{timestamp}.db"
    backup_path = os.path.join(backup_dir, backup_name)
    
    try:
        shutil.copy2(db_path, backup_path)
        return backup_path, "Backup creado exitosamente"
    except Exception as e:
        return None, f"Error al crear backup: {str(e)}"


import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Configuración de email (MODIFICA CON TUS DATOS)
EMAIL_REMITENTE = "tuclinica@gmail.com"
EMAIL_PASSWORD = "tu_contraseña_de_aplicacion"  # Usa contraseña de aplicación
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587

from flask import Flask, render_template_string, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from functools import wraps
import secrets
import io
import json
import os
from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm

app = Flask(__name__)
app.config['SECRET_KEY'] = 'clinica-medica-secreta-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///clinica.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
    
# ========== MODELOS ==========

class Usuario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    nombre_completo = db.Column(db.String(100), nullable=False)
    departamento = db.Column(db.String(50), default='General')
    cargo = db.Column(db.String(50), default='Médico')
    rol = db.Column(db.String(20), default='medico')
    fecha_registro = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y"))
    citas = db.relationship('Cita', backref='medico', lazy=True, foreign_keys='Cita.usuario_id')
    reset_token = db.Column(db.String(100), nullable=True)
    reset_token_expira = db.Column(db.String(20), nullable=True)

class Paciente(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100))
    telefono = db.Column(db.String(20))
    obra_social = db.Column(db.String(100))
    direccion = db.Column(db.String(200))
    fecha_nacimiento = db.Column(db.String(20))
    grupo_sanguineo = db.Column(db.String(5))
    alergias = db.Column(db.Text)
    notas = db.Column(db.Text)
    fecha_creacion = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y"))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    usuario = db.relationship('Usuario', backref='pacientes')

class Cita(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    fecha_cita = db.Column(db.String(20))
    hora_inicio = db.Column(db.String(5))
    hora_fin = db.Column(db.String(5))
    tipo_consulta = db.Column(db.String(50), default='Presencial')
    fecha_creacion = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y"))
    asistida = db.Column(db.Boolean, default=False)
    prioridad = db.Column(db.String(20), default='media')
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    paciente_id = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=True)
    paciente = db.relationship('Paciente', backref='citas')
    comentarios = db.relationship('Comentario', backref='cita', lazy=True, cascade='all, delete-orphan')

class TareaPersonal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    fecha_limite = db.Column(db.String(20))
    fecha_creacion = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y"))
    completada = db.Column(db.Boolean, default=False)
    prioridad = db.Column(db.String(20), default='media')
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    usuario = db.relationship('Usuario', backref='tareas_personales')

class NotaPersonal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), nullable=False)
    contenido = db.Column(db.Text)
    color = db.Column(db.String(20), default='#f39c12')
    fecha_creacion = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y %H:%M"))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    usuario = db.relationship('Usuario', backref='notas_personales')

class Comentario(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    texto = db.Column(db.Text, nullable=False)
    fecha = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y %H:%M"))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    cita_id = db.Column(db.Integer, db.ForeignKey('cita.id'), nullable=False)
    usuario = db.relationship('Usuario', backref='comentarios')

class Mensaje(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contenido = db.Column(db.Text, nullable=False)
    fecha = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y %H:%M"))
    leido = db.Column(db.Boolean, default=False)
    emisor_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    receptor_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    emisor = db.relationship('Usuario', foreign_keys=[emisor_id], backref='mensajes_enviados')
    receptor = db.relationship('Usuario', foreign_keys=[receptor_id], backref='mensajes_recibidos')

class Interaccion(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(50), default='Nota')
    descripcion = db.Column(db.Text, nullable=False)
    fecha = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y %H:%M"))
    paciente_id = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    paciente = db.relationship('Paciente', backref='interacciones')
    usuario = db.relationship('Usuario', backref='interacciones')

class MensajeGrupal(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    contenido = db.Column(db.Text, nullable=False)
    fecha = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y %H:%M"))
    departamento = db.Column(db.String(50), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    usuario = db.relationship('Usuario', backref='mensajes_grupales')

class Fichaje(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    tipo = db.Column(db.String(20), nullable=False)
    fecha_hora = db.Column(db.String(20), nullable=False)
    fecha = db.Column(db.String(20), nullable=False)
    usuario = db.relationship('Usuario', backref='fichajes')

class TipoHistoriaClinica(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    departamento = db.Column(db.String(50), default='General')
    color = db.Column(db.String(20), default='#3498db')
    admin_id = db.Column(db.Integer, db.ForeignKey('usuario.id'))
    admin = db.relationship('Usuario', backref='tipos_historia')

class HistoriaClinica(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    tipo_id = db.Column(db.Integer, db.ForeignKey('tipo_historia_clinica.id'), nullable=False)
    titulo = db.Column(db.String(200), nullable=False)
    descripcion = db.Column(db.Text)
    diagnostico = db.Column(db.Text)
    tratamiento = db.Column(db.Text)
    estado = db.Column(db.String(50), default='Abierto')
    fecha_creacion = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y"))
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    paciente_id = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=True)
    departamento = db.Column(db.String(50))
    
    tipo = db.relationship('TipoHistoriaClinica', backref='historias_clinicas')
    usuario = db.relationship('Usuario', backref='historias_clinicas')
    paciente = db.relationship('Paciente', backref='historias_clinicas')

class DocumentoClinico(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(200), nullable=False)
    nombre_original = db.Column(db.String(200))
    tipo_archivo = db.Column(db.String(50))
    tipo_documento = db.Column(db.String(50), default='Otro')
    fecha_subida = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y %H:%M"))
    historial_id = db.Column(db.Integer, db.ForeignKey('historia_clinica.id'), nullable=False)
    usuario_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    historial = db.relationship('HistoriaClinica', backref='documentos')
    usuario = db.relationship('Usuario', backref='documentos_clinicos')
    
class Receta(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre_medicamento = db.Column(db.String(200), nullable=False)
    dosis = db.Column(db.Text)
    instrucciones = db.Column(db.Text)
    fecha = db.Column(db.String(20), default=datetime.now().strftime("%d/%m/%Y"))
    historial_id = db.Column(db.Integer, db.ForeignKey('historia_clinica.id'), nullable=False)
    paciente_id = db.Column(db.Integer, db.ForeignKey('paciente.id'), nullable=True)
    medico_id = db.Column(db.Integer, db.ForeignKey('usuario.id'), nullable=False)
    
    historial = db.relationship('HistoriaClinica', backref='recetas')
    paciente = db.relationship('Paciente', backref='recetas')
    medico = db.relationship('Usuario', backref='recetas')


    # ========== FUNCIONES AUXILIARES ==========

def generar_pdf_receta(receta):
    """Genera el PDF de la receta y retorna el buffer"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    import io
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Estilo personalizado para el título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a5276'),
        alignment=1,
        spaceAfter=20
    )
    
    # Cabecera
    elements.append(Paragraph("CLÍNICA MÉDICA", title_style))
    elements.append(Paragraph(f"Receta Médica Nº {receta.id:06d}", styles['Heading2']))
    elements.append(Paragraph(f"Fecha de emisión: {receta.fecha_emision}", styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    # Datos del paciente y médico
    data = [
        ['Paciente:', receta.paciente.nombre, 'Médico:', f"Dr/a. {receta.medico.nombre_completo}"],
        ['Obra Social:', receta.paciente.obra_social or 'Particular', 'Especialidad:', receta.medico.departamento],
        ['Email:', receta.paciente.email or '-', 'Matrícula:', f"MN-{receta.medico.id:05d}"],
    ]
    table = Table(data, colWidths=[2.5*cm, 5*cm, 2.5*cm, 5*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e8f4f8')),
        ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#e8f4f8')),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    
    # Diagnóstico
    elements.append(Paragraph("<b>🔬 Diagnóstico:</b>", styles['Heading3']))
    elements.append(Paragraph(receta.diagnostico or 'No especificado', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Medicamentos
    elements.append(Paragraph("<b>💊 Medicamentos Recetados:</b>", styles['Heading3']))
    medicamentos = json.loads(receta.medicamentos) if receta.medicamentos else []
    if medicamentos:
        med_data = [['Medicamento', 'Dosis', 'Frecuencia', 'Duración']]
        for m in medicamentos:
            med_data.append([
                m.get('nombre', ''), 
                m.get('dosis', ''), 
                m.get('frecuencia', ''), 
                m.get('duracion', '')
            ])
        med_table = Table(med_data, colWidths=[4.5*cm, 3*cm, 3.5*cm, 3*cm])
        med_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(med_table)
    else:
        elements.append(Paragraph("No hay medicamentos recetados", styles['Normal']))
    
    elements.append(Spacer(1, 0.8*cm))
    
    # Indicaciones
    if receta.indicaciones:
        elements.append(Paragraph("<b>📝 Indicaciones:</b>", styles['Heading3']))
        elements.append(Paragraph(receta.indicaciones, styles['Normal']))
        elements.append(Spacer(1, 0.5*cm))
    
    # Próxima cita
    if receta.proxima_cita:
        elements.append(Paragraph(f"<b>📅 Próxima cita:</b> {receta.proxima_cita}", styles['Normal']))
        elements.append(Spacer(1, 0.8*cm))
    
    # Firma
    elements.append(Spacer(1, 1.5*cm))
    firma_data = [['_________________________', '_________________________']]
    firma_table = Table(firma_data, colWidths=[7*cm, 7*cm])
    firma_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
    ]))
    elements.append(firma_table)
    
    firma_texto = [['Firma del Médico', 'Sello de la Clínica']]
    firma_texto_table = Table(firma_texto, colWidths=[7*cm, 7*cm])
    firma_texto_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
    ]))
    elements.append(firma_texto_table)
    
    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        "Este documento es una receta médica oficial. Válido por 30 días desde la fecha de emisión.",
        styles['Normal']
    ))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer

def enviar_receta_email(destinatario, asunto, mensaje_html, pdf_buffer, nombre_archivo):
    """Envía un email con la receta en PDF adjunta"""
    try:
        # Crear mensaje
        msg = MIMEMultipart('alternative')
        msg['From'] = f"Clínica Médica <{EMAIL_REMITENTE}>"
        msg['To'] = destinatario
        msg['Subject'] = asunto
        
        # Adjuntar cuerpo HTML
        msg.attach(MIMEText(mensaje_html, 'html'))
        
        # Adjuntar PDF
        pdf_buffer.seek(0)
        part = MIMEBase('application', 'pdf')
        part.set_payload(pdf_buffer.read())
        encoders.encode_base64(part)
        part.add_header(
            'Content-Disposition',
            f'attachment; filename="{nombre_archivo}"'
        )
        msg.attach(part)
        
        # Enviar
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_REMITENTE, EMAIL_PASSWORD)
            server.send_message(msg)
        
        return True, "Email enviado correctamente"
    except Exception as e:
        return False, str(e)

def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect('/login')
        user = db.session.get(Usuario, session.get('user_id'))
        if not user:
            session.clear()
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated_function

def flash_html():
    from flask import get_flashed_messages
    messages = get_flashed_messages()
    return ''.join([f'<div class="flash flash-success">{m}</div>' for m in messages]) if messages else ''

def navbar_html():
    user_id = session.get('user_id')
    no_leidos = 0
    if user_id:
        no_leidos = Mensaje.query.filter_by(receptor_id=user_id, leido=False).count()
    
    badge_chat = f' <span style="background:#e74c3c;color:white;padding:2px 8px;border-radius:10px;">{no_leidos}</span>' if no_leidos > 0 else ''
    
    return f"""
    <div class="navbar">
        <span><strong>🏥 CLÍNICA MÉDICA</strong> | 👨‍⚕️ {session.get('nombre')} ({session.get('rol')})</span>
        <div style="display: flex; flex-wrap: wrap; gap: 5px; align-items: center;">
            <a href="/dashboard">📊 Dashboard</a>
            <a href="/mi-espacio">📌 Mi Espacio</a>
            <a href="/pacientes">🏥 Pacientes</a>
            <a href="/citas">📅 Citas</a>
            <a href="/recetas">📋 Recetas</a>
            <a href="/historias-clinicas">📁 Historias Clínicas</a>
            <a href="/chat">💬 Chat{badge_chat}</a>
            <a href="/chat-grupal">👥 Chat Grupal</a>
            <a href="/control-horario">⏰ Fichar</a>
            {'''<a href="/admin/medicos">👨‍⚕️ Médicos</a>
            <a href="/admin/panel-horario">⏰ Panel Horario</a>
            <a href="/admin/asignar-cita">📋 Asignar Cita</a>
            <a href="/admin/informe-mensual">📊 Informe PDF</a>
            <a href="/admin/tipos-historia">⚙️ Tipos Historia</a>''' if session.get('rol') == 'admin' else ''}
            <a href="/admin/backups">💾 Backups</a>
            <a href="/logout" style="background: #e74c3c; color: white; padding: 6px 12px; border-radius: 20px; font-size: 12px; text-decoration: none;">🚪</a>
        </div>
    </div>
    """

def base_html(content, titulo="Clínica Médica"):
    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <title>{titulo}</title>
        <meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, user-scalable=no">
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <style>
                      * {{ 
                margin: 0; 
                padding: 0; 
                box-sizing: border-box; 
            }}
            
            body {{ 
                font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Helvetica, Arial, sans-serif;
                background: linear-gradient(135deg, #e8f4f8 0%, #d1e9f0 100%);
                min-height: 100vh;
                padding: 20px;
                margin: 0;
            }}
            
            /* ========== NAVBAR PROFESIONAL ========== */
            .navbar {{ 
                background: linear-gradient(90deg, #0d3b66 0%, #1a5276 100%);
                color: white;
                padding: 16px 30px;
                margin-bottom: 30px;
                display: flex;
                flex-wrap: wrap;
                justify-content: space-between;
                align-items: center;
                gap: 20px;
                border-radius: 16px;
                box-shadow: 0 8px 20px rgba(13, 59, 102, 0.2);
                border-bottom: 3px solid #27ae60;
            }}
            
            .navbar span {{ 
                font-size: 18px;
                font-weight: 600;
                letter-spacing: 0.5px;
            }}
            
            .navbar span strong {{ 
                background: white;
                color: #0d3b66;
                padding: 6px 14px;
                border-radius: 30px;
                margin-right: 10px;
                font-size: 16px;
            }}
            
            .navbar a {{ 
                color: white;
                margin: 0 3px;
                text-decoration: none;
                font-size: 14px;
                font-weight: 500;
                white-space: nowrap;
                padding: 8px 14px;
                border-radius: 30px;
                transition: all 0.25s ease;
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(255, 255, 255, 0.1);
            }}
            
            .navbar a:hover {{ 
                background: rgba(255, 255, 255, 0.2);
                border-color: rgba(255, 255, 255, 0.3);
                transform: translateY(-2px);
            }}
            
            .navbar div {{ 
                display: flex;
                flex-wrap: wrap;
                gap: 5px;
                align-items: center;
            }}
            
            /* ========== CONTENEDOR PRINCIPAL ========== */
            .container {{ 
                max-width: 1400px;
                margin: 0 auto;
                background: rgba(255, 255, 255, 0.98);
                backdrop-filter: blur(10px);
                padding: 35px;
                border-radius: 24px;
                box-shadow: 0 20px 40px rgba(0, 0, 0, 0.08), 0 4px 12px rgba(0, 0, 0, 0.05);
                width: 100%;
                overflow-x: auto;
                word-wrap: break-word;
                border: 1px solid rgba(39, 174, 96, 0.15);
            }}
            
            /* ========== TIPOGRAFÍA ========== */
            h1, h2, h3, h4, h5, h6 {{
                font-weight: 600;
                letter-spacing: -0.02em;
                color: #0d3b66;
            }}
            
            h2 {{ 
                font-size: 28px;
                margin-bottom: 10px;
                position: relative;
                display: inline-block;
            }}
            
            h2:after {{
                content: '';
                position: absolute;
                bottom: -8px;
                left: 0;
                width: 60px;
                height: 4px;
                background: linear-gradient(90deg, #27ae60, #1a5276);
                border-radius: 4px;
            }}
            
            /* ========== TABLAS PROFESIONALES ========== */
            table {{ 
                width: 100%;
                border-collapse: separate;
                border-spacing: 0;
                margin: 25px 0;
                border-radius: 16px;
                overflow: hidden;
                box-shadow: 0 4px 12px rgba(0, 0, 0, 0.06);
            }}
            
            th, td {{ 
                padding: 16px 18px; 
                text-align: left; 
                border-bottom: 1px solid #e9ecef;
            }}
            
            th {{ 
                background: linear-gradient(90deg, #0d3b66 0%, #1a5276 100%);
                color: white;
                font-weight: 600;
                font-size: 14px;
                text-transform: uppercase;
                letter-spacing: 0.5px;
            }}
            
            tr {{ 
                background: white;
                transition: background 0.2s ease;
            }}
            
            tr:hover {{ 
                background: #f0f9f4;
            }}
            
            tr:last-child td {{ 
                border-bottom: none;
            }}
            
            /* ========== BOTONES ========== */
            .btn {{ 
                padding: 11px 22px; 
                border: none; 
                border-radius: 12px; 
                cursor: pointer; 
                text-decoration: none; 
                display: inline-block; 
                margin: 3px; 
                font-size: 14px; 
                font-weight: 600;
                transition: all 0.25s ease;
                box-shadow: 0 2px 6px rgba(0, 0, 0, 0.08);
                letter-spacing: 0.3px;
                border: 1px solid transparent;
            }}
            
            .btn:hover {{ 
                transform: translateY(-2px);
                box-shadow: 0 6px 14px rgba(0, 0, 0, 0.12);
            }}
            
            .btn-primary {{ 
                background: linear-gradient(135deg, #1a5276 0%, #0d3b66 100%);
                color: white;
                border: 1px solid rgba(255, 255, 255, 0.1);
            }}
            
            .btn-primary:hover {{
                background: linear-gradient(135deg, #0d3b66 0%, #092a4a 100%);
            }}
            
            .btn-success {{ 
                background: linear-gradient(135deg, #27ae60 0%, #1e8c4c 100%);
                color: white;
            }}
            
            .btn-success:hover {{
                background: linear-gradient(135deg, #1e8c4c 0%, #166b3a 100%);
            }}
            
            .btn-warning {{ 
                background: linear-gradient(135deg, #f39c12 0%, #e08e0b 100%);
                color: white;
            }}
            
            .btn-danger {{ 
                background: linear-gradient(135deg, #e74c3c 0%, #c0392b 100%);
                color: white;
            }}
            
            .btn-sm {{ 
                padding: 7px 14px; 
                font-size: 12px;
                border-radius: 8px;
            }}
            
            /* ========== FORMULARIOS ========== */
            .form-group {{ 
                margin-bottom: 22px; 
            }}
            
            .form-group label {{
                font-weight: 600;
                color: #0d3b66;
                margin-bottom: 8px;
                display: block;
                font-size: 14px;
                letter-spacing: 0.3px;
            }}
            
            .form-control {{ 
                width: 100%; 
                padding: 14px 16px; 
                border: 2px solid #e0e7ef;
                border-radius: 14px; 
                font-size: 15px;
                transition: all 0.25s ease;
                background: white;
                color: #2c3e50;
            }}
            
            .form-control:focus {{ 
                border-color: #1a5276;
                outline: none;
                box-shadow: 0 0 0 4px rgba(26, 82, 118, 0.1);
            }}
            
            select.form-control {{
                cursor: pointer;
                background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='16' height='16' viewBox='0 0 24 24' fill='none' stroke='%230d3b66' stroke-width='2' stroke-linecap='round' stroke-linejoin='round'%3E%3Cpolyline points='6 9 12 15 18 9'%3E%3C/polyline%3E%3C/svg%3E");
                background-repeat: no-repeat;
                background-position: right 16px center;
                appearance: none;
                padding-right: 45px;
            }}
            
            /* ========== TARJETAS (CARDS) ========== */
            .card {{ 
                background: white;
                border-radius: 20px;
                padding: 28px;
                box-shadow: 0 4px 16px rgba(0, 0, 0, 0.06);
                margin-bottom: 25px;
                border: 1px solid #eef2f6;
                transition: all 0.25s ease;
            }}
            
            .card:hover {{
                box-shadow: 0 8px 24px rgba(0, 0, 0, 0.1);
                border-color: #d1e0e8;
            }}
            
            .card h3 {{ 
                margin-bottom: 20px;
                color: #0d3b66;
                border-bottom: 2px solid #e8f4f8;
                padding-bottom: 14px;
                font-size: 20px;
            }}
            
            /* ========== BADGES ========== */
            .badge {{ 
                padding: 5px 12px; 
                border-radius: 30px; 
                font-size: 12px; 
                font-weight: 600;
                letter-spacing: 0.3px;
            }}
            
            .badge-admin {{ 
                background: linear-gradient(135deg, #e74c3c, #c0392b);
                color: white;
            }}
            
            .badge-medico {{ 
                background: linear-gradient(135deg, #1a5276, #0d3b66);
                color: white;
            }}
            
            .badge-alta {{ 
                background: #fdf0ed;
                color: #e74c3c;
                border: 1px solid #f5c6cb;
            }}
            
            .badge-media {{ 
                background: #fef9e7;
                color: #f39c12;
                border: 1px solid #fdebd0;
            }}
            
            .badge-baja {{ 
                background: #eafaf1;
                color: #27ae60;
                border: 1px solid #d5f5e3;
            }}
            
            /* ========== ESTADÍSTICAS ========== */
            .stats {{ 
                display: grid; 
                grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); 
                gap: 25px; 
                margin-bottom: 35px; 
            }}
            
            .stat-card {{ 
                background: linear-gradient(135deg, #1a5276 0%, #0d3b66 100%);
                padding: 28px 20px;
                border-radius: 20px;
                text-align: center;
                color: white;
                box-shadow: 0 10px 20px rgba(13, 59, 102, 0.25);
                transition: all 0.3s ease;
                border: 1px solid rgba(255, 255, 255, 0.1);
            }}
            
            .stat-card:hover {{
                transform: translateY(-5px);
                box-shadow: 0 15px 30px rgba(13, 59, 102, 0.35);
            }}
            
            .stat-number {{ 
                font-size: 42px; 
                font-weight: 700;
                line-height: 1.2;
            }}
            
            .stat-label {{ 
                opacity: 0.9; 
                font-size: 15px;
                margin-top: 8px;
                font-weight: 500;
            }}
            
            /* ========== TABS ========== */
            .tabs {{ 
                display: flex; 
                gap: 8px; 
                margin-bottom: 30px; 
                border-bottom: 2px solid #e0e7ef;
                padding-bottom: 5px;
            }}
            
            .tab {{ 
                padding: 14px 28px; 
                cursor: pointer; 
                border: none; 
                background: transparent;
                font-size: 16px; 
                font-weight: 600;
                color: #6c7a89;
                transition: all 0.25s ease;
                border-radius: 30px 30px 0 0;
                position: relative;
            }}
            
            .tab:hover {{ 
                color: #1a5276;
                background: rgba(26, 82, 118, 0.05);
            }}
            
            .tab.active {{ 
                color: #1a5276;
                font-weight: 700;
            }}
            
            .tab.active:after {{
                content: '';
                position: absolute;
                bottom: -7px;
                left: 0;
                right: 0;
                height: 4px;
                background: linear-gradient(90deg, #27ae60, #1a5276);
                border-radius: 4px 4px 0 0;
            }}
            
            .tab-content {{ 
                display: none;
                animation: fadeIn 0.4s ease;
            }}
            
            .tab-content.active {{ 
                display: block;
            }}
            
                        /* ========== CHAT - VERSIÓN CORREGIDA ========== */
            .chat-container {{
                display: flex;
                height: 500px;
                border: 1px solid #ddd;
                border-radius: 16px;
                overflow: hidden;
                background: white;
            }}
            
            .chat-sidebar {{
                width: 280px;
                background: #f5f7fa;
                border-right: 1px solid #e0e0e0;
                overflow-y: auto;
            }}
            
            .chat-sidebar-item {{
                padding: 15px 18px;
                border-bottom: 1px solid #e0e0e0;
                cursor: pointer;
                transition: background 0.2s;
                font-weight: 500;
                color: #333;
            }}
            
            .chat-sidebar-item:hover {{
                background: #e8ecf1;
            }}
            
            .chat-sidebar-item.active {{
                background: #1a5276;
                color: white;
            }}
            
            .chat-main {{
                flex: 1;
                display: flex;
                flex-direction: column;
                background: #fafafa;
            }}
            
            .chat-header {{
                padding: 16px 20px;
                background: white;
                border-bottom: 1px solid #e0e0e0;
                font-weight: 600;
                color: #1a5276;
            }}
            
            .chat-messages {{
                flex: 1;
                padding: 20px;
                overflow-y: auto;
                display: flex;
                flex-direction: column;
                gap: 12px;
            }}
            
            .chat-message {{
                display: flex;
                margin-bottom: 5px;
            }}
            
            .chat-message.sent {{
                justify-content: flex-end;
            }}
            
            .chat-message-bubble {{
                max-width: 70%;
                padding: 12px 16px;
                border-radius: 18px;
                background: white;
                box-shadow: 0 1px 2px rgba(0,0,0,0.1);
                word-wrap: break-word;
            }}
            
            .chat-message.sent .chat-message-bubble {{
                background: #1a5276;
                color: white;
            }}
            
            .chat-message-time {{
                font-size: 11px;
                color: #999;
                margin-top: 5px;
            }}
            
            .chat-message.sent .chat-message-time {{
                color: #d4e6f1;
            }}
            
            .chat-input {{
                padding: 15px 20px;
                background: white;
                border-top: 1px solid #e0e0e0;
                display: flex;
                gap: 10px;
            }}
            
            .chat-input input {{
                flex: 1;
                padding: 12px 16px;
                border: 1px solid #ddd;
                border-radius: 25px;
                font-size: 14px;
            }}
            
            .chat-input input:focus {{
                outline: none;
                border-color: #1a5276;
            }}
            
            .chat-input button {{
                padding: 12px 24px;
                background: #1a5276;
                color: white;
                border: none;
                border-radius: 25px;
                font-weight: 600;
                cursor: pointer;
            }}
            
            .chat-input button:hover {{
                background: #0d3b66;
            }}
            
            /* ========== PACIENTES GRID ========== */
            .pacientes-grid {{ 
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(340px, 1fr));
                gap: 25px;
                margin: 25px 0;
            }}
            
            .paciente-card {{ 
                background: white;
                border-radius: 20px;
                padding: 0;
                box-shadow: 0 4px 16px rgba(0, 0, 0, 0.06);
                border: 1px solid #eef2f6;
                transition: all 0.3s ease;
                overflow: hidden;
            }}
            
            .paciente-card:hover {{ 
                transform: translateY(-6px);
                box-shadow: 0 12px 28px rgba(0, 0, 0, 0.12);
                border-color: #27ae60;
            }}
            
            .paciente-header {{ 
                display: flex;
                align-items: center;
                gap: 16px;
                padding: 20px;
                background: linear-gradient(135deg, #f8fbfc, white);
                border-bottom: 1px solid #eef2f6;
            }}
            
            .paciente-avatar {{ 
                width: 60px;
                height: 60px;
                background: linear-gradient(135deg, #1a5276, #0d3b66);
                color: white;
                border-radius: 18px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 24px;
                font-weight: 700;
                box-shadow: 0 4px 10px rgba(13, 59, 102, 0.2);
            }}
            
            .paciente-info h3 {{ 
                margin: 0;
                color: #0d3b66;
                font-size: 18px;
            }}
            
            .paciente-info p {{ 
                margin: 5px 0 0;
                color: #6c7a89;
                font-size: 14px;
            }}
            
            .paciente-body {{ 
                padding: 20px;
            }}
            
            .paciente-contacto {{ 
                display: flex;
                flex-direction: column;
                gap: 8px;
                margin-bottom: 15px;
            }}
            
            .paciente-contacto span {{ 
                display: flex;
                align-items: center;
                gap: 10px;
                color: #2c3e50;
                font-size: 14px;
            }}
            
            .paciente-footer {{ 
                display: flex;
                justify-content: space-between;
                align-items: center;
                padding: 16px 20px;
                background: #fafcfd;
                border-top: 1px solid #eef2f6;
            }}
            
            /* ========== HISTORIAS CLÍNICAS ========== */
            .historias-grid {{ 
                display: grid;
                grid-template-columns: repeat(auto-fill, minmax(320px, 1fr));
                gap: 25px;
                margin: 25px 0;
            }}
            
            .historia-card {{ 
                background: white;
                border-radius: 20px;
                padding: 0;
                box-shadow: 0 4px 16px rgba(0, 0, 0, 0.06);
                border-top: 5px solid #1a5276;
                transition: all 0.3s ease;
                overflow: hidden;
            }}
            
            .historia-card:hover {{ 
                transform: translateY(-6px);
                box-shadow: 0 12px 28px rgba(0, 0, 0, 0.12);
            }}
            
            .historia-header {{ 
                display: flex;
                justify-content: space-between;
                align-items: flex-start;
                padding: 20px 20px 10px;
            }}
            
            .historia-tipo {{ 
                padding: 6px 16px;
                border-radius: 30px;
                font-size: 13px;
                font-weight: 600;
                color: white;
            }}
            
            .historia-estado {{ 
                font-size: 12px;
                padding: 5px 12px;
                border-radius: 30px;
                font-weight: 600;
            }}
            
            .estado-abierto {{ 
                background: #eafaf1;
                color: #27ae60;
            }}
            
            .estado-proceso {{ 
                background: #fef9e7;
                color: #f39c12;
            }}
            
            .estado-cerrado {{ 
                background: #fdf0ed;
                color: #e74c3c;
            }}
            
            /* ========== CONTROL HORARIO ========== */
            .reloj-container {{ 
                background: linear-gradient(135deg, #0d3b66 0%, #1a2a4a 100%);
                color: white;
                padding: 45px;
                border-radius: 28px;
                text-align: center;
                margin-bottom: 30px;
                box-shadow: 0 20px 30px rgba(0, 0, 0, 0.15);
            }}
            
            .reloj-tiempo {{ 
                font-size: 80px;
                font-weight: 700;
                font-family: 'SF Mono', 'Monaco', 'Inconsolata', 'Fira Code', monospace;
                margin: 25px 0;
                text-shadow: 0 4px 20px rgba(0, 0, 0, 0.2);
            }}
            
            .reloj-estado {{ 
                font-size: 20px;
                opacity: 0.95;
                margin-bottom: 20px;
                font-weight: 500;
            }}
            
            .fichaje-buttons {{ 
                display: flex;
                gap: 20px;
                justify-content: center;
                flex-wrap: wrap;
            }}
            
            .btn-fichaje {{ 
                padding: 18px 40px;
                font-size: 18px;
                border: none;
                border-radius: 50px;
                cursor: pointer;
                font-weight: 700;
                transition: all 0.25s ease;
                box-shadow: 0 6px 16px rgba(0, 0, 0, 0.2);
            }}
            
            .btn-fichaje:hover {{
                transform: translateY(-3px);
                box-shadow: 0 10px 24px rgba(0, 0, 0, 0.25);
            }}
            
            .btn-entrada {{ 
                background: linear-gradient(135deg, #27ae60, #1e8c4c);
                color: white;
            }}
            
            .btn-salida {{ 
                background: linear-gradient(135deg, #e74c3c, #c0392b);
                color: white;
            }}
            
            .btn-pausa {{ 
                background: linear-gradient(135deg, #f39c12, #e08e0b);
                color: white;
            }}
            
            .btn-fichaje:disabled {{ 
                opacity: 0.5;
                cursor: not-allowed;
                transform: none;
                box-shadow: none;
            }}
            
            /* ========== FLASH MESSAGES ========== */
            .flash {{ 
                padding: 18px 22px;
                margin: 20px 0;
                border-radius: 16px;
                animation: slideDown 0.4s ease;
                font-weight: 500;
            }}
            
            .flash-success {{ 
                background: linear-gradient(135deg, #eafaf1, #d5f5e3);
                color: #166b3a;
                border-left: 6px solid #27ae60;
            }}
            
            /* ========== ANIMACIONES ========== */
            @keyframes fadeIn {{ 
                from {{ opacity: 0; transform: translateY(10px); }}
                to {{ opacity: 1; transform: translateY(0); }}
            }}
            
            @keyframes slideDown {{
                from {{ opacity: 0; transform: translateY(-20px); }}
                to {{ opacity: 1; transform: translateY(0); }}
            }}


                    /* ========== CALENDARIO MEJORADO ========== */
            .calendario-semana {{
                display: grid;
                grid-template-columns: repeat(7, 1fr);
                gap: 12px;
                margin: 30px 0;
            }}
            
            .calendario-dia {{
                background: white;
                border-radius: 20px;
                padding: 18px 12px;
                box-shadow: 0 5px 15px rgba(0, 0, 0, 0.05);
                border: 1px solid #eef2f6;
                transition: all 0.25s ease;
                min-height: 200px;
            }}
            
            .calendario-dia:hover {{
                transform: translateY(-5px);
                box-shadow: 0 15px 30px rgba(26, 82, 118, 0.1);
                border-color: #1a5276;
            }}
            
            .calendario-dia.hoy {{
                background: linear-gradient(135deg, #fff9e6 0%, #ffffff 100%);
                border: 2px solid #27ae60;
                box-shadow: 0 8px 20px rgba(39, 174, 96, 0.15);
            }}
            
            .calendario-dia-header {{
                text-align: center;
                margin-bottom: 15px;
                padding-bottom: 10px;
                border-bottom: 2px solid #e0e7ef;
            }}
            
            .calendario-dia-header strong {{
                font-size: 16px;
                color: #0d3b66;
                display: block;
            }}
            
            .calendario-dia-header span {{
                font-size: 24px;
                font-weight: 700;
                color: #1a5276;
            }}
            
            .calendario-tareas {{
                display: flex;
                flex-direction: column;
                gap: 6px;
            }}
            
            .calendario-tarea {{
                background: linear-gradient(90deg, #f8fbfc, white);
                padding: 8px 10px;
                border-radius: 10px;
                font-size: 12px;
                border-left: 4px solid;
                box-shadow: 0 2px 5px rgba(0, 0, 0, 0.03);
                transition: all 0.2s;
                cursor: pointer;
                animation: popIn 0.3s ease;
            }}
            
            .calendario-tarea:hover {{
                transform: scale(1.02);
                box-shadow: 0 5px 12px rgba(0, 0, 0, 0.08);
            }}
            
            .calendario-tarea.alta {{ border-left-color: #e74c3c; }}
            .calendario-tarea.media {{ border-left-color: #f39c12; }}
            .calendario-tarea.baja {{ border-left-color: #27ae60; }}
            
            @keyframes popIn {{
                from {{
                    opacity: 0;
                    transform: scale(0.8);
                }}
                to {{
                    opacity: 1;
                    transform: scale(1);
                }}
            }}


            
            /* ========== RESPONSIVE ========== */
            @media (max-width: 768px) {{ 
                body {{ padding: 10px; }}
                
                .container {{ 
                    padding: 20px;
                    border-radius: 20px;
                }}
                
                .navbar {{ 
                    flex-direction: column;
                    gap: 15px;
                    text-align: center;
                    padding: 15px 20px;
                }}
                
                .navbar div {{ 
                    justify-content: center;
                }}
                
                .stats {{ 
                    grid-template-columns: 1fr 1fr;
                }}
                
                .reloj-tiempo {{ 
                    font-size: 50px;
                }}
                
                .btn-fichaje {{ 
                    padding: 14px 24px;
                    font-size: 16px;
                }}
            }}
            
            @media (max-width: 480px) {{ 
                h1 {{ font-size: 24px; }}
                h2 {{ font-size: 22px; }}
                
                .stats {{ 
                    grid-template-columns: 1fr;
                }}
                
                .navbar a {{ 
                    padding: 6px 10px;
                    font-size: 13px;
                }}
                
                .stat-number {{ 
                    font-size: 32px;
                }}
            }}
        </style>
    </head>
    <body>
        {navbar_html() if session.get('user_id') else ''}
        <div class="container">
            {flash_html()}
            {content}
        </div>
    </body>
    </html>
    """


    # ========== RUTAS DE AUTENTICACIÓN ==========

@app.route('/')
def index():
    if 'user_id' in session:
        return redirect('/dashboard')
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = Usuario.query.filter_by(username=request.form['username']).first()
        if user and check_password_hash(user.password_hash, request.form['password']):
            session['user_id'] = user.id
            session['username'] = user.username
            session['nombre'] = user.nombre_completo
            session['rol'] = user.rol
            flash(f'✅ Bienvenido/a Dr/a. {user.nombre_completo}')
            return redirect('/dashboard')
        flash('❌ Usuario o contraseña incorrectos')
    
    content = """
        <div style="max-width: 450px; margin: 50px auto;">
            <div style="text-align: center; margin-bottom: 30px;">
                <h1 style="color: #1a5276; font-size: 32px;">🏥 Clínica Médica</h1>
                <p style="color: #666;">Sistema de Gestión Integral</p>
            </div>
            <div class="card">
                <h2 style="text-align: center; margin-bottom: 20px;">🔐 Acceso al Sistema</h2>
                <form method="POST">
                    <div class="form-group">
                        <label>👤 Usuario</label>
                        <input type="text" name="username" class="form-control" placeholder="Ej: admin" required autofocus>
                    </div>
                    <div class="form-group">
                        <label>🔑 Contraseña</label>
                        <input type="password" name="password" class="form-control" placeholder="••••••••" required>
                    </div>
                    <button type="submit" class="btn btn-primary" style="width: 100%; padding: 12px; font-size: 16px;">🚪 Entrar al Sistema</button>
                </form>
                <div style="margin-top: 20px; text-align: center;">
                    <a href="/registro" style="color: #2980b9; text-decoration: none;">📝 Registrar nuevo médico</a>
                    <br><br>
                    <a href="/olvide-password" style="color: #999; text-decoration: none; font-size: 14px;">¿Olvidaste tu contraseña?</a>
                </div>
            </div>
            <div style="text-align: center; margin-top: 20px; color: #999; font-size: 12px;">
                <p>👨‍⚕️ Admin por defecto: admin / admin123</p>
            </div>
        </div>
    """
    return base_html(content, "Login - Clínica Médica")

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        if Usuario.query.filter_by(username=request.form['username']).first():
            flash('❌ El usuario ya existe')
        else:
            es_admin = Usuario.query.count() == 0
            user = Usuario(
                username=request.form['username'],
                password_hash=generate_password_hash(request.form['password'], method='pbkdf2:sha256'),
                nombre_completo=request.form['nombre_completo'],
                departamento=request.form.get('departamento', 'Medicina General'),
                cargo=request.form.get('cargo', 'Médico'),
                rol='admin' if es_admin else 'medico'
            )
            db.session.add(user)
            db.session.commit()
            
            if 'user_id' not in session:
                session['user_id'] = user.id
                session['username'] = user.username
                session['nombre'] = user.nombre_completo
                session['rol'] = user.rol
            
            flash('✅ Médico registrado correctamente')
            return redirect('/dashboard')
    
    content = """
        <div style="max-width: 600px; margin: 30px auto;">
            <div class="card">
                <h2>📝 Registrar Nuevo Profesional</h2>
                <form method="POST">
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                        <div class="form-group">
                            <label>👤 Usuario *</label>
                            <input type="text" name="username" class="form-control" required>
                        </div>
                        <div class="form-group">
                            <label>🔑 Contraseña *</label>
                            <input type="password" name="password" class="form-control" required>
                        </div>
                    </div>
                    <div class="form-group">
                        <label>👨‍⚕️ Nombre Completo *</label>
                        <input type="text" name="nombre_completo" class="form-control" placeholder="Ej: Dr. Juan Pérez" required>
                    </div>
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                        <div class="form-group">
                            <label>🏥 Especialidad / Departamento</label>
                            <select name="departamento" class="form-control">
                                <option value="Medicina General">Medicina General</option>
                                <option value="Cardiología">Cardiología</option>
                                <option value="Pediatría">Pediatría</option>
                                <option value="Traumatología">Traumatología</option>
                                <option value="Ginecología">Ginecología</option>
                                <option value="Dermatología">Dermatología</option>
                                <option value="Oftalmología">Oftalmología</option>
                                <option value="Psiquiatría">Psiquiatría</option>
                                <option value="Administración">Administración</option>
                            </select>
                        </div>
                        <div class="form-group">
                            <label>💼 Cargo</label>
                            <input type="text" name="cargo" class="form-control" value="Médico">
                        </div>
                    </div>
                    <button type="submit" class="btn btn-success" style="width: 100%; padding: 12px; font-size: 16px;">✅ Registrar Profesional</button>
                    <a href="/login" class="btn" style="background: #95a5a6; color: white; width: 100%; margin-top: 10px; text-align: center;">← Volver al Login</a>
                </form>
            </div>
        </div>
    """
    return base_html(content, "Registro - Clínica Médica")

@app.route('/olvide-password', methods=['GET', 'POST'])
def olvide_password():
    if request.method == 'POST':
        username = request.form.get('username')
        user = Usuario.query.filter_by(username=username).first()
        
        if user:
            token = secrets.token_urlsafe(32)
            user.reset_token = token
            user.reset_token_expira = (datetime.now() + timedelta(hours=1)).strftime("%d/%m/%Y %H:%M")
            db.session.commit()
            
            reset_url = url_for('reset_password', token=token, _external=True)
            flash(f'🔗 Link de recuperación (válido por 1 hora): {reset_url}')
        else:
            flash('❌ Usuario no encontrado')
    
    content = """
        <div style="max-width: 450px; margin: 50px auto;">
            <div class="card">
                <h2>🔐 Recuperar Contraseña</h2>
                <p style="color: #666; margin-bottom: 20px;">Ingresa tu nombre de usuario y te enviaremos un link para restablecer tu contraseña.</p>
                <form method="POST">
                    <div class="form-group">
                        <label>👤 Usuario</label>
                        <input type="text" name="username" class="form-control" required>
                    </div>
                    <button type="submit" class="btn btn-primary" style="width: 100%;">Enviar Link de Recuperación</button>
                </form>
                <p style="margin-top: 20px; text-align: center;">
                    <a href="/login">← Volver al login</a>
                </p>
            </div>
        </div>
    """
    return base_html(content, "Recuperar Contraseña")

@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    user = Usuario.query.filter_by(reset_token=token).first()
    
    if not user:
        flash('❌ Link inválido o expirado')
        return redirect('/login')
    
    expira = datetime.strptime(user.reset_token_expira, "%d/%m/%Y %H:%M")
    if datetime.now() > expira:
        flash('❌ El link ha expirado')
        return redirect('/login')
    
    if request.method == 'POST':
        password = request.form.get('password')
        password_confirm = request.form.get('password_confirm')
        
        if password != password_confirm:
            flash('❌ Las contraseñas no coinciden')
            return redirect(request.url)
        
        user.password_hash = generate_password_hash(password, method='pbkdf2:sha256')
        user.reset_token = None
        user.reset_token_expira = None
        db.session.commit()
        flash('✅ Contraseña actualizada. Ya puedes iniciar sesión.')
        return redirect('/login')
    
    content = f"""
        <div style="max-width: 450px; margin: 50px auto;">
            <div class="card">
                <h2>🔐 Nueva Contraseña</h2>
                <p style="color: #666; margin-bottom: 20px;">Usuario: <strong>{user.nombre_completo}</strong></p>
                <form method="POST">
                    <div class="form-group">
                        <label>🔑 Nueva contraseña</label>
                        <input type="password" name="password" class="form-control" required minlength="6">
                    </div>
                    <div class="form-group">
                        <label>🔑 Confirmar contraseña</label>
                        <input type="password" name="password_confirm" class="form-control" required minlength="6">
                    </div>
                    <button type="submit" class="btn btn-success" style="width: 100%;">Guardar Nueva Contraseña</button>
                </form>
            </div>
        </div>
    """
    return base_html(content, "Nueva Contraseña")

@app.route('/logout')
def logout():
    session.clear()
    flash('👋 Has cerrado sesión correctamente')
    return redirect('/login')




# ========== DASHBOARD ==========

@app.route('/dashboard')
@login_required
def dashboard():
    user = db.session.get(Usuario, session.get('user_id'))
    hoy = datetime.now()
    
    if session.get('rol') == 'admin':
        # Dashboard de ADMIN
        total_medicos = Usuario.query.filter_by(rol='medico').count()
        total_pacientes = Paciente.query.count()
        total_citas = Cita.query.count()
        citas_pendientes = Cita.query.filter_by(asistida=False).count()
        citas_completadas = Cita.query.filter_by(asistida=True).count()
        
        # Pacientes nuevos este mes
        inicio_mes = hoy.replace(day=1).strftime('%d/%m/%Y')
        pacientes_mes = Paciente.query.filter(Paciente.fecha_creacion >= inicio_mes).count()
        
        # Citas por especialidad
        especialidades = db.session.query(Usuario.departamento, db.func.count(Cita.id)).join(Cita).filter(Cita.asistida == False).group_by(Usuario.departamento).all()
        
        # Top 5 médicos por productividad
        medicos = Usuario.query.filter_by(rol='medico').all()
        ranking = []
        for m in medicos:
            total = len(m.citas)
            completadas = sum(1 for c in m.citas if c.asistida)
            prod = int((completadas / total * 100)) if total > 0 else 0
            ranking.append((m.nombre_completo, prod, completadas, total, m.departamento))
        ranking.sort(key=lambda x: x[1], reverse=True)
        top5 = ranking[:5]
        
        # Citas urgentes (hoy/mañana)
        hoy_str = hoy.strftime('%Y-%m-%d')
        manana = (hoy + timedelta(days=1)).strftime('%Y-%m-%d')
        citas_urgentes = Cita.query.filter(
            Cita.asistida == False,
            Cita.fecha_cita.in_([hoy_str, manana])
        ).limit(5).all()
        
        # Pacientes recientes
        pacientes_recientes = Paciente.query.order_by(Paciente.id.desc()).limit(5).all()
        
        # Productividad general
        productividad_general = int((citas_completadas / total_citas * 100)) if total_citas > 0 else 0
        
        content = f"""
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 25px;">
                <h2 style="margin:0;">📊 Panel de Control - Dirección Médica</h2>
                <div>
                    <a href="/admin/exportar-todo" class="btn btn-success" style="margin-right:10px;">📊 Exportar Todo</a>
                    <span style="color:#666;">{hoy.strftime('%d de %B, %Y')}</span>
                </div>
            </div>
            
            <!-- KPIs -->
            <div style="display: grid; grid-template-columns: repeat(5, 1fr); gap: 15px; margin-bottom: 25px;">
                <div style="background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">👨‍⚕️ Médicos</div>
                    <div style="font-size: 32px; font-weight: bold;">{total_medicos}</div>
                </div>
                <div style="background: linear-gradient(135deg, #f093fb, #f5576c); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">🏥 Pacientes</div>
                    <div style="font-size: 32px; font-weight: bold;">{total_pacientes}</div>
                </div>
                <div style="background: linear-gradient(135deg, #4facfe, #00f2fe); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">📅 Citas Pendientes</div>
                    <div style="font-size: 32px; font-weight: bold;">{citas_pendientes}</div>
                </div>
                <div style="background: linear-gradient(135deg, #43e97b, #38f9d7); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">✅ Citas Realizadas</div>
                    <div style="font-size: 32px; font-weight: bold;">{citas_completadas}</div>
                </div>
                <div style="background: linear-gradient(135deg, #fa709a, #fee140); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">📈 Productividad</div>
                    <div style="font-size: 32px; font-weight: bold;">{productividad_general}%</div>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                <!-- Citas por Especialidad -->
                <div class="card">
                    <h3>📋 Citas Pendientes por Especialidad</h3>
                    <div style="max-height: 250px; overflow-y: auto;">
                        <table style="width:100%; min-width:0;">
                            <thead><tr><th>Especialidad</th><th>Pendientes</th><th>Barra</th></tr></thead>
                            <tbody>
                                {''.join([f'''
                                <tr>
                                    <td>{d[0]}</td>
                                    <td><strong>{d[1]}</strong></td>
                                    <td>
                                        <div style="background:#e0e0e0; height:8px; border-radius:4px; width:100px;">
                                            <div style="background:#1a5276; height:8px; border-radius:4px; width:{min(d[1]*10, 100)}px;"></div>
                                        </div>
                                    </td>
                                </tr>
                                ''' for d in especialidades]) if especialidades else '<tr><td colspan="3">No hay datos</td></tr>'}
                            </tbody>
                        </table>
                    </div>
                </div>
                
                <!-- Top Médicos -->
                <div class="card">
                    <h3>🏆 Top Productividad Médica</h3>
                    <div style="max-height: 250px; overflow-y: auto;">
                        {''.join([f'''
                        <div style="display: flex; align-items: center; justify-content: space-between; padding: 10px 0; border-bottom: 1px solid #eee;">
                            <div style="display: flex; align-items: center; gap: 10px;">
                                <div style="width: 40px; height: 40px; background: {"#ffd700" if i==0 else "#c0c0c0" if i==1 else "#cd7f32" if i==2 else "#1a5276"}; color: white; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: bold;">
                                    👨‍⚕️
                                </div>
                                <div>
                                    <strong>Dr/a. {nombre}</strong>
                                    <div style="font-size: 12px; color: #666;">{depto}</div>
                                </div>
                            </div>
                            <div style="text-align: right;">
                                <div style="font-weight: bold; color: #1a5276;">{prod}%</div>
                                <div style="font-size: 12px; color: #666;">{comp}/{total} citas</div>
                            </div>
                        </div>
                        ''' for i, (nombre, prod, comp, total, depto) in enumerate(top5)]) if top5 else '<p>No hay médicos registrados</p>'}
                    </div>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                <!-- Citas Urgentes -->
                <div class="card">
                    <h3 style="color: #e74c3c;">⚡ Citas Urgentes (hoy/mañana)</h3>
                    {''.join([f'''
                    <div style="padding: 12px; background: #fff5f5; border-radius: 8px; margin-bottom: 8px; border-left: 4px solid #e74c3c;">
                        <div style="display: flex; justify-content: space-between;">
                            <strong>{c.titulo[:40]}...</strong>
                            <span style="color:#e74c3c; font-size:12px;">{c.fecha_cita} {c.hora_inicio or ''}</span>
                        </div>
                        <div style="font-size: 13px; color: #666;">👨‍⚕️ Dr/a. {c.medico.nombre_completo if c.medico else 'Sin asignar'}</div>
                    </div>
                    ''' for c in citas_urgentes]) if citas_urgentes else '<p style="color:#27ae60;">✅ No hay citas urgentes</p>'}
                </div>
                
                <!-- Pacientes Recientes -->
                <div class="card">
                    <h3>🆕 Pacientes Recientes</h3>
                    {''.join([f'''
                    <div style="padding: 12px; border-bottom: 1px solid #eee;">
                        <div style="display: flex; align-items: center; gap: 10px;">
                            <div style="width: 35px; height: 35px; background: #1a5276; color: white; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 14px;">
                                {p.nombre[0].upper()}
                            </div>
                            <div style="flex:1;">
                                <strong>{p.nombre}</strong>
                                <div style="font-size: 12px; color: #666;">{p.obra_social or 'Particular'} • {p.fecha_creacion}</div>
                            </div>
                            <span style="font-size: 12px; color: #999;">👨‍⚕️ {p.usuario.nombre_completo.split()[0]}</span>
                        </div>
                    </div>
                    ''' for p in pacientes_recientes]) if pacientes_recientes else '<p>No hay pacientes aún</p>'}
                    <a href="/pacientes" style="display: block; text-align: center; margin-top: 15px; color: #1a5276;">Ver todos los pacientes →</a>
                </div>
            </div>
        """
    else:
        # Dashboard de MÉDICO
        citas = Cita.query.filter_by(usuario_id=user.id).all()
        pendientes = sum(1 for c in citas if not c.asistida)
        completadas = sum(1 for c in citas if c.asistida)
        productividad = int((completadas / len(citas) * 100)) if citas else 0
        
        # Citas urgentes (hoy/mañana)
        hoy_str = hoy.strftime('%Y-%m-%d')
        manana = (hoy + timedelta(days=1)).strftime('%Y-%m-%d')
        citas_urgentes = [c for c in citas if not c.asistida and c.fecha_cita in [hoy_str, manana]]
        
        # Próximas citas
        citas_proximas = sorted([c for c in citas if not c.asistida], key=lambda x: x.fecha_cita or '9999-12-31')[:5]
        
        # Mis pacientes
        mis_pacientes = Paciente.query.filter_by(usuario_id=user.id).count()
        pacientes_recientes = Paciente.query.filter_by(usuario_id=user.id).order_by(Paciente.id.desc()).limit(3).all()
        
        # Mensajes no leídos
        mensajes_no_leidos = Mensaje.query.filter_by(receptor_id=user.id, leido=False).count()
        
        # Citas por prioridad
        citas_alta = sum(1 for c in citas if not c.asistida and c.prioridad == 'alta')
        citas_media = sum(1 for c in citas if not c.asistida and c.prioridad == 'media')
        citas_baja = sum(1 for c in citas if not c.asistida and c.prioridad == 'baja')
        
        # Frase motivacional
        if productividad >= 80:
            frase = "🌟 ¡Excelente trabajo, doctor/a!"
            emoji = "🏆"
        elif productividad >= 50:
            frase = "💪 ¡Buen ritmo de consultas!"
            emoji = "🚀"
        else:
            frase = "📋 ¡Cada paciente atendido es un paso adelante!"
            emoji = "🎯"
        
        content = f"""
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 25px;">
                <div>
                    <h2 style="margin:0;">👋 ¡Hola, Dr/a. {user.nombre_completo}!</h2>
                    <p style="color:#666; margin:5px 0 0 0;">{user.departamento} • {user.cargo}</p>
                </div>
                <div style="text-align: right;">
                    <div style="font-size: 32px;">{emoji}</div>
                    <p style="color:#666; font-style:italic; max-width:300px;">{frase}</p>
                </div>
            </div>
            
            <!-- KPIs Personales -->
            <div style="display: grid; grid-template-columns: repeat(5, 1fr); gap: 15px; margin-bottom: 25px;">
                <div style="background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">📅 Mis Citas</div>
                    <div style="font-size: 32px; font-weight: bold;">{len(citas)}</div>
                    <div style="font-size: 12px; opacity:0.8;">{pendientes} pendientes</div>
                </div>
                <div style="background: linear-gradient(135deg, #f093fb, #f5576c); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">✅ Realizadas</div>
                    <div style="font-size: 32px; font-weight: bold;">{completadas}</div>
                </div>
                <div style="background: linear-gradient(135deg, #4facfe, #00f2fe); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">📈 Productividad</div>
                    <div style="font-size: 32px; font-weight: bold;">{productividad}%</div>
                    <div style="margin-top:5px; background:rgba(255,255,255,0.3); height:6px; border-radius:3px;">
                        <div style="background:white; width:{productividad}%; height:6px; border-radius:3px;"></div>
                    </div>
                </div>
                <div style="background: linear-gradient(135deg, #43e97b, #38f9d7); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">🏥 Mis Pacientes</div>
                    <div style="font-size: 32px; font-weight: bold;">{mis_pacientes}</div>
                </div>
                <div style="background: linear-gradient(135deg, #fa709a, #fee140); color: white; padding: 20px; border-radius: 12px;">
                    <div style="font-size: 14px; opacity: 0.9;">💬 Mensajes</div>
                    <div style="font-size: 32px; font-weight: bold;">{mensajes_no_leidos}</div>
                    <div style="font-size: 12px; opacity:0.8;">sin leer</div>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px;">
                <!-- Citas Urgentes -->
                <div class="card">
                    <h3 style="color: #e74c3c;">⚡ Citas Urgentes (hoy/mañana)</h3>
                    {''.join([f'''
                    <div style="padding: 12px; background: #fff5f5; border-radius: 8px; margin-bottom: 8px; border-left: 4px solid #e74c3c;">
                        <div style="display: flex; justify-content: space-between;">
                            <strong>{c.titulo[:35]}...</strong>
                            <span style="color:#e74c3c; font-size:12px;">{c.fecha_cita} {c.hora_inicio or ''}</span>
                        </div>
                        <div style="margin-top:8px;">
                            <a href="/cita/completar/{c.id}" class="btn btn-success btn-sm">✅ Marcar realizada</a>
                            <button class="btn btn-primary btn-sm" onclick="verCita({c.id})">👁️ Ver</button>
                        </div>
                    </div>
                    ''' for c in citas_urgentes]) if citas_urgentes else '<p style="color:#27ae60; padding:12px; background:#f0fff0; border-radius:8px;">✅ No tienes citas urgentes</p>'}
                </div>
                
                <!-- Resumen por Prioridad -->
                <div class="card">
                    <h3>📊 Citas Pendientes por Prioridad</h3>
                    <div style="margin-bottom: 20px;">
                        <div style="display: flex; justify-content: space-between; margin-bottom: 8px;">
                            <span>🔴 Alta</span>
                            <span><strong>{citas_alta}</strong> citas</span>
                        </div>
                        <div style="background:#e0e0e0; height:8px; border-radius:4px;">
                            <div style="background:#e74c3c; width:{min(citas_alta*20, 100)}%; height:8px; border-radius:4px;"></div>
                        </div>
                    </div>
                    <div style="margin-bottom: 20px;">
                        <div style="display: flex; justify-content: space-between; margin-bottom: 8px;">
                            <span>🟡 Media</span>
                            <span><strong>{citas_media}</strong> citas</span>
                        </div>
                        <div style="background:#e0e0e0; height:8px; border-radius:4px;">
                            <div style="background:#f39c12; width:{min(citas_media*20, 100)}%; height:8px; border-radius:4px;"></div>
                        </div>
                    </div>
                    <div>
                        <div style="display: flex; justify-content: space-between; margin-bottom: 8px;">
                            <span>🟢 Baja</span>
                            <span><strong>{citas_baja}</strong> citas</span>
                        </div>
                        <div style="background:#e0e0e0; height:8px; border-radius:4px;">
                            <div style="background:#27ae60; width:{min(citas_baja*20, 100)}%; height:8px; border-radius:4px;"></div>
                        </div>
                    </div>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 2fr 1fr; gap: 20px; margin-bottom: 20px;">
                <!-- Próximas Citas -->
                <div class="card">
                    <h3>📌 Próximas Citas</h3>
                    <div style="max-height: 250px; overflow-y: auto;">
                        {''.join([f'''
                        <div style="padding: 12px; border-bottom: 1px solid #eee; display: flex; align-items: center; gap: 10px;">
                            <div style="width: 8px; height: 8px; background: {"#e74c3c" if c.prioridad=="alta" else "#f39c12" if c.prioridad=="media" else "#27ae60"}; border-radius: 50%;"></div>
                            <div style="flex:1;">
                                <strong>{c.titulo[:40]}...</strong>
                                <div style="font-size: 12px; color: #666;">
                                    📅 {c.fecha_cita or 'Sin fecha'} 
                                    {f"🕐 {c.hora_inicio}" if c.hora_inicio else ""}
                                    {f" • 🏥 {c.paciente.nombre}" if c.paciente else ""}
                                </div>
                            </div>
                            <button class="btn btn-primary btn-sm" onclick="verCita({c.id})">👁️</button>
                        </div>
                        ''' for c in citas_proximas]) if citas_proximas else '<p style="color:#999; text-align:center; padding:20px;">No hay citas pendientes</p>'}
                    </div>
                    <a href="/citas" style="display: block; text-align: center; margin-top: 15px; color: #1a5276;">Ver todas mis citas →</a>
                </div>
                
                <!-- Accesos Rápidos -->
                <div class="card">
                    <h3>⚡ Accesos Rápidos</h3>
                    <div style="display: flex; flex-direction: column; gap: 10px;">
                        <a href="/citas" style="padding: 15px; background: #f8f9fa; border-radius: 8px; text-decoration: none; color: #333; display: flex; align-items: center; gap: 10px;">
                            <span style="font-size: 24px;">📅</span>
                            <div>
                                <strong>Gestionar Citas</strong>
                                <div style="font-size: 12px; color: #666;">{pendientes} citas pendientes</div>
                            </div>
                        </a>
                        <a href="/pacientes" style="padding: 15px; background: #f8f9fa; border-radius: 8px; text-decoration: none; color: #333; display: flex; align-items: center; gap: 10px;">
                            <span style="font-size: 24px;">🏥</span>
                            <div>
                                <strong>Mis Pacientes</strong>
                                <div style="font-size: 12px; color: #666;">{mis_pacientes} pacientes asignados</div>
                            </div>
                        </a>
                        <a href="/chat" style="padding: 15px; background: #f8f9fa; border-radius: 8px; text-decoration: none; color: #333; display: flex; align-items: center; gap: 10px;">
                            <span style="font-size: 24px;">💬</span>
                            <div>
                                <strong>Chat con Dirección</strong>
                                {f'<span style="background:#e74c3c; color:white; padding:2px 8px; border-radius:10px; font-size:11px; margin-left:10px;">{mensajes_no_leidos} nuevo</span>' if mensajes_no_leidos > 0 else ''}
                            </div>
                        </a>
                        <a href="/chat-grupal" style="padding: 15px; background: #f8f9fa; border-radius: 8px; text-decoration: none; color: #333; display: flex; align-items: center; gap: 10px;">
                            <span style="font-size: 24px;">👥</span>
                            <div>
                                <strong>Chat {user.departamento}</strong>
                                <div style="font-size: 12px; color: #666;">Equipo médico</div>
                            </div>
                        </a>
                    </div>
                </div>
            </div>
            
            <!-- Modal para ver cita -->
            <div id="modalCita" style="display:none; position:fixed; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.5); z-index:1000; justify-content:center; align-items:center;">
                <div style="background:white; padding:30px; border-radius:12px; max-width:500px; width:90%; max-height:80vh; overflow-y:auto;">
                    <h3 id="modalTitulo"></h3>
                    <p><strong>📅 Fecha:</strong> <span id="modalFecha"></span> <span id="modalHora"></span></p>
                    <p><strong>🩺 Tipo:</strong> <span id="modalTipo"></span></p>
                    <p><strong>🔴 Prioridad:</strong> <span id="modalPrioridad"></span></p>
                    <p><strong>🏥 Paciente:</strong> <span id="modalPaciente"></span></p>
                    <p><strong>📝 Descripción:</strong></p>
                    <div id="modalDescripcion" style="background:#f5f5f5; padding:15px; border-radius:8px; margin:10px 0;"></div>
                    <div style="display: flex; gap: 10px;">
                        <button class="btn btn-primary" onclick="document.getElementById('modalCita').style.display='none'">Cerrar</button>
                        <a href="#" id="modalCompletar" class="btn btn-success">✅ Marcar realizada</a>
                    </div>
                </div>
            </div>
            
            <script>
            const citasData = """ + json.dumps([{
                'id': c.id,
                'titulo': c.titulo,
                'descripcion': c.descripcion or 'Sin descripción',
                'fecha_cita': c.fecha_cita or 'Sin fecha',
                'hora_inicio': c.hora_inicio or '',
                'tipo_consulta': c.tipo_consulta or 'Presencial',
                'prioridad': c.prioridad,
                'asistida': c.asistida,
                'paciente': c.paciente.nombre if c.paciente else 'Sin paciente'
            } for c in citas]) + """;
            
            function verCita(id) {
                const cita = citasData.find(c => c.id === id);
                if (cita) {
                    document.getElementById('modalTitulo').textContent = cita.titulo;
                    document.getElementById('modalFecha').textContent = cita.fecha_cita;
                    document.getElementById('modalHora').textContent = cita.hora_inicio ? '🕐 ' + cita.hora_inicio : '';
                    document.getElementById('modalTipo').textContent = cita.tipo_consulta;
                    document.getElementById('modalPrioridad').textContent = cita.prioridad.toUpperCase();
                    document.getElementById('modalPaciente').textContent = cita.paciente;
                    document.getElementById('modalDescripcion').textContent = cita.descripcion;
                    document.getElementById('modalCompletar').href = '/cita/completar/' + cita.id;
                    if (cita.asistida) {
                        document.getElementById('modalCompletar').style.display = 'none';
                    } else {
                        document.getElementById('modalCompletar').style.display = 'inline-block';
                    }
                    document.getElementById('modalCita').style.display = 'flex';
                }
            }
            
            document.getElementById('modalCita').addEventListener('click', function(e) {
                if (e.target === this) this.style.display = 'none';
            });
            </script>
        """
    
    return base_html(content, "Dashboard - Clínica Médica")

    # ========== MI ESPACIO PERSONAL ==========

@app.route('/mi-espacio')
@login_required
def mi_espacio():
    user = db.session.get(Usuario, session.get('user_id'))
    notas = NotaPersonal.query.filter_by(usuario_id=user.id).order_by(NotaPersonal.id.desc()).all()
    tareas_personales = TareaPersonal.query.filter_by(usuario_id=user.id).order_by(TareaPersonal.id.desc()).all()
    
    # ========== CALENDARIO (CÓDIGO NUEVO) ==========
    hoy = datetime.now().date()
    lunes = hoy - timedelta(days=hoy.weekday())
    semana = [lunes + timedelta(days=i) for i in range(7)]
    
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    meses = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    
    tareas_semana = []
    for t in tareas_personales:
        if t.fecha_limite and not t.completada:
            try:
                fecha_t = datetime.strptime(t.fecha_limite, '%Y-%m-%d').date()
                if lunes <= fecha_t <= lunes + timedelta(days=6):
                    tareas_semana.append(t)
            except:
                pass
    
    dias_html = ""
    for i, dia in enumerate(semana):
        tareas_dia = [t for t in tareas_semana if datetime.strptime(t.fecha_limite, '%Y-%m-%d').date() == dia]
        es_hoy = dia == hoy
        
        tareas_dia_html = ""
        for t in tareas_dia[:4]:
            prioridad_class = t.prioridad
            prioridad_icono = {'alta': '🔴', 'media': '🟡', 'baja': '🟢'}.get(t.prioridad, '📌')
            
            tareas_dia_html += f"""
            <div class="calendario-tarea {prioridad_class}" onclick="window.location.href='/tarea-personal/completar/{t.id}'">
                <div style="display: flex; align-items: center; gap: 5px;">
                    <span>{prioridad_icono}</span>
                    <span style="flex: 1; font-weight: 500;">{t.titulo[:20]}{'...' if len(t.titulo) > 20 else ''}</span>
                </div>
                <small style="display: block; margin-top: 3px; opacity: 0.7; font-size: 10px;">
                    ⏰ {t.fecha_limite}
                </small>
            </div>
            """
        
        tareas_extra = len(tareas_dia) - 4
        if tareas_extra > 0:
            tareas_dia_html += f"""
            <div class="calendario-tarea" style="background: #e0e7ef; text-align: center; font-weight: 600; border-left-color: #8898aa;">
                +{tareas_extra} más
            </div>
            """
        
        hoy_class = 'hoy' if es_hoy else ''
        nombre_dia = dias_semana[dia.weekday()]
        dia_numero = dia.day
        mes_actual = meses[dia.month - 1]
        
        dias_html += f"""
        <div class="calendario-dia {hoy_class}">
            <div class="calendario-dia-header">
                <strong>{nombre_dia}</strong>
                <span>{dia_numero}</span>
                <small style="display: block; color: #8898aa; font-size: 10px; margin-top: 3px;">{mes_actual[:3]}</small>
            </div>
            <div class="calendario-tareas">
                {tareas_dia_html if tareas_dia else '<div style="padding: 15px 5px; text-align: center; color: #bdc3c7; font-size: 12px;"><span style="font-size: 20px; display: block; margin-bottom: 5px;">✨</span>Libre</div>'}
            </div>
        </div>
        """
    
    calendario_html = f"""
    <div style="margin-bottom: 25px;">
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
            <h3 style="margin: 0; display: flex; align-items: center; gap: 10px;">
                <span style="font-size: 28px;">📅</span> 
                Semana del {lunes.strftime('%d')} al {(lunes + timedelta(days=6)).strftime('%d de %B, %Y')}
            </h3>
            <div style="display: flex; gap: 8px;">
                <span style="display: flex; align-items: center; gap: 5px;"><span style="width: 12px; height: 12px; background: #e74c3c; border-radius: 4px;"></span> Alta</span>
                <span style="display: flex; align-items: center; gap: 5px;"><span style="width: 12px; height: 12px; background: #f39c12; border-radius: 4px;"></span> Media</span>
                <span style="display: flex; align-items: center; gap: 5px;"><span style="width: 12px; height: 12px; background: #27ae60; border-radius: 4px;"></span> Baja</span>
            </div>
        </div>
        <div class="calendario-semana">
            {dias_html}
        </div>
        <div style="display: flex; justify-content: center; margin-top: 20px;">
            <a href="/tarea-personal/nueva" class="btn btn-success" style="padding: 14px 32px; font-size: 16px;">
                ➕ Añadir Nueva Tarea
            </a>
        </div>
    </div>
    """
    # ========== FIN CALENDARIO ==========
    
  
    # Notas HTML
    notas_html = ""
    colores = ['#f39c12', '#e74c3c', '#3498db', '#2ecc71', '#9b59b6', '#1abc9c']
    for nota in notas:
        notas_html += f"""
        <div style="background: {nota.color}15; border-left: 5px solid {nota.color}; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); min-height: 150px; position: relative;">
            <h4 style="margin: 0 0 10px 0;">{nota.titulo}</h4>
            <p style="margin: 0; color: #555;">{nota.contenido[:100]}{'...' if len(nota.contenido) > 100 else ''}</p>
            <small style="color: #999; display: block; margin-top: 10px;">{nota.fecha_creacion}</small>
            <div style="position: absolute; bottom: 10px; right: 10px; display: flex; gap: 5px;">
                <a href="/nota/editar/{nota.id}" class="btn btn-warning btn-sm">✏️</a>
                <a href="/nota/eliminar/{nota.id}" class="btn btn-danger btn-sm" onclick="return confirm('¿Eliminar nota?')">🗑️</a>
            </div>
        </div>
        """
    
    # Tareas personales HTML
    tareas_html = ""
    for t in tareas_personales:
        prioridad_badge = {'alta': '🔴', 'media': '🟡', 'baja': '🟢'}.get(t.prioridad, '')
        tareas_html += f"""
        <tr>
            <td>{prioridad_badge} {t.titulo[:40]}...</td>
            <td>{t.fecha_limite or '-'}</td>
            <td>{'✅' if t.completada else '⏳'}</td>
            <td>
                {f'<a href="/tarea-personal/completar/{t.id}" class="btn btn-success btn-sm">✅</a>' if not t.completada else ''}
                <a href="/tarea-personal/eliminar/{t.id}" class="btn btn-danger btn-sm" onclick="return confirm(\'¿Eliminar?\')">🗑️</a>
            </td>
        </tr>
        """
    
    content = f"""
        <h2>📌 Mi Espacio Personal - Dr/a. {user.nombre_completo}</h2>
        
        <div class="tabs">
            <button class="tab active" onclick="showTab('notas')">📓 Mis Notas</button>
            <button class="tab" onclick="showTab('tareas')">✅ Mis Tareas</button>
            <button class="tab" onclick="showTab('calendario')">📅 Mi Calendario</button>
        </div>
        
        <!-- NOTAS -->
        <div id="tab-notas" class="tab-content active">
            <p style="margin-bottom: 20px;">
                <a href="/nota/nueva" class="btn btn-primary">➕ Nueva Nota</a>
            </p>
            <div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(250px, 1fr)); gap: 15px;">
                {notas_html if notas else '<p style="grid-column: 1/-1; text-align: center; color: #999;">📓 No hay notas. ¡Crea una!</p>'}
            </div>
        </div>
        
        <!-- TAREAS PERSONALES -->
        <div id="tab-tareas" class="tab-content">
            <p style="margin-bottom: 20px;">
                <a href="/tarea-personal/nueva" class="btn btn-primary">➕ Nueva Tarea</a>
            </p>
            <table>
                <thead>
                    <tr>
                        <th>Título</th>
                        <th>Fecha Límite</th>
                        <th>Estado</th>
                        <th>Acciones</th>
                    </tr>
                </thead>
                <tbody>
                    {tareas_html if tareas_personales else '<tr><td colspan="4" style="text-align:center;">No hay tareas personales</td></tr>'}
                </tbody>
            </table>
        </div>
        
        <!-- CALENDARIO PERSONAL -->
        <div id="tab-calendario" class="tab-content">
            <div class="card">
                <p style="margin-bottom: 15px;">
                    📅 Semana del <strong>{lunes.strftime('%d/%m/%Y')}</strong> al <strong>{(lunes + timedelta(days=6)).strftime('%d/%m/%Y')}</strong>
                </p>
                <table style="width: 100%; border-collapse: collapse;">
                    <tr>{dias_html}</tr>
                </table>
                <p style="margin-top: 20px;">
                    <a href="/tarea-personal/nueva" class="btn btn-primary">➕ Añadir tarea al calendario</a>
                </p>
            </div>
        </div>
        
        <script>
            function showTab(tabName) {{
                document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
                document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
                event.target.classList.add('active');
                document.getElementById('tab-' + tabName).classList.add('active');
            }}
        </script>
    """
    return base_html(content, "Mi Espacio Personal")

# ========== GESTIÓN DE PACIENTES ==========

@app.route('/pacientes')
@login_required
def pacientes():
    user = db.session.get(Usuario, session.get('user_id'))
    
    # Admin ve todos los pacientes, médicos solo los suyos
    if session.get('rol') == 'admin':
        pacientes_list = Paciente.query.order_by(Paciente.nombre).all()
    else:
        pacientes_list = Paciente.query.filter_by(usuario_id=user.id).order_by(Paciente.nombre).all()
    
    busqueda = request.args.get('busqueda', '')
    if busqueda:
        pacientes_list = [p for p in pacientes_list if 
                        busqueda.lower() in p.nombre.lower() or 
                        (p.obra_social and busqueda.lower() in p.obra_social.lower()) or
                        (p.email and busqueda.lower() in p.email.lower())]
    
    # Estadísticas rápidas
    total_pacientes = len(pacientes_list)
    pacientes_con_email = sum(1 for p in pacientes_list if p.email)
    pacientes_con_telefono = sum(1 for p in pacientes_list if p.telefono)
    
    # Agrupar por obra social
    obras_sociales = {}
    for p in pacientes_list:
        os = p.obra_social or 'Particular'
        if os not in obras_sociales:
            obras_sociales[os] = []
        obras_sociales[os].append(p)
    
    pacientes_html = ""
    for p in pacientes_list[:20]:
        inicial = p.nombre[0].upper() if p.nombre else '?'
        interacciones_count = len(p.interacciones)
        ultima_interaccion = p.interacciones[-1].fecha if p.interacciones else 'Sin consultas'
        
        # Color según actividad
        if interacciones_count > 5:
            border_color = '#27ae60'
        elif interacciones_count > 2:
            border_color = '#f39c12'
        else:
            border_color = '#e74c3c'
        
        # Grupo sanguíneo badge
        grupo_badge = f'<span style="background:#e74c3c; color:white; padding:2px 8px; border-radius:10px; font-size:11px; margin-left:5px;">{p.grupo_sanguineo}</span>' if p.grupo_sanguineo else ''
        
        pacientes_html += f"""
        <div class="paciente-card" style="border-top: 4px solid {border_color};">
            <div class="paciente-header">
                <div class="paciente-avatar" style="background: {border_color};">{inicial}</div>
                <div class="paciente-info">
                    <h3>{p.nombre} {grupo_badge}</h3>
                    <p>{p.obra_social or 'Particular'}</p>
                </div>
            </div>
            <div class="paciente-body">
                <div class="paciente-contacto">
                    {f'<span>📧 {p.email}</span>' if p.email else ''}
                    {f'<span>📱 {p.telefono}</span>' if p.telefono else ''}
                    {f'<span>🎂 {p.fecha_nacimiento}</span>' if p.fecha_nacimiento else ''}
                    {f'<span>📍 {p.direccion[:30]}...</span>' if p.direccion else ''}
                </div>
                {f'<p style="background:#fff3e0; padding:8px; border-radius:6px; margin-top:10px; font-size:13px;"><strong>⚠️ Alergias:</strong> {p.alergias[:50]}...</p>' if p.alergias else ''}
                <div style="display: flex; justify-content: space-between; margin-top: 10px; font-size: 12px; color: #666;">
                    <span>🩺 {interacciones_count} consultas</span>
                    <span>🕐 {ultima_interaccion}</span>
                </div>
                {f'<p style="background:#f5f5f5; padding:8px; border-radius:6px; margin-top:10px; font-size:13px;">📝 {p.notas[:60]}...</p>' if p.notas else ''}
            </div>
            <div class="paciente-footer">
                <small>👨‍⚕️ Dr/a. {p.usuario.nombre_completo}</small>
                <div>
                    <a href="/paciente/{p.id}" class="btn btn-primary btn-sm">👁️ Ver</a>
                    <a href="/paciente/editar/{p.id}" class="btn btn-warning btn-sm">✏️</a>
                    <a href="/paciente/eliminar/{p.id}" class="btn btn-danger btn-sm" onclick="return confirm('¿Eliminar paciente?')">🗑️</a>
                </div>
            </div>
        </div>
        """
    
    content = f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 25px;">
            <h2 style="margin:0;">🏥 Gestión de Pacientes</h2>
            <a href="/paciente/nuevo" class="btn btn-success">➕ Nuevo Paciente</a>
        </div>
        
        <!-- KPIs de pacientes -->
        <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 25px;">
            <div style="background: linear-gradient(135deg, #667eea, #764ba2); color: white; padding: 15px; border-radius: 12px;">
                <div style="font-size: 14px; opacity: 0.9;">👥 Total Pacientes</div>
                <div style="font-size: 28px; font-weight: bold;">{total_pacientes}</div>
            </div>
            <div style="background: linear-gradient(135deg, #f093fb, #f5576c); color: white; padding: 15px; border-radius: 12px;">
                <div style="font-size: 14px; opacity: 0.9;">📧 Con Email</div>
                <div style="font-size: 28px; font-weight: bold;">{pacientes_con_email}</div>
            </div>
            <div style="background: linear-gradient(135deg, #4facfe, #00f2fe); color: white; padding: 15px; border-radius: 12px;">
                <div style="font-size: 14px; opacity: 0.9;">📱 Con Teléfono</div>
                <div style="font-size: 28px; font-weight: bold;">{pacientes_con_telefono}</div>
            </div>
            <div style="background: linear-gradient(135deg, #43e97b, #38f9d7); color: white; padding: 15px; border-radius: 12px;">
                <div style="font-size: 14px; opacity: 0.9;">🏥 Obras Sociales</div>
                <div style="font-size: 28px; font-weight: bold;">{len(obras_sociales)}</div>
            </div>
        </div>
        
        <!-- Buscador -->
        <div style="display: flex; gap: 10px; margin-bottom: 20px;">
            <form method="GET" style="flex: 1; display: flex; gap: 10px;">
                <input type="text" name="busqueda" placeholder="🔍 Buscar por nombre, email u obra social..." value="{busqueda}" 
                       style="flex:1; padding:12px; border:1px solid #ddd; border-radius:8px; font-size:14px;">
                <button type="submit" class="btn btn-primary">Buscar</button>
                {f'<a href="/pacientes" class="btn" style="background:#95a5a6;color:white;">Limpiar</a>' if busqueda else ''}
            </form>
        </div>
        
        <!-- Lista de pacientes -->
        <div class="pacientes-grid">
            {pacientes_html if pacientes_list else '<div style="grid-column:1/-1; text-align:center; padding:40px; color:#999;"><p style="font-size:48px;">🏥</p><p>No hay pacientes registrados. ¡Añade tu primer paciente!</p></div>'}
        </div>
        
        {f'<p style="text-align:center; margin-top:20px; color:#666;">Mostrando {min(20, len(pacientes_list))} de {len(pacientes_list)} pacientes</p>' if len(pacientes_list) > 20 else ''}
    """
    return base_html(content, "Pacientes")

@app.route('/paciente/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_paciente():
    user = db.session.get(Usuario, session.get('user_id'))
    
    if request.method == 'POST':
        paciente = Paciente(
            nombre=request.form['nombre'],
            email=request.form.get('email', ''),
            telefono=request.form.get('telefono', ''),
            obra_social=request.form.get('obra_social', ''),
            direccion=request.form.get('direccion', ''),
            fecha_nacimiento=request.form.get('fecha_nacimiento', ''),
            grupo_sanguineo=request.form.get('grupo_sanguineo', ''),
            alergias=request.form.get('alergias', ''),
            notas=request.form.get('notas', ''),
            usuario_id=user.id
        )
        db.session.add(paciente)
        db.session.commit()
        flash('✅ Paciente registrado correctamente')
        return redirect(f'/paciente/{paciente.id}')
    
    content = """
        <h2>🏥 Nuevo Paciente</h2>
        <form method="POST" style="max-width: 700px;">
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Nombre completo *</label>
                    <input type="text" name="nombre" class="form-control" required>
                </div>
                <div class="form-group">
                    <label>Fecha de nacimiento</label>
                    <input type="date" name="fecha_nacimiento" class="form-control">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Email</label>
                    <input type="email" name="email" class="form-control">
                </div>
                <div class="form-group">
                    <label>Teléfono</label>
                    <input type="tel" name="telefono" class="form-control">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Obra Social</label>
                    <input type="text" name="obra_social" class="form-control" placeholder="Ej: OSDE, Swiss Medical, Particular...">
                </div>
                <div class="form-group">
                    <label>Grupo Sanguíneo</label>
                    <select name="grupo_sanguineo" class="form-control">
                        <option value="">Desconocido</option>
                        <option value="A+">A+</option>
                        <option value="A-">A-</option>
                        <option value="B+">B+</option>
                        <option value="B-">B-</option>
                        <option value="AB+">AB+</option>
                        <option value="AB-">AB-</option>
                        <option value="O+">O+</option>
                        <option value="O-">O-</option>
                    </select>
                </div>
            </div>
            
            <div class="form-group">
                <label>Dirección</label>
                <input type="text" name="direccion" class="form-control">
            </div>
            
            <div class="form-group">
                <label>⚠️ Alergias</label>
                <textarea name="alergias" class="form-control" rows="2" placeholder="Ej: Penicilina, Látex, Polen..."></textarea>
            </div>
            
            <div class="form-group">
                <label>Notas adicionales</label>
                <textarea name="notas" class="form-control" rows="3" placeholder="Antecedentes relevantes, observaciones..."></textarea>
            </div>
            
            <button type="submit" class="btn btn-primary">💾 Guardar Paciente</button>
            <a href="/pacientes" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Nuevo Paciente")

@app.route('/paciente/<int:id>')
@login_required
def ver_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    # Verificar permisos
    if session.get('rol') != 'admin' and paciente.usuario_id != user.id:
        flash('❌ No tienes permiso para ver este paciente')
        return redirect('/pacientes')
    
    interacciones = Interaccion.query.filter_by(paciente_id=paciente.id).order_by(Interaccion.id.desc()).all()
    
    interacciones_html = ""
    for i in interacciones:
        icono = {'Llamada': '📞', 'Email': '📧', 'Consulta': '🩺', 'Nota': '📝'}.get(i.tipo, '📌')
        interacciones_html += f"""
        <div style="display: flex; gap: 15px; padding: 15px; background: #f8f9fa; border-radius: 8px; margin-bottom: 10px;">
            <div style="width: 40px; height: 40px; background: #1a5276; color: white; border-radius: 50%; display: flex; align-items: center; justify-content: center;">
                {icono}
            </div>
            <div style="flex: 1;">
                <div style="display: flex; justify-content: space-between; margin-bottom: 5px;">
                    <span style="font-weight: bold; color: #1a5276;">{i.tipo}</span>
                    <span style="color: #999; font-size: 12px;">{i.fecha}</span>
                </div>
                <p>{i.descripcion}</p>
                <small>👨‍⚕️ Dr/a. {i.usuario.nombre_completo}</small>
            </div>
        </div>
        """
    
    # Badge de grupo sanguíneo
    grupo_badge = f'<span style="background:#e74c3c; color:white; padding:5px 15px; border-radius:20px; font-size:18px; font-weight:bold;">{paciente.grupo_sanguineo}</span>' if paciente.grupo_sanguineo else '<span style="color:#999;">No registrado</span>'
    
    # Historias clínicas del paciente
    historias = HistoriaClinica.query.filter_by(paciente_id=paciente.id).order_by(HistoriaClinica.id.desc()).limit(5).all()
    historias_html = ""
    for h in historias:
        historias_html += f"""
        <div style="padding: 12px; border-bottom: 1px solid #eee;">
            <div style="display: flex; justify-content: space-between;">
                <strong>{h.titulo[:40]}...</strong>
                <span style="color: #666; font-size: 12px;">{h.fecha_creacion}</span>
            </div>
            <p style="font-size: 13px; color: #666; margin-top: 5px;">Diagnóstico: {h.diagnostico[:50] if h.diagnostico else 'Pendiente'}...</p>
            <a href="/historia-clinica/{h.id}" class="btn btn-primary btn-sm" style="margin-top: 5px;">👁️ Ver historia</a>
        </div>
        """
    
    # Recetas del paciente
    recetas = Receta.query.filter_by(paciente_id=paciente.id).order_by(Receta.id.desc()).limit(5).all()
    recetas_html = ""
    for r in recetas:
        medicamentos_count = len(json.loads(r.medicamentos)) if r.medicamentos else 0
        recetas_html += f"""
        <div style="padding: 12px; border-bottom: 1px solid #eee;">
            <div style="display: flex; justify-content: space-between;">
                <strong>📋 Receta #{r.id:06d}</strong>
                <span style="color: #666; font-size: 12px;">{r.fecha_emision}</span>
            </div>
            <p style="font-size: 13px; color: #666; margin-top: 5px;">
                💊 {medicamentos_count} medicamento(s) | Estado: <span style="color: #27ae60;">{r.estado}</span>
            </p>
            <a href="/receta/{r.id}" class="btn btn-success btn-sm" style="margin-top: 5px;">👁️ Ver receta</a>
        </div>
        """
    
    content = f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
            <h2>🏥 {paciente.nombre}</h2>
            <div style="display: flex; gap: 8px;">
                <a href="/cita/nueva?paciente_id={paciente.id}" class="btn btn-success">📅 Agendar Cita</a>
                <a href="/historia-clinica/nueva?paciente_id={paciente.id}" class="btn btn-primary">📁 Nueva Historia</a>
                <a href="/receta/nueva?paciente_id={paciente.id}" class="btn btn-warning">📋 Nueva Receta</a>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
            <div class="card">
                <h3>📋 Información del Paciente</h3>
                <p><strong>Obra Social:</strong> {paciente.obra_social or 'Particular'}</p>
                <p><strong>Email:</strong> {paciente.email or '-'}</p>
                <p><strong>Teléfono:</strong> {paciente.telefono or '-'}</p>
                <p><strong>Dirección:</strong> {paciente.direccion or '-'}</p>
                <p><strong>Fecha de nacimiento:</strong> {paciente.fecha_nacimiento or '-'}</p>
                <p><strong>Grupo Sanguíneo:</strong> {grupo_badge}</p>
                
                <div style="background:#fff3e0; padding:15px; border-radius:8px; margin:15px 0;">
                    <strong>⚠️ Alergias:</strong><br>
                    {paciente.alergias or 'No se registran alergias'}
                </div>
                
                <p><strong>Notas:</strong><br>{paciente.notas or '-'}</p>
                <p><strong>Registrado:</strong> {paciente.fecha_creacion} por Dr/a. {paciente.usuario.nombre_completo}</p>
                
                <div style="margin-top:20px;">
                    <a href="/paciente/editar/{paciente.id}" class="btn btn-warning">✏️ Editar</a>
                    <a href="/pacientes" class="btn" style="background:#95a5a6;color:white;">← Volver</a>
                </div>
            </div>
            
            <div class="card">
                <h3>🩺 Registrar Consulta</h3>
                <form action="/paciente/{paciente.id}/interaccion" method="POST">
                    <div class="form-group">
                        <label>Tipo</label>
                        <select name="tipo" class="form-control">
                            <option value="Consulta">🩺 Consulta Médica</option>
                            <option value="Llamada">📞 Llamada</option>
                            <option value="Email">📧 Email</option>
                            <option value="Nota">📝 Nota</option>
                        </select>
                    </div>
                    <div class="form-group">
                        <label>Descripción / Diagnóstico</label>
                        <textarea name="descripcion" class="form-control" rows="4" required></textarea>
                    </div>
                    <button type="submit" class="btn btn-primary">💾 Guardar Consulta</button>
                </form>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; margin-top: 20px;">
            <div class="card">
                <h3>📁 Historias Clínicas</h3>
                {historias_html if historias else '<p style="color:#999;">No hay historias clínicas</p>'}
                <a href="/historias-clinicas?paciente_id={paciente.id}" style="display: block; text-align: center; margin-top: 15px;">Ver todas →</a>
            </div>
            
            <div class="card">
                <h3>📋 Recetas Médicas</h3>
                {recetas_html if recetas else '<p style="color:#999;">No hay recetas emitidas</p>'}
                <a href="/recetas" style="display: block; text-align: center; margin-top: 15px;">Ver todas →</a>
            </div>
            
            <div class="card">
                <h3>📅 Historial de Consultas</h3>
                <div style="max-height: 250px; overflow-y: auto;">
                    {interacciones_html if interacciones else '<p style="color:#999;">No hay consultas registradas</p>'}
                </div>
            </div>
        </div>
    """
    return base_html(content, f"Paciente: {paciente.nombre}")

@app.route('/paciente/<int:id>/interaccion', methods=['POST'])
@login_required
def añadir_interaccion_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    interaccion = Interaccion(
        tipo=request.form['tipo'],
        descripcion=request.form['descripcion'],
        paciente_id=paciente.id,
        usuario_id=user.id
    )
    db.session.add(interaccion)
    db.session.commit()
    flash('✅ Consulta registrada correctamente')
    return redirect(f'/paciente/{paciente.id}')

@app.route('/paciente/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') != 'admin' and paciente.usuario_id != user.id:
        flash('❌ No tienes permiso para editar este paciente')
        return redirect('/pacientes')
    
    if request.method == 'POST':
        paciente.nombre = request.form['nombre']
        paciente.email = request.form.get('email', '')
        paciente.telefono = request.form.get('telefono', '')
        paciente.obra_social = request.form.get('obra_social', '')
        paciente.direccion = request.form.get('direccion', '')
        paciente.fecha_nacimiento = request.form.get('fecha_nacimiento', '')
        paciente.grupo_sanguineo = request.form.get('grupo_sanguineo', '')
        paciente.alergias = request.form.get('alergias', '')
        paciente.notas = request.form.get('notas', '')
        db.session.commit()
        flash('✅ Paciente actualizado correctamente')
        return redirect(f'/paciente/{paciente.id}')
    
    content = f"""
        <h2>✏️ Editar Paciente</h2>
        <form method="POST" style="max-width: 700px;">
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Nombre completo *</label>
                    <input type="text" name="nombre" class="form-control" value="{paciente.nombre}" required>
                </div>
                <div class="form-group">
                    <label>Fecha de nacimiento</label>
                    <input type="date" name="fecha_nacimiento" class="form-control" value="{paciente.fecha_nacimiento or ''}">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Email</label>
                    <input type="email" name="email" class="form-control" value="{paciente.email or ''}">
                </div>
                <div class="form-group">
                    <label>Teléfono</label>
                    <input type="tel" name="telefono" class="form-control" value="{paciente.telefono or ''}">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Obra Social</label>
                    <input type="text" name="obra_social" class="form-control" value="{paciente.obra_social or ''}">
                </div>
                <div class="form-group">
                    <label>Grupo Sanguíneo</label>
                    <select name="grupo_sanguineo" class="form-control">
                        <option value="">Desconocido</option>
                        {''.join([f'<option value="{g}" {"selected" if paciente.grupo_sanguineo == g else ""}>{g}</option>' for g in ['A+', 'A-', 'B+', 'B-', 'AB+', 'AB-', 'O+', 'O-']])}
                    </select>
                </div>
            </div>
            
            <div class="form-group">
                <label>Dirección</label>
                <input type="text" name="direccion" class="form-control" value="{paciente.direccion or ''}">
            </div>
            
            <div class="form-group">
                <label>⚠️ Alergias</label>
                <textarea name="alergias" class="form-control" rows="2">{paciente.alergias or ''}</textarea>
            </div>
            
            <div class="form-group">
                <label>Notas adicionales</label>
                <textarea name="notas" class="form-control" rows="3">{paciente.notas or ''}</textarea>
            </div>
            
            <button type="submit" class="btn btn-primary">💾 Actualizar</button>
            <a href="/paciente/{paciente.id}" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Editar Paciente")

@app.route('/paciente/eliminar/<int:id>')
@login_required
def eliminar_paciente(id):
    paciente = Paciente.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') != 'admin' and paciente.usuario_id != user.id:
        flash('❌ No tienes permiso para eliminar este paciente')
        return redirect('/pacientes')
    
    # Eliminar interacciones asociadas
    Interaccion.query.filter_by(paciente_id=paciente.id).delete()
    
    db.session.delete(paciente)
    db.session.commit()
    flash('✅ Paciente eliminado correctamente')
    return redirect('/pacientes')


# ========== GESTIÓN DE CITAS ==========

@app.route('/citas')
@login_required
def citas():
    user = db.session.get(Usuario, session.get('user_id'))
    hoy = datetime.now().strftime('%Y-%m-%d')
    
    # Filtros
    estado = request.args.get('estado', 'pendientes')
    prioridad = request.args.get('prioridad', 'todas')
    fecha = request.args.get('fecha', '')
    
    if session.get('rol') == 'admin':
        query = Cita.query
    else:
        query = Cita.query.filter_by(usuario_id=user.id)
    
    if estado == 'pendientes':
        query = query.filter_by(asistida=False)
    elif estado == 'completadas':
        query = query.filter_by(asistida=True)
    
    if prioridad != 'todas':
        query = query.filter_by(prioridad=prioridad)
    
    if fecha:
        query = query.filter_by(fecha_cita=fecha)
    
    citas_list = query.order_by(Cita.fecha_cita.desc(), Cita.hora_inicio).all()
    
    # Estadísticas
    total_citas = len(citas_list)
    citas_pendientes = sum(1 for c in citas_list if not c.asistida)
    citas_completadas = sum(1 for c in citas_list if c.asistida)
    citas_hoy = sum(1 for c in citas_list if c.fecha_cita == hoy and not c.asistida)
    
    citas_html = ""
    for c in citas_list[:30]:
        prioridad_color = {'alta': '#e74c3c', 'media': '#f39c12', 'baja': '#27ae60'}.get(c.prioridad, '#3498db')
        estado_icono = '✅' if c.asistida else '⏳'
        estado_texto = 'Realizada' if c.asistida else 'Pendiente'
        
        citas_html += f"""
        <tr>
            <td><span style="background: {prioridad_color}; width: 10px; height: 10px; display: inline-block; border-radius: 50%; margin-right: 8px;"></span> {c.titulo[:40]}...</td>
            <td>{c.fecha_cita or '-'}</td>
            <td>{c.hora_inicio or '-'} - {c.hora_fin or ''}</td>
            <td>{c.tipo_consulta}</td>
            <td>{c.paciente.nombre if c.paciente else 'Sin paciente'}</td>
            <td>{c.medico.nombre_completo if c.medico else 'Sin asignar'}</td>
            <td><span style="color: {'#27ae60' if c.asistida else '#e74c3c'};">{estado_icono} {estado_texto}</span></td>
            <td>
                <a href="/cita/{c.id}" class="btn btn-primary btn-sm">👁️</a>
                {f'<a href="/cita/completar/{c.id}" class="btn btn-success btn-sm">✅</a>' if not c.asistida else ''}
                <a href="/cita/eliminar/{c.id}" class="btn btn-danger btn-sm" onclick="return confirm(\'¿Eliminar cita?\')">🗑️</a>
            </td>
        </tr>
        """
    
    # Selector de médicos para admin
    selector_medico = ""
    if session.get('rol') == 'admin':
        medicos = Usuario.query.filter_by(rol='medico').all()
        medico_id = request.args.get('medico_id', type=int)
        if medico_id:
            citas_list = [c for c in citas_list if c.usuario_id == medico_id]
        selector_medico = f"""
        <select name="medico_id" class="form-control" style="width: auto;" onchange="window.location.href='?medico_id='+this.value">
            <option value="">Todos los médicos</option>
            {''.join([f'<option value="{m.id}" {"selected" if medico_id == m.id else ""}>{m.nombre_completo}</option>' for m in medicos])}
        </select>
        """
    
    content = f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 25px;">
            <h2 style="margin:0;">📅 Gestión de Citas</h2>
            <div>
                <a href="/cita/nueva" class="btn btn-success">➕ Nueva Cita</a>
                <a href="/admin/calendario-citas" class="btn btn-primary">📅 Vista Calendario</a>
            </div>
        </div>
        
        <!-- KPIs de Citas -->
        <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin-bottom: 25px;">
            <div class="stat-card" style="background: linear-gradient(135deg, #667eea, #764ba2);">
                <div class="stat-number">{total_citas}</div>
                <div class="stat-label">Total Citas</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #f093fb, #f5576c);">
                <div class="stat-number">{citas_pendientes}</div>
                <div class="stat-label">Pendientes</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #4facfe, #00f2fe);">
                <div class="stat-number">{citas_hoy}</div>
                <div class="stat-label">Para Hoy</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #43e97b, #38f9d7);">
                <div class="stat-number">{citas_completadas}</div>
                <div class="stat-label">Realizadas</div>
            </div>
        </div>
        
        <!-- Filtros -->
        <div class="card">
            <form method="GET" style="display: flex; gap: 10px; flex-wrap: wrap; align-items: center;">
                <select name="estado" class="form-control" style="width: auto;">
                    <option value="todas" {'selected' if estado == 'todas' else ''}>Todos los estados</option>
                    <option value="pendientes" {'selected' if estado == 'pendientes' else ''}>⏳ Pendientes</option>
                    <option value="completadas" {'selected' if estado == 'completadas' else ''}>✅ Realizadas</option>
                </select>
                <select name="prioridad" class="form-control" style="width: auto;">
                    <option value="todas" {'selected' if prioridad == 'todas' else ''}>Todas las prioridades</option>
                    <option value="alta">🔴 Alta</option>
                    <option value="media">🟡 Media</option>
                    <option value="baja">🟢 Baja</option>
                </select>
                <input type="date" name="fecha" class="form-control" style="width: auto;" value="{fecha}" placeholder="Filtrar por fecha">
                {selector_medico}
                <button type="submit" class="btn btn-primary">🔍 Filtrar</button>
                <a href="/citas" class="btn" style="background:#95a5a6;color:white;">Limpiar</a>
            </form>
        </div>
        
        <!-- Tabla de Citas -->
        <div class="card" style="margin-top: 20px;">
            <table>
                <thead>
                    <tr>
                        <th>Título</th>
                        <th>Fecha</th>
                        <th>Hora</th>
                        <th>Tipo</th>
                        <th>Paciente</th>
                        <th>Médico</th>
                        <th>Estado</th>
                        <th>Acciones</th>
                    </tr>
                </thead>
                <tbody>
                    {citas_html if citas_list else '<tr><td colspan="8" style="text-align:center; padding:30px;">No hay citas registradas</td></tr>'}
                </tbody>
            </table>
        </div>
    """
    return base_html(content, "Citas")

@app.route('/cita/nueva', methods=['GET', 'POST'])
@login_required
def nueva_cita():
    user = db.session.get(Usuario, session.get('user_id'))
    
    if request.method == 'POST':
        cita = Cita(
            titulo=request.form['titulo'],
            descripcion=request.form.get('descripcion', ''),
            fecha_cita=request.form.get('fecha_cita', ''),
            hora_inicio=request.form.get('hora_inicio', ''),
            hora_fin=request.form.get('hora_fin', ''),
            tipo_consulta=request.form.get('tipo_consulta', 'Presencial'),
            prioridad=request.form.get('prioridad', 'media'),
            usuario_id=request.form.get('usuario_id', user.id),
            paciente_id=request.form.get('paciente_id') or None
        )
        db.session.add(cita)
        db.session.commit()
        flash('✅ Cita agendada correctamente')
        return redirect('/citas')
    
    # Obtener pacientes y médicos para los selects
    if session.get('rol') == 'admin':
        pacientes = Paciente.query.order_by(Paciente.nombre).all()
        medicos = Usuario.query.filter_by(rol='medico').all()
    else:
        pacientes = Paciente.query.filter_by(usuario_id=user.id).order_by(Paciente.nombre).all()
        medicos = [user]
    
    paciente_id = request.args.get('paciente_id', '')
    
    pacientes_options = "".join([f'<option value="{p.id}" {"selected" if str(p.id) == paciente_id else ""}>{p.nombre}</option>' for p in pacientes])
    medicos_options = "".join([f'<option value="{m.id}" {"selected" if m.id == user.id else ""}>{m.nombre_completo} ({m.departamento})</option>' for m in medicos])
    
    content = f"""
        <h2>📅 Nueva Cita</h2>
        <form method="POST" style="max-width: 700px;">
            <div class="form-group">
                <label>Título / Motivo *</label>
                <input type="text" name="titulo" class="form-control" placeholder="Ej: Consulta general, Control..." required>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Paciente</label>
                    <select name="paciente_id" class="form-control">
                        <option value="">Seleccionar paciente...</option>
                        {pacientes_options}
                    </select>
                </div>
                <div class="form-group">
                    <label>Médico asignado</label>
                    <select name="usuario_id" class="form-control" required>
                        {medicos_options}
                    </select>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Fecha *</label>
                    <input type="date" name="fecha_cita" class="form-control" required>
                </div>
                <div class="form-group">
                    <label>Hora inicio</label>
                    <input type="time" name="hora_inicio" class="form-control">
                </div>
                <div class="form-group">
                    <label>Hora fin</label>
                    <input type="time" name="hora_fin" class="form-control">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Tipo de consulta</label>
                    <select name="tipo_consulta" class="form-control">
                        <option value="Presencial">🏥 Presencial</option>
                        <option value="Virtual">💻 Virtual</option>
                        <option value="Telefónica">📞 Telefónica</option>
                        <option value="Domiciliaria">🏠 Domiciliaria</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Prioridad</label>
                    <select name="prioridad" class="form-control">
                        <option value="alta">🔴 Alta</option>
                        <option value="media" selected>🟡 Media</option>
                        <option value="baja">🟢 Baja</option>
                    </select>
                </div>
            </div>
            
            <div class="form-group">
                <label>Descripción / Notas</label>
                <textarea name="descripcion" class="form-control" rows="3" placeholder="Motivo de consulta, observaciones..."></textarea>
            </div>
            
            <button type="submit" class="btn btn-success">💾 Agendar Cita</button>
            <a href="/citas" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Nueva Cita")

@app.route('/cita/<int:id>')
@login_required
def ver_cita(id):
    cita = Cita.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') != 'admin' and cita.usuario_id != user.id:
        flash('❌ No tienes permiso para ver esta cita')
        return redirect('/citas')
    
    comentarios_html = ""
    for com in cita.comentarios:
        comentarios_html += f"""
        <div style="background: #f8f9fa; padding: 12px; border-radius: 8px; margin-bottom: 10px;">
            <strong>{com.usuario.nombre_completo}</strong>
            <small style="color: #666;">{com.fecha}</small>
            <p style="margin-top: 5px;">{com.texto}</p>
        </div>
        """
    
    content = f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
            <h2>📋 {cita.titulo}</h2>
            <div>
                {f'<a href="/cita/completar/{cita.id}" class="btn btn-success">✅ Marcar como Realizada</a>' if not cita.asistida else ''}
                <a href="/cita/editar/{cita.id}" class="btn btn-warning">✏️ Editar</a>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
            <div class="card">
                <h3>📅 Información de la Cita</h3>
                <p><strong>Fecha:</strong> {cita.fecha_cita or 'Sin fecha'}</p>
                <p><strong>Hora:</strong> {cita.hora_inicio or '--:--'} - {cita.hora_fin or '--:--'}</p>
                <p><strong>Tipo de consulta:</strong> {cita.tipo_consulta}</p>
                <p><strong>Prioridad:</strong> 
                    <span style="background: {'#e74c3c' if cita.prioridad == 'alta' else '#f39c12' if cita.prioridad == 'media' else '#27ae60'}; color: white; padding: 3px 10px; border-radius: 20px;">
                        {cita.prioridad.upper()}
                    </span>
                </p>
                <p><strong>Estado:</strong> {'✅ Realizada' if cita.asistida else '⏳ Pendiente'}</p>
                <p><strong>Descripción:</strong><br>{cita.descripcion or 'Sin descripción'}</p>
            </div>
            
            <div class="card">
                <h3>👥 Participantes</h3>
                <div style="display: flex; align-items: center; gap: 15px; margin-bottom: 20px;">
                    <div style="width: 50px; height: 50px; background: #1a5276; color: white; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 20px;">
                        👨‍⚕️
                    </div>
                    <div>
                        <strong>Médico:</strong><br>
                        Dr/a. {cita.medico.nombre_completo if cita.medico else 'No asignado'}<br>
                        <small>{cita.medico.departamento if cita.medico else ''}</small>
                    </div>
                </div>
                
                {f'''
                <div style="display: flex; align-items: center; gap: 15px;">
                    <div style="width: 50px; height: 50px; background: #27ae60; color: white; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 20px;">
                        🏥
                    </div>
                    <div>
                        <strong>Paciente:</strong><br>
                        <a href="/paciente/{cita.paciente.id}">{cita.paciente.nombre}</a><br>
                        <small>{cita.paciente.obra_social or 'Particular'}</small>
                    </div>
                </div>
                ''' if cita.paciente else '<p>Sin paciente asignado</p>'}
            </div>
        </div>
        
        <div class="card" style="margin-top: 20px;">
            <h3>💬 Comentarios</h3>
            <div style="margin-bottom: 20px; max-height: 300px; overflow-y: auto;">
                {comentarios_html if cita.comentarios else '<p style="color:#999;">No hay comentarios</p>'}
            </div>
            
            <form method="POST" action="/cita/{cita.id}/comentario">
                <div class="form-group">
                    <label>Añadir comentario</label>
                    <textarea name="texto" class="form-control" rows="3" required placeholder="Escribe una nota sobre esta cita..."></textarea>
                </div>
                <button type="submit" class="btn btn-primary">💬 Enviar comentario</button>
            </form>
        </div>
        
        <p style="margin-top: 20px;">
            <a href="/citas" class="btn" style="background:#95a5a6;color:white;">← Volver a Citas</a>
        </p>
    """
    return base_html(content, f"Cita: {cita.titulo[:30]}")

@app.route('/cita/<int:id>/comentario', methods=['POST'])
@login_required
def añadir_comentario_cita(id):
    cita = Cita.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    comentario = Comentario(
        texto=request.form['texto'],
        usuario_id=user.id,
        cita_id=cita.id
    )
    db.session.add(comentario)
    db.session.commit()
    flash('✅ Comentario añadido')
    return redirect(f'/cita/{id}')

@app.route('/cita/completar/<int:id>')
@login_required
def completar_cita(id):
    cita = Cita.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') != 'admin' and cita.usuario_id != user.id:
        flash('❌ No tienes permiso')
        return redirect('/citas')
    
    cita.asistida = True
    db.session.commit()
    flash('✅ Cita marcada como realizada')
    return redirect('/citas')

@app.route('/cita/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cita(id):
    cita = Cita.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') != 'admin' and cita.usuario_id != user.id:
        flash('❌ No tienes permiso')
        return redirect('/citas')
    
    if request.method == 'POST':
        cita.titulo = request.form['titulo']
        cita.descripcion = request.form.get('descripcion', '')
        cita.fecha_cita = request.form.get('fecha_cita', '')
        cita.hora_inicio = request.form.get('hora_inicio', '')
        cita.hora_fin = request.form.get('hora_fin', '')
        cita.tipo_consulta = request.form.get('tipo_consulta', 'Presencial')
        cita.prioridad = request.form.get('prioridad', 'media')
        cita.paciente_id = request.form.get('paciente_id') or None
        
        if session.get('rol') == 'admin':
            cita.usuario_id = request.form.get('usuario_id', cita.usuario_id)
        
        db.session.commit()
        flash('✅ Cita actualizada')
        return redirect(f'/cita/{cita.id}')
    
    pacientes = Paciente.query.order_by(Paciente.nombre).all()
    pacientes_options = "".join([f'<option value="{p.id}" {"selected" if p.id == cita.paciente_id else ""}>{p.nombre}</option>' for p in pacientes])
    
    medicos = Usuario.query.filter_by(rol='medico').all()
    medicos_options = "".join([f'<option value="{m.id}" {"selected" if m.id == cita.usuario_id else ""}>{m.nombre_completo}</option>' for m in medicos])
    
    content = f"""
        <h2>✏️ Editar Cita</h2>
        <form method="POST" style="max-width: 700px;">
            <div class="form-group">
                <label>Título *</label>
                <input type="text" name="titulo" class="form-control" value="{cita.titulo}" required>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Paciente</label>
                    <select name="paciente_id" class="form-control">
                        <option value="">Sin paciente</option>
                        {pacientes_options}
                    </select>
                </div>
                {f'''
                <div class="form-group">
                    <label>Médico</label>
                    <select name="usuario_id" class="form-control">
                        {medicos_options}
                    </select>
                </div>
                ''' if session.get('rol') == 'admin' else ''}
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Fecha</label>
                    <input type="date" name="fecha_cita" class="form-control" value="{cita.fecha_cita or ''}">
                </div>
                <div class="form-group">
                    <label>Hora inicio</label>
                    <input type="time" name="hora_inicio" class="form-control" value="{cita.hora_inicio or ''}">
                </div>
                <div class="form-group">
                    <label>Hora fin</label>
                    <input type="time" name="hora_fin" class="form-control" value="{cita.hora_fin or ''}">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Tipo de consulta</label>
                    <select name="tipo_consulta" class="form-control">
                        {''.join([f'<option value="{t}" {"selected" if t == cita.tipo_consulta else ""}>{t}</option>' for t in ['Presencial', 'Virtual', 'Telefónica', 'Domiciliaria']])}
                    </select>
                </div>
                <div class="form-group">
                    <label>Prioridad</label>
                    <select name="prioridad" class="form-control">
                        {''.join([f'<option value="{p}" {"selected" if p == cita.prioridad else ""}>{p.upper()}</option>' for p in ['alta', 'media', 'baja']])}
                    </select>
                </div>
            </div>
            
            <div class="form-group">
                <label>Descripción</label>
                <textarea name="descripcion" class="form-control" rows="3">{cita.descripcion or ''}</textarea>
            </div>
            
            <button type="submit" class="btn btn-primary">💾 Guardar Cambios</button>
            <a href="/cita/{cita.id}" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Editar Cita")

@app.route('/cita/eliminar/<int:id>')
@login_required
def eliminar_cita(id):
    cita = Cita.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') != 'admin' and cita.usuario_id != user.id:
        flash('❌ No tienes permiso')
        return redirect('/citas')
    
    # Eliminar comentarios asociados
    Comentario.query.filter_by(cita_id=cita.id).delete()
    db.session.delete(cita)
    db.session.commit()
    flash('✅ Cita eliminada')
    return redirect('/citas')

@app.route('/admin/asignar-cita', methods=['GET', 'POST'])
@login_required
def admin_asignar_cita():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    if request.method == 'POST':
        cita = Cita(
            titulo=request.form['titulo'],
            descripcion=request.form.get('descripcion', ''),
            fecha_cita=request.form.get('fecha_cita', ''),
            hora_inicio=request.form.get('hora_inicio', ''),
            hora_fin=request.form.get('hora_fin', ''),
            tipo_consulta=request.form.get('tipo_consulta', 'Presencial'),
            prioridad=request.form.get('prioridad', 'media'),
            usuario_id=request.form['usuario_id'],
            paciente_id=request.form.get('paciente_id') or None
        )
        db.session.add(cita)
        db.session.commit()
        flash('✅ Cita asignada correctamente')
        return redirect('/admin/asignar-cita')
    
    medicos = Usuario.query.filter_by(rol='medico').all()
    pacientes = Paciente.query.order_by(Paciente.nombre).all()
    
    medicos_options = "".join([f'<option value="{m.id}">{m.nombre_completo} ({m.departamento})</option>' for m in medicos])
    pacientes_options = "".join([f'<option value="{p.id}">{p.nombre}</option>' for p in pacientes])
    
    content = f"""
        <h2>📋 Asignar Nueva Cita</h2>
        <form method="POST" style="max-width: 700px;">
            <div class="form-group">
                <label>Médico *</label>
                <select name="usuario_id" class="form-control" required>
                    <option value="">Seleccionar médico...</option>
                    {medicos_options}
                </select>
            </div>
            
            <div class="form-group">
                <label>Paciente</label>
                <select name="paciente_id" class="form-control">
                    <option value="">Sin paciente (solo bloqueo horario)</option>
                    {pacientes_options}
                </select>
            </div>
            
            <div class="form-group">
                <label>Título / Motivo *</label>
                <input type="text" name="titulo" class="form-control" required>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Fecha *</label>
                    <input type="date" name="fecha_cita" class="form-control" required>
                </div>
                <div class="form-group">
                    <label>Hora inicio</label>
                    <input type="time" name="hora_inicio" class="form-control">
                </div>
                <div class="form-group">
                    <label>Hora fin</label>
                    <input type="time" name="hora_fin" class="form-control">
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Tipo de consulta</label>
                    <select name="tipo_consulta" class="form-control">
                        <option value="Presencial">🏥 Presencial</option>
                        <option value="Virtual">💻 Virtual</option>
                        <option value="Telefónica">📞 Telefónica</option>
                        <option value="Domiciliaria">🏠 Domiciliaria</option>
                    </select>
                </div>
                <div class="form-group">
                    <label>Prioridad</label>
                    <select name="prioridad" class="form-control">
                        <option value="alta">🔴 Alta</option>
                        <option value="media" selected>🟡 Media</option>
                        <option value="baja">🟢 Baja</option>
                    </select>
                </div>
            </div>
            
            <div class="form-group">
                <label>Descripción</label>
                <textarea name="descripcion" class="form-control" rows="3"></textarea>
            </div>
            
            <button type="submit" class="btn btn-primary">Asignar Cita</button>
            <a href="/citas" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Asignar Cita")

@app.route('/admin/calendario-citas')
@login_required
def calendario_citas():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    hoy = datetime.now().date()
    lunes = hoy - timedelta(days=hoy.weekday())
    semana = [lunes + timedelta(days=i) for i in range(7)]
    
    # Obtener todas las citas de la semana
    citas_semana = []
    for c in Cita.query.all():
        if c.fecha_cita:
            try:
                fecha_c = datetime.strptime(c.fecha_cita, '%Y-%m-%d').date()
                if lunes <= fecha_c <= lunes + timedelta(days=6):
                    citas_semana.append(c)
            except:
                pass
    
    # Construir tabla de calendario
    dias_html = ""
    for dia in semana:
        citas_dia = [c for c in citas_semana if datetime.strptime(c.fecha_cita, '%Y-%m-%d').date() == dia]
        es_hoy = dia == hoy
        
        citas_dia_html = ""
        for c in citas_dia:
            prioridad_color = {'alta': '#e74c3c', 'media': '#f39c12', 'baja': '#27ae60'}.get(c.prioridad, '#3498db')
            estado_icono = '✅' if c.asistida else '⏳'
            citas_dia_html += f"""
            <div style="background: {prioridad_color}; color: white; padding: 5px 8px; margin: 3px 0; border-radius: 4px; font-size: 11px;">
                <a href="/cita/{c.id}" style="color: white; text-decoration: none;">
                    {estado_icono} {c.hora_inicio or '--:--'} - {c.medico.nombre_completo.split()[0]}<br>
                    <small>{c.titulo[:20]}...</small>
                </a>
            </div>
            """
        
        dias_html += f"""
        <td style="vertical-align: top; padding: 10px; border: 1px solid #ddd; {'background: #e8f4f8;' if es_hoy else ''}; width: 14%;">
            <strong>{dia.strftime('%a')}<br>{dia.strftime('%d/%m')}</strong>
            <div style="margin-top: 10px; max-height: 400px; overflow-y: auto;">
                {citas_dia_html if citas_dia else '<p style="color: #999; font-size: 11px;">Sin citas</p>'}
            </div>
        </td>
        """
    
    content = f"""
        <h2>📅 Calendario Semanal de Citas</h2>
        <p>Semana del <strong>{lunes.strftime('%d/%m/%Y')}</strong> al <strong>{(lunes + timedelta(days=6)).strftime('%d/%m/%Y')}</strong></p>
        
        <table style="width: 100%; border-collapse: collapse; table-layout: fixed;">
            <tr>{dias_html}</tr>
        </table>
        
        <p style="margin-top: 20px;">
            <a href="/citas" class="btn btn-primary">← Volver a Citas</a>
            <a href="/cita/nueva" class="btn btn-success">➕ Nueva Cita</a>
        </p>
    """
    return base_html(content, "Calendario de Citas")


# ========== GESTIÓN DE NOTAS ==========

@app.route('/nota/nueva', methods=['GET', 'POST'])
@login_required
def nueva_nota():
    if request.method == 'POST':
        nota = NotaPersonal(
            titulo=request.form['titulo'],
            contenido=request.form['contenido'],
            color=request.form.get('color', '#f39c12'),
            usuario_id=session['user_id']
        )
        db.session.add(nota)
        db.session.commit()
        flash('✅ Nota creada correctamente')
        return redirect('/mi-espacio')
    
    colores = ['#f39c12', '#e74c3c', '#3498db', '#2ecc71', '#9b59b6', '#1abc9c']
    colores_html = "".join([f'<span class="color-option" style="background: {c}; width: 30px; height: 30px; border-radius: 50%; cursor: pointer; border: 2px solid transparent;" onclick="selectColor(\'{c}\')"></span>' for c in colores])
    
    content = f"""
        <h2>📓 Nueva Nota</h2>
        <form method="POST" style="max-width: 500px;">
            <div class="form-group">
                <label>Título</label>
                <input type="text" name="titulo" class="form-control" required>
            </div>
            <div class="form-group">
                <label>Contenido</label>
                <textarea name="contenido" class="form-control" rows="6" required></textarea>
            </div>
            <div class="form-group">
                <label>Color</label>
                <div style="display: flex; gap: 10px; margin: 10px 0;">
                    {colores_html}
                </div>
                <input type="hidden" name="color" id="colorSelected" value="#f39c12">
            </div>
            <button type="submit" class="btn btn-primary">💾 Guardar Nota</button>
            <a href="/mi-espacio" class="btn" style="background: #95a5a6; color: white;">Cancelar</a>
        </form>
        <script>
            function selectColor(color) {{
                document.querySelectorAll('.color-option').forEach(c => c.style.borderColor = 'transparent');
                event.target.style.borderColor = '#333';
                document.getElementById('colorSelected').value = color;
            }}
        </script>
    """
    return base_html(content, "Nueva Nota")

@app.route('/nota/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_nota(id):
    nota = NotaPersonal.query.get_or_404(id)
    if nota.usuario_id != session['user_id']:
        flash('❌ No tienes permiso para editar esta nota')
        return redirect('/mi-espacio')
    
    if request.method == 'POST':
        nota.titulo = request.form['titulo']
        nota.contenido = request.form['contenido']
        nota.color = request.form.get('color', nota.color)
        db.session.commit()
        flash('✅ Nota actualizada')
        return redirect('/mi-espacio')
    
    colores = ['#f39c12', '#e74c3c', '#3498db', '#2ecc71', '#9b59b6', '#1abc9c']
    colores_html = "".join([f'<span class="color-option" style="background: {c}; width: 30px; height: 30px; border-radius: 50%; cursor: pointer; border: 2px solid {"#333" if c == nota.color else "transparent"};" onclick="selectColor(\'{c}\')"></span>' for c in colores])
    
    content = f"""
        <h2>✏️ Editar Nota</h2>
        <form method="POST" style="max-width: 500px;">
            <div class="form-group">
                <label>Título</label>
                <input type="text" name="titulo" class="form-control" value="{nota.titulo}" required>
            </div>
            <div class="form-group">
                <label>Contenido</label>
                <textarea name="contenido" class="form-control" rows="6" required>{nota.contenido}</textarea>
            </div>
            <div class="form-group">
                <label>Color</label>
                <div style="display: flex; gap: 10px; margin: 10px 0;">
                    {colores_html}
                </div>
                <input type="hidden" name="color" id="colorSelected" value="{nota.color}">
            </div>
            <button type="submit" class="btn btn-primary">💾 Actualizar</button>
            <a href="/mi-espacio" class="btn" style="background: #95a5a6; color: white;">Cancelar</a>
        </form>
        <script>
            function selectColor(color) {{
                document.querySelectorAll('.color-option').forEach(c => c.style.borderColor = 'transparent');
                event.target.style.borderColor = '#333';
                document.getElementById('colorSelected').value = color;
            }}
        </script>
    """
    return base_html(content, "Editar Nota")

@app.route('/nota/eliminar/<int:id>')
@login_required
def eliminar_nota(id):
    nota = NotaPersonal.query.get_or_404(id)
    if nota.usuario_id == session['user_id']:
        db.session.delete(nota)
        db.session.commit()
        flash('✅ Nota eliminada')
    return redirect('/mi-espacio')

# ========== GESTIÓN DE HISTORIAS CLÍNICAS ==========

@app.route('/historias-clinicas')
@login_required
def historias_clinicas():
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') == 'admin':
        tipos = TipoHistoriaClinica.query.all()
        historias = HistoriaClinica.query.order_by(HistoriaClinica.id.desc()).all()
    else:
        tipos = TipoHistoriaClinica.query.filter(
            (TipoHistoriaClinica.departamento == user.departamento) | 
            (TipoHistoriaClinica.departamento == 'General')
        ).all()
        historias = HistoriaClinica.query.filter_by(usuario_id=user.id).order_by(HistoriaClinica.id.desc()).all()
    
    # Filtrar por tipo
    tipo_id = request.args.get('tipo', type=int)
    if tipo_id:
        historias = [h for h in historias if h.tipo_id == tipo_id]
    
    # Filtrar por paciente
    paciente_id = request.args.get('paciente_id', type=int)
    if paciente_id:
        historias = [h for h in historias if h.paciente_id == paciente_id]
    
    historias_html = ""
    for h in historias[:20]:
        docs_count = len(h.documentos)
        color_tipo = h.tipo.color if h.tipo else '#3498db'
        estado_class = {'Abierto': 'estado-abierto', 'En proceso': 'estado-proceso', 'Cerrado': 'estado-cerrado'}.get(h.estado, '')
        
        historias_html += f"""
        <div class="historia-card" style="border-top-color: {color_tipo}; cursor: pointer;" onclick="window.location.href='/historia-clinica/{h.id}'">
            <div class="historia-header">
                <span class="historia-tipo" style="background: {color_tipo};">{h.tipo.nombre if h.tipo else 'General'}</span>
                <span class="historia-estado {estado_class}">{h.estado}</span>
            </div>
            <h3 style="margin: 10px 0;">{h.titulo[:40]}{'...' if len(h.titulo) > 40 else ''}</h3>
            <p style="color: #666; font-size: 14px;"><strong>Paciente:</strong> {h.paciente.nombre if h.paciente else 'Sin asignar'}</p>
            <p style="color: #666; font-size: 14px;"><strong>Diagnóstico:</strong> {h.diagnostico[:60] if h.diagnostico else 'Pendiente'}...</p>
            <div style="display: flex; justify-content: space-between; margin-top: 15px; font-size: 12px; color: #999;">
                <span>👨‍⚕️ Dr/a. {h.usuario.nombre_completo}</span>
                <span>📎 {docs_count} archivos</span>
            </div>
            <div style="margin-top: 10px; font-size: 12px; color: #999;">
                📅 {h.fecha_creacion}
            </div>
        </div>
        """
    
    # Filtros por tipo
    tipos_html = "".join([f'<a href="?tipo={t.id}" class="btn btn-sm" style="background: {t.color}; color: white; margin: 2px;">{t.nombre}</a>' for t in tipos])
    
    content = f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
            <h2>📁 Historias Clínicas</h2>
            <a href="/historia-clinica/nueva" class="btn btn-success">➕ Nueva Historia Clínica</a>
        </div>
        
        <div style="margin-bottom: 20px;">
            <a href="/historias-clinicas" class="btn btn-primary btn-sm">Todas</a>
            {tipos_html}
        </div>
        
        <div class="historias-grid">
            {historias_html if historias else '<p style="grid-column:1/-1; text-align:center; color:#999; padding:40px;">📁 No hay historias clínicas registradas</p>'}
        </div>
        
        {f'<p style="text-align:center; margin-top:20px; color:#666;">Mostrando {min(20, len(historias))} de {len(historias)} historias</p>' if len(historias) > 20 else ''}
    """
    return base_html(content, "Historias Clínicas")

@app.route('/historia-clinica/nueva', methods=['GET', 'POST'])
@login_required
def nueva_historia_clinica():
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') == 'admin':
        tipos = TipoHistoriaClinica.query.all()
    else:
        tipos = TipoHistoriaClinica.query.filter(
            (TipoHistoriaClinica.departamento == user.departamento) | 
            (TipoHistoriaClinica.departamento == 'General')
        ).all()
    
    if request.method == 'POST':
        historia = HistoriaClinica(
            tipo_id=request.form['tipo_id'],
            titulo=request.form['titulo'],
            descripcion=request.form.get('descripcion', ''),
            diagnostico=request.form.get('diagnostico', ''),
            tratamiento=request.form.get('tratamiento', ''),
            estado=request.form.get('estado', 'Abierto'),
            usuario_id=user.id,
            paciente_id=request.form.get('paciente_id') or None,
            departamento=user.departamento
        )
        db.session.add(historia)
        db.session.commit()
        flash('✅ Historia clínica creada correctamente')
        return redirect(f'/historia-clinica/{historia.id}')
    
    # Obtener pacientes para el select
    if session.get('rol') == 'admin':
        pacientes = Paciente.query.order_by(Paciente.nombre).all()
    else:
        pacientes = Paciente.query.filter_by(usuario_id=user.id).order_by(Paciente.nombre).all()
    
    paciente_id = request.args.get('paciente_id', '')
    
    tipos_options = "".join([f'<option value="{t.id}">{t.nombre}</option>' for t in tipos])
    pacientes_options = "".join([f'<option value="{p.id}" {"selected" if str(p.id) == paciente_id else ""}>{p.nombre}</option>' for p in pacientes])
    
    content = f"""
        <h2>📁 Nueva Historia Clínica</h2>
        <form method="POST" style="max-width: 800px;">
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 15px;">
                <div class="form-group">
                    <label>Tipo de historia *</label>
                    <select name="tipo_id" class="form-control" required>
                        <option value="">Seleccionar...</option>
                        {tipos_options}
                    </select>
                </div>
                <div class="form-group">
                    <label>Paciente</label>
                    <select name="paciente_id" class="form-control">
                        <option value="">Sin paciente asignado</option>
                        {pacientes_options}
                    </select>
                </div>
            </div>
            
            <div class="form-group">
                <label>Título *</label>
                <input type="text" name="titulo" class="form-control" maxlength="200" required placeholder="Ej: Historia clínica general, Control cardiológico...">
            </div>
            
            <div class="form-group">
                <label>Descripción / Motivo de consulta</label>
                <textarea name="descripcion" class="form-control" rows="3" maxlength="1000" placeholder="Describe el motivo de la consulta..."></textarea>
            </div>
            
            <div class="form-group">
                <label>🔬 Diagnóstico</label>
                <textarea name="diagnostico" class="form-control" rows="4" placeholder="Diagnóstico médico..."></textarea>
            </div>
            
            <div class="form-group">
                <label>💊 Tratamiento</label>
                <textarea name="tratamiento" class="form-control" rows="4" placeholder="Tratamiento indicado, medicación, dosis..."></textarea>
            </div>
            
            <div class="form-group">
                <label>Estado</label>
                <select name="estado" class="form-control">
                    <option value="Abierto">Abierto</option>
                    <option value="En proceso">En proceso</option>
                    <option value="Cerrado">Cerrado</option>
                </select>
            </div>
            
            <button type="submit" class="btn btn-primary">💾 Crear Historia Clínica</button>
            <a href="/historias-clinicas" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Nueva Historia Clínica")

@app.route('/historia-clinica/<int:id>', methods=['GET', 'POST'])
@login_required
def ver_historia_clinica(id):
    user = db.session.get(Usuario, session.get('user_id'))
    historia = HistoriaClinica.query.get_or_404(id)
    
    # Verificar permisos
    if session.get('rol') != 'admin' and historia.usuario_id != user.id:
        flash('❌ No tienes permiso para ver esta historia clínica')
        return redirect('/historias-clinicas')
    
    if request.method == 'POST':
        if 'archivo' not in request.files:
            flash('❌ No se seleccionó archivo')
            return redirect(request.url)
        
        file = request.files['archivo']
        if file.filename == '':
            flash('❌ No se seleccionó archivo')
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            unique_name = f"{datetime.now().strftime('%Y%m%d%H%M%S')}_{filename}"
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
            file.save(filepath)
            
            doc = Documento(
                nombre=unique_name,
                nombre_original=filename,
                tipo_archivo=filename.rsplit('.', 1)[1].lower(),
                historia_clinica_id=historia.id,
                usuario_id=user.id
            )
            db.session.add(doc)
            db.session.commit()
            flash('✅ Archivo subido correctamente')
        else:
            flash('❌ Tipo de archivo no permitido')
        
        return redirect(f'/historia-clinica/{id}')
    
    # Lista de documentos
    docs_html = ""
    for d in historia.documentos:
        icono = {'pdf': '📄', 'doc': '📝', 'docx': '📝', 'xls': '📊', 'xlsx': '📊', 'png': '🖼️', 'jpg': '🖼️', 'jpeg': '🖼️'}.get(d.tipo_archivo, '📎')
        docs_html += f"""
        <div style="display: flex; align-items: center; gap: 10px; padding: 10px; background: #f8f9fa; border-radius: 8px; margin-bottom: 8px;">
            <span style="font-size: 24px;">{icono}</span>
            <div style="flex:1;">
                <a href="/uploads/{d.nombre}" target="_blank">{d.nombre_original}</a>
                <div style="font-size:11px; color:#999;">{d.fecha_subida} - {d.usuario.nombre_completo}</div>
            </div>
            <a href="/documento/eliminar/{d.id}" class="btn btn-danger btn-sm" onclick="return confirm('¿Eliminar archivo?')">🗑️</a>
        </div>
        """
    
    content = f"""
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 20px;">
            <h2>📁 {historia.titulo}</h2>
            <div>
                <a href="/historia-clinica/editar/{historia.id}" class="btn btn-warning">✏️ Editar</a>
                <a href="/historia-clinica/eliminar/{historia.id}" class="btn btn-danger" onclick="return confirm('¿Eliminar esta historia clínica?')">🗑️ Eliminar</a>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
            <div class="card">
                <h3>📋 Información General</h3>
                <p><strong>Tipo:</strong> 
                    <span style="background:{historia.tipo.color if historia.tipo else '#3498db'}; color:white; padding:3px 12px; border-radius:20px;">
                        {historia.tipo.nombre if historia.tipo else 'General'}
                    </span>
                </p>
                <p><strong>Estado:</strong> {historia.estado}</p>
                <p><strong>Departamento:</strong> {historia.departamento}</p>
                <p><strong>Médico responsable:</strong> Dr/a. {historia.usuario.nombre_completo}</p>
                <p><strong>Fecha de creación:</strong> {historia.fecha_creacion}</p>
                
                {f'''
                <div style="margin-top: 20px; padding: 15px; background: #e8f4f8; border-radius: 8px;">
                    <strong>🏥 Paciente:</strong><br>
                    <a href="/paciente/{historia.paciente.id}" style="font-size: 18px;">{historia.paciente.nombre}</a><br>
                    <small>{historia.paciente.obra_social or 'Particular'} • {historia.paciente.telefono or 'Sin teléfono'}</small>
                </div>
                ''' if historia.paciente else '<p>Sin paciente asignado</p>'}
            </div>
            
            <div class="card">
                <h3>📄 Descripción</h3>
                <p>{historia.descripcion or 'Sin descripción'}</p>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-top: 20px;">
            <div class="card">
                <h3>🔬 Diagnóstico</h3>
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; min-height: 100px;">
                    {historia.diagnostico or '<span style="color:#999;">Diagnóstico pendiente</span>'}
                </div>
            </div>
            
            <div class="card">
                <h3>💊 Tratamiento</h3>
                <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; min-height: 100px;">
                    {historia.tratamiento or '<span style="color:#999;">Tratamiento pendiente</span>'}
                </div>
            </div>
        </div>
        
        <div class="card" style="margin-top: 20px;">
            <h3>📎 Documentos Adjuntos ({len(historia.documentos)})</h3>
            
            <form method="POST" enctype="multipart/form-data" style="margin-bottom:20px; padding:15px; background:#f8f9fa; border-radius:8px;">
                <div class="form-group">
                    <label>Adjuntar archivo (PDF, Word, Excel, Imagen)</label>
                    <input type="file" name="archivo" class="form-control" required accept=".pdf,.doc,.docx,.xls,.xlsx,.png,.jpg,.jpeg">
                </div>
                <button type="submit" class="btn btn-primary">📤 Subir Archivo</button>
            </form>
            
            <div style="max-height: 300px; overflow-y: auto;">
                {docs_html if docs_html else '<p style="color:#999;">No hay documentos adjuntos</p>'}
            </div>
        </div>
        
        <p style="margin-top:20px;">
            <a href="/historias-clinicas" class="btn" style="background:#95a5a6;color:white;">← Volver a Historias Clínicas</a>
        </p>
    """
    return base_html(content, f"Historia: {historia.titulo[:30]}")

@app.route('/historia-clinica/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_historia_clinica(id):
    user = db.session.get(Usuario, session.get('user_id'))
    historia = HistoriaClinica.query.get_or_404(id)
    
    if session.get('rol') != 'admin' and historia.usuario_id != user.id:
        flash('❌ No tienes permiso para editar esta historia')
        return redirect('/historias-clinicas')
    
    if request.method == 'POST':
        historia.titulo = request.form['titulo']
        historia.descripcion = request.form.get('descripcion', '')
        historia.diagnostico = request.form.get('diagnostico', '')
        historia.tratamiento = request.form.get('tratamiento', '')
        historia.estado = request.form.get('estado', historia.estado)
        historia.paciente_id = request.form.get('paciente_id') or None
        db.session.commit()
        flash('✅ Historia clínica actualizada')
        return redirect(f'/historia-clinica/{id}')
    
    # Obtener pacientes
    if session.get('rol') == 'admin':
        pacientes = Paciente.query.order_by(Paciente.nombre).all()
    else:
        pacientes = Paciente.query.filter_by(usuario_id=user.id).order_by(Paciente.nombre).all()
    
    pacientes_options = "".join([f'<option value="{p.id}" {"selected" if p.id == historia.paciente_id else ""}>{p.nombre}</option>' for p in pacientes])
    
    content = f"""
        <h2>✏️ Editar Historia Clínica</h2>
        <form method="POST" style="max-width: 800px;">
            <div class="form-group">
                <label>Paciente</label>
                <select name="paciente_id" class="form-control">
                    <option value="">Sin paciente asignado</option>
                    {pacientes_options}
                </select>
            </div>
            
            <div class="form-group">
                <label>Título *</label>
                <input type="text" name="titulo" class="form-control" value="{historia.titulo}" maxlength="200" required>
            </div>
            
            <div class="form-group">
                <label>Descripción / Motivo de consulta</label>
                <textarea name="descripcion" class="form-control" rows="3" maxlength="1000">{historia.descripcion or ''}</textarea>
            </div>
            
            <div class="form-group">
                <label>🔬 Diagnóstico</label>
                <textarea name="diagnostico" class="form-control" rows="4">{historia.diagnostico or ''}</textarea>
            </div>
            
            <div class="form-group">
                <label>💊 Tratamiento</label>
                <textarea name="tratamiento" class="form-control" rows="4">{historia.tratamiento or ''}</textarea>
            </div>
            
            <div class="form-group">
                <label>Estado</label>
                <select name="estado" class="form-control">
                    <option value="Abierto" {'selected' if historia.estado == 'Abierto' else ''}>Abierto</option>
                    <option value="En proceso" {'selected' if historia.estado == 'En proceso' else ''}>En proceso</option>
                    <option value="Cerrado" {'selected' if historia.estado == 'Cerrado' else ''}>Cerrado</option>
                </select>
            </div>
            
            <button type="submit" class="btn btn-primary">💾 Guardar Cambios</button>
            <a href="/historia-clinica/{id}" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Editar Historia Clínica")

@app.route('/historia-clinica/eliminar/<int:id>')
@login_required
def eliminar_historia_clinica(id):
    user = db.session.get(Usuario, session.get('user_id'))
    historia = HistoriaClinica.query.get_or_404(id)
    
    if session.get('rol') != 'admin' and historia.usuario_id != user.id:
        flash('❌ No tienes permiso para eliminar esta historia')
        return redirect('/historias-clinicas')
    
    # Eliminar archivos físicos
    for doc in historia.documentos:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], doc.nombre)
        if os.path.exists(filepath):
            os.remove(filepath)
        db.session.delete(doc)
    
    db.session.delete(historia)
    db.session.commit()
    flash('✅ Historia clínica eliminada')
    return redirect('/historias-clinicas')

@app.route('/documento/eliminar/<int:id>')
@login_required
def eliminar_documento(id):
    doc = Documento.query.get_or_404(id)
    historia_id = doc.historia_clinica_id
    
    # Eliminar archivo físico
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], doc.nombre)
    if os.path.exists(filepath):
        os.remove(filepath)
    
    db.session.delete(doc)
    db.session.commit()
    flash('✅ Documento eliminado')
    return redirect(f'/historia-clinica/{historia_id}')

@app.route('/uploads/<filename>')
@login_required
def uploaded_file(filename):
    from flask import send_from_directory
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

# ========== ADMIN: GESTIÓN DE TIPOS DE HISTORIA CLÍNICA ==========

@app.route('/admin/tipos-historia')
@login_required
def admin_tipos_historia():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    tipos = TipoHistoriaClinica.query.all()
    
    tipos_html = ""
    for t in tipos:
        tipos_html += f"""
        <tr>
            <td><span style="background:{t.color}; padding:5px 12px; border-radius:20px; color:white;">{t.nombre}</span></td>
            <td>{t.descripcion[:50]}...</td>
            <td>{t.departamento}</td>
            <td>
                <a href="/admin/tipo-historia/editar/{t.id}" class="btn btn-warning btn-sm">✏️</a>
                <a href="/admin/tipo-historia/eliminar/{t.id}" class="btn btn-danger btn-sm" onclick="return confirm('¿Eliminar tipo?')">🗑️</a>
            </td>
        </tr>
        """
    
    content = f"""
        <h2>⚙️ Configurar Tipos de Historia Clínica</h2>
        <p><a href="/admin/tipo-historia/nuevo" class="btn btn-success">➕ Nuevo Tipo</a></p>
        
        <table>
            <thead>
                <tr><th>Nombre</th><th>Descripción</th><th>Departamento</th><th>Acciones</th></tr>
            </thead>
            <tbody>
                {tipos_html if tipos else '<tr><td colspan="4">No hay tipos configurados</td></tr>'}
            </tbody>
        </table>
        
        <p style="margin-top:20px;"><a href="/historias-clinicas" class="btn btn-primary">← Ir a Historias Clínicas</a></p>
    """
    return base_html(content, "Tipos de Historia Clínica")

@app.route('/admin/tipo-historia/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_tipo_historia():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    if request.method == 'POST':
        tipo = TipoHistoriaClinica(
            nombre=request.form['nombre'],
            descripcion=request.form.get('descripcion', ''),
            departamento=request.form.get('departamento', 'General'),
            color=request.form.get('color', '#3498db'),
            admin_id=session['user_id']
        )
        db.session.add(tipo)
        db.session.commit()
        flash('✅ Tipo de historia clínica creado')
        return redirect('/admin/tipos-historia')
    
    departamentos = db.session.query(Usuario.departamento).distinct().all()
    deptos_options = "".join([f'<option value="{d[0]}">{d[0]}</option>' for d in departamentos])
    
    content = f"""
        <h2>➕ Nuevo Tipo de Historia Clínica</h2>
        <form method="POST" style="max-width: 500px;">
            <div class="form-group">
                <label>Nombre *</label>
                <input type="text" name="nombre" class="form-control" maxlength="100" required placeholder="Ej: Historia Clínica General">
            </div>
            <div class="form-group">
                <label>Descripción</label>
                <textarea name="descripcion" class="form-control" rows="3"></textarea>
            </div>
            <div class="form-group">
                <label>Departamento</label>
                <select name="departamento" class="form-control">
                    <option value="General">General (todos)</option>
                    {deptos_options}
                </select>
            </div>
            <div class="form-group">
                <label>Color</label>
                <select name="color" class="form-control">
                    <option value="#3498db">🔵 Azul</option>
                    <option value="#27ae60">🟢 Verde</option>
                    <option value="#e74c3c">🔴 Rojo</option>
                    <option value="#f39c12">🟡 Naranja</option>
                    <option value="#9b59b6">🟣 Morado</option>
                    <option value="#1abc9c">🟢 Turquesa</option>
                </select>
            </div>
            <button type="submit" class="btn btn-primary">Crear Tipo</button>
            <a href="/admin/tipos-historia" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Nuevo Tipo Historia")

# ========== CHAT INTERNO ==========

@app.route('/chat')
@login_required
def chat():
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') == 'admin':
        # Admin ve lista de médicos
        medicos = Usuario.query.filter(Usuario.id != user.id).all()
        
        medicos_html = ""
        for m in medicos:
            no_leidos = Mensaje.query.filter_by(
                emisor_id=m.id, 
                receptor_id=user.id, 
                leido=False
            ).count()
            
            badge = f'<span style="background:#e74c3c; color:white; padding:2px 8px; border-radius:10px;">{no_leidos}</span>' if no_leidos > 0 else ''
            
            medicos_html += f"""
            <div class="chat-sidebar-item" onclick="window.location.href='/chat/{m.id}'">
                👨‍⚕️ Dr/a. {m.nombre_completo} {badge}
                <small style="display:block;color:#666;">{m.departamento}</small>
            </div>
            """
        
        content = f"""
            <h2>💬 Chat con el Equipo Médico</h2>
            <div class="chat-container">
                <div class="chat-sidebar">
                    {medicos_html if medicos else '<p style="padding:15px;">No hay médicos registrados</p>'}
                </div>
                <div class="chat-main">
                    <div class="chat-header">👈 Selecciona un médico para chatear</div>
                    <div class="chat-messages" style="display:flex;align-items:center;justify-content:center;color:#999;">
                        Elige un contacto para comenzar
                    </div>
                </div>
            </div>
        """
    else:
        # Médico: chat directo con admin
        admin = Usuario.query.filter_by(rol='admin').first()
        if admin:
            return redirect(f'/chat/{admin.id}')
        else:
            flash('❌ No hay administrador registrado')
            return redirect('/dashboard')
    
    return base_html(content, "Chat")

@app.route('/chat/<int:otro_id>', methods=['GET', 'POST'])
@login_required
def chat_con(otro_id):
    user = db.session.get(Usuario, session.get('user_id'))
    otro = Usuario.query.get_or_404(otro_id)
    
    # Verificar permisos
    if session.get('rol') != 'admin' and otro.rol != 'admin':
        flash('❌ No tienes permiso')
        return redirect('/chat')
    
    if request.method == 'POST':
        contenido = request.form.get('mensaje')
        if contenido:
            msg = Mensaje(
                contenido=contenido,
                emisor_id=user.id,
                receptor_id=otro.id
            )
            db.session.add(msg)
            db.session.commit()
        return redirect(f'/chat/{otro_id}')
    
    # Marcar mensajes como leídos
    Mensaje.query.filter_by(
        emisor_id=otro.id,
        receptor_id=user.id,
        leido=False
    ).update({'leido': True})
    db.session.commit()
    
    # Obtener conversación
    mensajes = Mensaje.query.filter(
        ((Mensaje.emisor_id == user.id) & (Mensaje.receptor_id == otro.id)) |
        ((Mensaje.emisor_id == otro.id) & (Mensaje.receptor_id == user.id))
    ).order_by(Mensaje.id).all()
    
    mensajes_html = ""
    for m in mensajes:
        sent_class = 'sent' if m.emisor_id == user.id else ''
        mensajes_html += f"""
        <div class="chat-message {sent_class}">
            <div class="chat-message-bubble">
                {m.contenido}
                <div class="chat-message-time">{m.fecha}</div>
            </div>
        </div>
        """
    
    # Sidebar para admin
    sidebar_html = ""
    if session.get('rol') == 'admin':
        medicos = Usuario.query.filter(Usuario.id != user.id).all()
        for m in medicos:
            no_leidos = Mensaje.query.filter_by(
                emisor_id=m.id,
                receptor_id=user.id,
                leido=False
            ).count()
            badge = f'<span style="background:#e74c3c; color:white; padding:2px 8px; border-radius:10px;">{no_leidos}</span>' if no_leidos > 0 else ''
            active_class = 'active' if m.id == otro_id else ''
            sidebar_html += f"""
            <div class="chat-sidebar-item {active_class}" onclick="window.location.href='/chat/{m.id}'">
                👨‍⚕️ Dr/a. {m.nombre_completo} {badge}
            </div>
            """
    
    content = f"""
        <h2>💬 Chat con Dr/a. {otro.nombre_completo}</h2>
        <div class="chat-container">
            {f'<div class="chat-sidebar">{sidebar_html}</div>' if session.get("rol") == "admin" else ''}
            <div class="chat-main">
                <div class="chat-header">👨‍⚕️ Dr/a. {otro.nombre_completo} ({otro.departamento})</div>
                <div class="chat-messages" id="chat-messages">
                    {mensajes_html if mensajes else '<p style="text-align:center;color:#999;padding:20px;">No hay mensajes. ¡Empieza la conversación!</p>'}
                </div>
                <form method="POST" class="chat-input">
                    <input type="text" name="mensaje" placeholder="Escribe un mensaje..." required autocomplete="off">
                    <button type="submit" class="btn btn-primary">📤 Enviar</button>
                </form>
            </div>
        </div>
        
        <script>
            const messagesDiv = document.getElementById('chat-messages');
            if (messagesDiv) messagesDiv.scrollTop = messagesDiv.scrollHeight;
        </script>
    """
    return base_html(content, f"Chat - {otro.nombre_completo}")

# ========== CHAT GRUPAL POR DEPARTAMENTOS ==========

@app.route('/chat-grupal')
@login_required
def chat_grupal():
    user = db.session.get(Usuario, session.get('user_id'))
    
    # Determinar departamentos disponibles
    if session.get('rol') == 'admin':
        departamentos = db.session.query(Usuario.departamento).distinct().all()
        deptos_disponibles = [d[0] for d in departamentos]
    else:
        deptos_disponibles = [user.departamento] if user.departamento else ['General']
    
    # Sidebar con departamentos
    sidebar_html = ""
    for depto in deptos_disponibles:
        sidebar_html += f"""
        <div class="chat-sidebar-item" onclick="window.location.href='/chat-grupal/{depto}'">
            🏥 {depto}
        </div>
        """
    
    content = f"""
        <h2>👥 Chat por Especialidades</h2>
        <div class="chat-container">
            <div class="chat-sidebar">
                {sidebar_html if sidebar_html else '<p style="padding:15px;">No hay departamentos</p>'}
            </div>
            <div class="chat-main">
                <div class="chat-header">👈 Selecciona una especialidad</div>
                <div class="chat-messages" style="display:flex;align-items:center;justify-content:center;color:#999;">
                    Elige un departamento para ver la conversación
                </div>
            </div>
        </div>
    """
    return base_html(content, "Chat Grupal")

@app.route('/chat-grupal/<depto>', methods=['GET', 'POST'])
@login_required
def chat_grupal_depto(depto):
    user = db.session.get(Usuario, session.get('user_id'))
    
    # Verificar permisos
    if session.get('rol') != 'admin' and user.departamento != depto:
        flash('❌ No tienes acceso a este departamento')
        return redirect('/chat-grupal')
    
    if request.method == 'POST':
        contenido = request.form.get('mensaje')
        if contenido:
            msg = MensajeGrupal(
                contenido=contenido,
                departamento=depto,
                usuario_id=user.id
            )
            db.session.add(msg)
            db.session.commit()
        return redirect(f'/chat-grupal/{depto}')
    
    # Obtener mensajes del departamento
    mensajes = MensajeGrupal.query.filter_by(departamento=depto).order_by(MensajeGrupal.id.desc()).limit(50).all()
    mensajes = list(reversed(mensajes))
    
    mensajes_html = ""
    for m in mensajes:
        es_mio = m.usuario_id == user.id
        mensajes_html += f"""
        <div class="chat-message {'sent' if es_mio else ''}">
            <div class="chat-message-bubble">
                <strong>{m.usuario.nombre_completo}</strong><br>
                {m.contenido}
                <div class="chat-message-time">{m.fecha}</div>
            </div>
        </div>
        """
    
    # Sidebar
    if session.get('rol') == 'admin':
        departamentos = db.session.query(Usuario.departamento).distinct().all()
        deptos_disponibles = [d[0] for d in departamentos]
    else:
        deptos_disponibles = [user.departamento]
    
    sidebar_html = ""
    for d in deptos_disponibles:
        active_class = 'active' if d == depto else ''
        sidebar_html += f"""
        <div class="chat-sidebar-item {active_class}" onclick="window.location.href='/chat-grupal/{d}'">
            🏥 {d}
        </div>
        """
    
    content = f"""
        <h2>👥 Chat de {depto}</h2>
        <div class="chat-container">
            <div class="chat-sidebar">
                {sidebar_html}
            </div>
            <div class="chat-main">
                <div class="chat-header">
                    🏥 {depto} | 👥 {Usuario.query.filter_by(departamento=depto).count()} miembros
                </div>
                <div class="chat-messages" id="chat-messages">
                    {mensajes_html if mensajes else '<p style="text-align:center;color:#999;padding:20px;">No hay mensajes. ¡Empieza la conversación!</p>'}
                </div>
                <form method="POST" class="chat-input">
                    <input type="text" name="mensaje" placeholder="Escribe un mensaje para {depto}..." required autocomplete="off">
                    <button type="submit" class="btn btn-primary">📤 Enviar</button>
                </form>
            </div>
        </div>
        
        <script>
            const messagesDiv = document.getElementById('chat-messages');
            if (messagesDiv) messagesDiv.scrollTop = messagesDiv.scrollHeight;
        </script>
    """
    return base_html(content, f"Chat - {depto}")

# ========== CONTROL HORARIO ==========

@app.route('/control-horario')
@login_required
def control_horario():
    user = db.session.get(Usuario, session['user_id'])
    hoy = datetime.now().strftime('%d/%m/%Y')
    
    # Obtener fichajes de hoy
    fichajes_hoy = Fichaje.query.filter_by(
        usuario_id=user.id, 
        fecha=hoy
    ).order_by(Fichaje.fecha_hora).all()
    
    # Determinar estado actual
    estado = 'fuera'
    if fichajes_hoy:
        ultimo = fichajes_hoy[-1]
        if ultimo.tipo == 'entrada' or ultimo.tipo == 'pausa_fin':
            estado = 'trabajando'
        elif ultimo.tipo == 'pausa_inicio':
            estado = 'pausa'
        elif ultimo.tipo == 'salida':
            estado = 'finalizado'
    
    # Calcular tiempo trabajado hoy
    tiempo_trabajado = timedelta()
    entrada_time = None
    for f in fichajes_hoy:
        if f.tipo == 'entrada':
            entrada_time = datetime.strptime(f.fecha_hora, '%H:%M:%S')
        elif f.tipo == 'salida' and entrada_time:
            salida_time = datetime.strptime(f.fecha_hora, '%H:%M:%S')
            tiempo_trabajado += (salida_time - entrada_time)
            entrada_time = None
        elif f.tipo == 'pausa_inicio' and entrada_time:
            pausa_inicio = datetime.strptime(f.fecha_hora, '%H:%M:%S')
            tiempo_trabajado += (pausa_inicio - entrada_time)
            entrada_time = None
        elif f.tipo == 'pausa_fin':
            entrada_time = datetime.strptime(f.fecha_hora, '%H:%M:%S')
    
    if entrada_time:
        tiempo_trabajado += (datetime.now() - entrada_time)
    
    horas = tiempo_trabajado.seconds // 3600
    minutos = (tiempo_trabajado.seconds % 3600) // 60
    segundos = tiempo_trabajado.seconds % 60
    
    content = f"""
        <h2>⏰ Control Horario</h2>
        
        <div class="reloj-container">
            <div class="reloj-estado" id="estado-texto">
                {'🟢 En consulta' if estado == 'trabajando' else '🟡 En pausa' if estado == 'pausa' else '⚪ Jornada finalizada' if estado == 'finalizado' else '🔴 Sin fichar'}
            </div>
            <div class="reloj-tiempo" id="tiempo-trabajado">
                {horas:02d}:{minutos:02d}:{segundos:02d}
            </div>
            <div class="fichaje-buttons">
                <form method="POST" action="/fichar/entrada" style="display:inline;">
                    <button type="submit" class="btn-fichaje btn-entrada" {'disabled' if estado != 'fuera' else ''}>
                        🟢 Entrada
                    </button>
                </form>
                <form method="POST" action="/fichar/pausa" style="display:inline;">
                    <button type="submit" class="btn-fichaje btn-pausa" {'disabled' if estado not in ['trabajando', 'pausa'] else ''}>
                        {'☕ Iniciar Pausa' if estado == 'trabajando' else '▶️ Finalizar Pausa'}
                    </button>
                </form>
                <form method="POST" action="/fichar/salida" style="display:inline;">
                    <button type="submit" class="btn-fichaje btn-salida" {'disabled' if estado not in ['trabajando', 'pausa'] else ''}>
                        🔴 Salida
                    </button>
                </form>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; margin: 20px 0;">
            <div style="background: #f8f9fa; padding: 15px; border-radius: 12px; text-align: center;">
                <div style="font-size: 24px;">🕐</div>
                <div style="font-weight: bold;">Entrada</div>
                <div>{next((f.fecha_hora[:5] for f in fichajes_hoy if f.tipo == 'entrada'), '--:--')}</div>
            </div>
            <div style="background: #f8f9fa; padding: 15px; border-radius: 12px; text-align: center;">
                <div style="font-size: 24px;">☕</div>
                <div style="font-weight: bold;">Pausa</div>
                <div>{'30 min' if any(f.tipo == 'pausa_inicio' for f in fichajes_hoy) else '--'}</div>
            </div>
            <div style="background: #f8f9fa; padding: 15px; border-radius: 12px; text-align: center;">
                <div style="font-size: 24px;">🏁</div>
                <div style="font-weight: bold;">Salida</div>
                <div>{next((f.fecha_hora[:5] for f in fichajes_hoy if f.tipo == 'salida'), '--:--')}</div>
            </div>
        </div>
        
        <h3>📋 Registro de hoy</h3>
        <table>
            <thead><tr><th>Hora</th><th>Tipo</th></tr></thead>
            <tbody>
                {''.join([f'<tr><td>{f.fecha_hora[:5]}</td><td>{f.tipo.replace("_", " ").title()}</td></tr>' for f in fichajes_hoy])}
            </tbody>
        </table>
        
        <p style="margin-top:20px;">
            <a href="/control-horario/historial" class="btn btn-primary">📊 Ver historial completo</a>
        </p>
        
        <script>
            let segundos = {tiempo_trabajado.seconds};
            const tiempoElement = document.getElementById('tiempo-trabajado');
            const estado = '{estado}';
            
            if (estado === 'trabajando') {{
                setInterval(() => {{
                    segundos++;
                    const h = Math.floor(segundos / 3600);
                    const m = Math.floor((segundos % 3600) / 60);
                    const s = segundos % 60;
                    tiempoElement.textContent = `${{h.toString().padStart(2, '0')}}:${{m.toString().padStart(2, '0')}}:${{s.toString().padStart(2, '0')}}`;
                }}, 1000);
            }}
        </script>
    """
    return base_html(content, "Control Horario")

@app.route('/fichar/<tipo>', methods=['POST'])
@login_required
def fichar(tipo):
    user = db.session.get(Usuario, session['user_id'])
    ahora = datetime.now()
    hoy = ahora.strftime('%d/%m/%Y')
    
    if tipo == 'entrada':
        existe = Fichaje.query.filter_by(usuario_id=user.id, fecha=hoy, tipo='entrada').first()
        if not existe:
            fichaje = Fichaje(
                usuario_id=user.id,
                tipo='entrada',
                fecha_hora=ahora.strftime('%H:%M:%S'),
                fecha=hoy
            )
            db.session.add(fichaje)
            db.session.commit()
            flash('✅ Entrada registrada a las ' + ahora.strftime('%H:%M'))
    
    elif tipo == 'salida':
        ultimo = Fichaje.query.filter_by(usuario_id=user.id, fecha=hoy).order_by(Fichaje.id.desc()).first()
        if ultimo and ultimo.tipo in ['entrada', 'pausa_fin']:
            fichaje = Fichaje(
                usuario_id=user.id,
                tipo='salida',
                fecha_hora=ahora.strftime('%H:%M:%S'),
                fecha=hoy
            )
            db.session.add(fichaje)
            db.session.commit()
            flash('🔴 Salida registrada a las ' + ahora.strftime('%H:%M'))
    
    elif tipo == 'pausa':
        ultimo = Fichaje.query.filter_by(usuario_id=user.id, fecha=hoy).order_by(Fichaje.id.desc()).first()
        if ultimo:
            nuevo_tipo = 'pausa_inicio' if ultimo.tipo in ['entrada', 'pausa_fin'] else 'pausa_fin'
            fichaje = Fichaje(
                usuario_id=user.id,
                tipo=nuevo_tipo,
                fecha_hora=ahora.strftime('%H:%M:%S'),
                fecha=hoy
            )
            db.session.add(fichaje)
            db.session.commit()
            flash('☕ ' + ('Pausa iniciada' if nuevo_tipo == 'pausa_inicio' else 'Pausa finalizada'))
    
    return redirect('/control-horario')

@app.route('/control-horario/historial')
@login_required
def historial_fichajes():
    user = db.session.get(Usuario, session['user_id'])
    empleado_id = request.args.get('empleado_id', type=int)
    
    if session.get('rol') == 'admin':
        if empleado_id:
            user = db.session.get(Usuario, empleado_id)
            fichajes = Fichaje.query.filter_by(usuario_id=empleado_id).order_by(Fichaje.fecha.desc(), Fichaje.fecha_hora.desc()).all() if user else []
        else:
            fichajes = Fichaje.query.order_by(Fichaje.fecha.desc(), Fichaje.fecha_hora.desc()).all()
            medicos = Usuario.query.filter_by(rol='medico').all()
    else:
        fichajes = Fichaje.query.filter_by(usuario_id=user.id).order_by(Fichaje.fecha.desc(), Fichaje.fecha_hora.desc()).all()
        medicos = None
    
    # Agrupar por fecha
    fichajes_por_dia = {}
    for f in fichajes:
        if f.fecha not in fichajes_por_dia:
            fichajes_por_dia[f.fecha] = []
        fichajes_por_dia[f.fecha].append(f)
    
    # Calcular horas por día
    horas_por_dia = {}
    for fecha, fichas in fichajes_por_dia.items():
        tiempo = timedelta()
        entrada = None
        for f in fichas:
            if f.tipo == 'entrada':
                entrada = datetime.strptime(f.fecha_hora, '%H:%M:%S')
            elif f.tipo == 'salida' and entrada:
                salida = datetime.strptime(f.fecha_hora, '%H:%M:%S')
                tiempo += (salida - entrada)
                entrada = None
            elif f.tipo == 'pausa_inicio' and entrada:
                pausa = datetime.strptime(f.fecha_hora, '%H:%M:%S')
                tiempo += (pausa - entrada)
                entrada = None
            elif f.tipo == 'pausa_fin':
                entrada = datetime.strptime(f.fecha_hora, '%H:%M:%S')
        horas_por_dia[fecha] = tiempo
    
    historial_html = ""
    for fecha, fichas in fichajes_por_dia.items():
        tiempo_dia = horas_por_dia[fecha]
        horas = tiempo_dia.seconds // 3600
        minutos = (tiempo_dia.seconds % 3600) // 60
        
        fichas_html = ""
        for f in fichas:
            icono = {'entrada': '🟢', 'salida': '🔴', 'pausa_inicio': '☕', 'pausa_fin': '▶️'}.get(f.tipo, '📌')
            fichas_html += f"<div>{icono} {f.fecha_hora[:5]} - {f.tipo.replace('_', ' ').title()}</div>"
        
        historial_html += f"""
        <div style="background: white; border-radius: 12px; padding: 20px; margin-bottom: 15px; box-shadow: 0 2px 8px rgba(0,0,0,0.1);">
            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px;">
                <h3 style="margin:0;">📅 {fecha}</h3>
                <div style="font-size: 20px; font-weight: bold; color: #1a5276;">{horas:02d}:{minutos:02d} h</div>
            </div>
            <div style="display: flex; flex-wrap: wrap; gap: 15px 30px;">
                {fichas_html}
            </div>
        </div>
        """
    
    selector_html = ""
    if session.get('rol') == 'admin' and medicos:
        selector_html = f"""
        <div style="margin-bottom: 20px;">
            <form method="GET" style="display: flex; gap: 10px;">
                <select name="empleado_id" class="form-control" style="width: auto;">
                    <option value="">Todos los médicos</option>
                    {''.join([f'<option value="{m.id}" {"selected" if empleado_id == m.id else ""}>{m.nombre_completo}</option>' for m in medicos])}
                </select>
                <button type="submit" class="btn btn-primary">Filtrar</button>
            </form>
        </div>
        """
    
    content = f"""
        <h2>📊 Historial de Fichajes</h2>
        {selector_html}
        
        <div style="margin-top: 20px;">
            {historial_html if historial_html else '<p style="text-align:center; color:#999;">No hay fichajes registrados</p>'}
        </div>
        
        <p style="margin-top:20px;">
            <a href="/control-horario" class="btn btn-primary">← Volver al control horario</a>
        </p>
    """
    return base_html(content, "Historial de Fichajes")

# ========== EXPORTACIONES ==========

@app.route('/admin/exportar-todo')
@login_required
def exportar_todo():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    wb = Workbook()
    
    # Hoja 1: Médicos
    ws1 = wb.active
    ws1.title = "Médicos"
    headers1 = ['ID', 'Usuario', 'Nombre', 'Especialidad', 'Cargo', 'Citas Totales', 'Realizadas', 'Productividad']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    
    medicos = Usuario.query.filter_by(rol='medico').all()
    for row, m in enumerate(medicos, 2):
        total = len(m.citas)
        completadas = sum(1 for c in m.citas if c.asistida)
        prod = f"{(completadas/total*100):.1f}%" if total > 0 else "0%"
        ws1.cell(row=row, column=1, value=m.id)
        ws1.cell(row=row, column=2, value=m.username)
        ws1.cell(row=row, column=3, value=m.nombre_completo)
        ws1.cell(row=row, column=4, value=m.departamento)
        ws1.cell(row=row, column=5, value=m.cargo)
        ws1.cell(row=row, column=6, value=total)
        ws1.cell(row=row, column=7, value=completadas)
        ws1.cell(row=row, column=8, value=prod)
    
    # Hoja 2: Pacientes
    ws2 = wb.create_sheet("Pacientes")
    headers2 = ['ID', 'Nombre', 'Email', 'Teléfono', 'Obra Social', 'Fecha Nac.', 'Grupo Sanguíneo', 'Médico', 'Fecha Registro']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    
    pacientes = Paciente.query.all()
    for row, p in enumerate(pacientes, 2):
        ws2.cell(row=row, column=1, value=p.id)
        ws2.cell(row=row, column=2, value=p.nombre)
        ws2.cell(row=row, column=3, value=p.email or '')
        ws2.cell(row=row, column=4, value=p.telefono or '')
        ws2.cell(row=row, column=5, value=p.obra_social or '')
        ws2.cell(row=row, column=6, value=p.fecha_nacimiento or '')
        ws2.cell(row=row, column=7, value=p.grupo_sanguineo or '')
        ws2.cell(row=row, column=8, value=p.usuario.nombre_completo)
        ws2.cell(row=row, column=9, value=p.fecha_creacion)
    
    # Hoja 3: Citas
    ws3 = wb.create_sheet("Citas")
    headers3 = ['ID', 'Título', 'Fecha', 'Hora', 'Tipo', 'Paciente', 'Médico', 'Estado', 'Prioridad']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a5276", end_color="1a5276", fill_type="solid")
    
    citas = Cita.query.all()
    for row, c in enumerate(citas, 2):
        ws3.cell(row=row, column=1, value=c.id)
        ws3.cell(row=row, column=2, value=c.titulo)
        ws3.cell(row=row, column=3, value=c.fecha_cita or '')
        ws3.cell(row=row, column=4, value=f"{c.hora_inicio or ''} - {c.hora_fin or ''}")
        ws3.cell(row=row, column=5, value=c.tipo_consulta)
        ws3.cell(row=row, column=6, value=c.paciente.nombre if c.paciente else '')
        ws3.cell(row=row, column=7, value=c.medico.nombre_completo if c.medico else '')
        ws3.cell(row=row, column=8, value='Realizada' if c.asistida else 'Pendiente')
        ws3.cell(row=row, column=9, value=c.prioridad)
    
    # Ajustar ancho
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'Clinica_Medica_Reporte_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
    )

# ========== GESTIÓN DE TAREAS PERSONALES ==========

@app.route('/tarea-personal/nueva', methods=['GET', 'POST'])
@login_required
def nueva_tarea_personal():
    if request.method == 'POST':
        tarea = TareaPersonal(
            titulo=request.form['titulo'],
            descripcion=request.form.get('descripcion', ''),
            fecha_limite=request.form.get('fecha_limite', ''),
            prioridad=request.form.get('prioridad', 'media'),
            usuario_id=session['user_id']
        )
        db.session.add(tarea)
        db.session.commit()
        flash('✅ Tarea personal creada')
        return redirect('/mi-espacio')
    
    content = """
        <h2>✅ Nueva Tarea Personal</h2>
        <form method="POST" style="max-width: 500px;">
            <div class="form-group">
                <label>Título *</label>
                <input type="text" name="titulo" class="form-control" required>
            </div>
            <div class="form-group">
                <label>Descripción</label>
                <textarea name="descripcion" class="form-control" rows="3"></textarea>
            </div>
            <div class="form-group">
                <label>Fecha límite</label>
                <input type="date" name="fecha_limite" class="form-control">
            </div>
            <div class="form-group">
                <label>Prioridad</label>
                <select name="prioridad" class="form-control">
                    <option value="alta">🔴 Alta</option>
                    <option value="media" selected>🟡 Media</option>
                    <option value="baja">🟢 Baja</option>
                </select>
            </div>
            <button type="submit" class="btn btn-primary">Crear Tarea</button>
            <a href="/mi-espacio" class="btn" style="background: #95a5a6; color: white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Nueva Tarea Personal")

@app.route('/tarea-personal/completar/<int:id>')
@login_required
def completar_tarea_personal(id):
    tarea = TareaPersonal.query.get_or_404(id)
    if tarea.usuario_id == session['user_id']:
        tarea.completada = not tarea.completada
        db.session.commit()
        flash('✅ Estado de tarea actualizado')
    return redirect('/mi-espacio')

@app.route('/tarea-personal/eliminar/<int:id>')
@login_required
def eliminar_tarea_personal(id):
    tarea = TareaPersonal.query.get_or_404(id)
    if tarea.usuario_id == session['user_id']:
        db.session.delete(tarea)
        db.session.commit()
        flash('✅ Tarea eliminada')
    return redirect('/mi-espacio')


# ========== ADMIN: GESTIÓN DE MÉDICOS ==========

@app.route('/admin/medicos')
@login_required
def admin_medicos():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    medicos = Usuario.query.filter_by(rol='medico').all()
    
    filas = ""
    for m in medicos:
        citas_totales = len(m.citas)
        citas_realizadas = sum(1 for c in m.citas if c.asistida)
        productividad = int((citas_realizadas / citas_totales * 100)) if citas_totales > 0 else 0
        pacientes_asignados = Paciente.query.filter_by(usuario_id=m.id).count()
        
        filas += f"""
        <tr>
            <td>{m.nombre_completo}</td>
            <td>{m.username}</td>
            <td>{m.departamento}</td>
            <td>{m.cargo}</td>
            <td><span class="badge badge-medico">MÉDICO</span></td>
            <td>{citas_totales} ({citas_realizadas} ✅)</td>
            <td>{pacientes_asignados}</td>
            <td>
                <div style="background: #ddd; height: 6px; width: 100px; border-radius: 3px;">
                    <div style="background: #27ae60; height: 6px; width: {productividad}px; border-radius: 3px;"></div>
                </div>
                {productividad}%
            </td>
            <td>
                <a href="/admin/medico/editar/{m.id}" class="btn btn-warning btn-sm">✏️</a>
                <a href="/admin/medico/eliminar/{m.id}" class="btn btn-danger btn-sm" onclick="return confirm('¿Eliminar a {m.nombre_completo}?')">🗑️</a>
            </td>
        </tr>
        """
    
    content = f"""
        <h2>👨‍⚕️ Gestión de Médicos</h2>
        <p><a href="/registro" class="btn btn-success">➕ Nuevo Médico</a></p>
        
        <table>
            <thead>
                <tr>
                    <th>Nombre</th>
                    <th>Usuario</th>
                    <th>Especialidad</th>
                    <th>Cargo</th>
                    <th>Rol</th>
                    <th>Citas</th>
                    <th>Pacientes</th>
                    <th>Productividad</th>
                    <th>Acciones</th>
                </tr>
            </thead>
            <tbody>
                {filas if filas else '<tr><td colspan="9" style="text-align:center;">No hay médicos registrados</td></tr>'}
            </tbody>
        </table>
    """
    return base_html(content, "Gestión de Médicos")

@app.route('/admin/medico/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_medico(id):
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    medico = Usuario.query.get_or_404(id)
    
    if request.method == 'POST':
        medico.nombre_completo = request.form['nombre_completo']
        medico.departamento = request.form['departamento']
        medico.cargo = request.form['cargo']
        if request.form.get('password'):
            medico.password_hash = generate_password_hash(request.form['password'], method='pbkdf2:sha256')
        db.session.commit()
        flash('✅ Médico actualizado correctamente')
        return redirect('/admin/medicos')
    
    especialidades = ['Medicina General', 'Cardiología', 'Pediatría', 'Traumatología', 'Ginecología', 'Dermatología', 'Oftalmología', 'Psiquiatría', 'Administración']
    
    content = f"""
        <h2>✏️ Editar Médico: {medico.nombre_completo}</h2>
        <form method="POST" style="max-width: 500px;">
            <div class="form-group">
                <label>Nombre Completo</label>
                <input type="text" name="nombre_completo" class="form-control" value="{medico.nombre_completo}" required>
            </div>
            <div class="form-group">
                <label>Especialidad / Departamento</label>
                <select name="departamento" class="form-control">
                    {''.join([f'<option value="{e}" {"selected" if e == medico.departamento else ""}>{e}</option>' for e in especialidades])}
                </select>
            </div>
            <div class="form-group">
                <label>Cargo</label>
                <input type="text" name="cargo" class="form-control" value="{medico.cargo}">
            </div>
            <div class="form-group">
                <label>Nueva Contraseña (dejar en blanco para no cambiar)</label>
                <input type="password" name="password" class="form-control">
            </div>
            <button type="submit" class="btn btn-primary">Guardar Cambios</button>
            <a href="/admin/medicos" class="btn" style="background:#95a5a6;color:white;">Cancelar</a>
        </form>
    """
    return base_html(content, "Editar Médico")

@app.route('/admin/medico/eliminar/<int:id>')
@login_required
def eliminar_medico(id):
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    medico = db.session.get(Usuario, id)
    if not medico:
        flash('❌ Médico no encontrado')
        return redirect('/admin/medicos')
    
    if medico.rol == 'admin':
        flash('❌ No se puede eliminar al administrador')
        return redirect('/admin/medicos')
    
    # Eliminar datos asociados
    Cita.query.filter_by(usuario_id=id).delete()
    Paciente.query.filter_by(usuario_id=id).update({'usuario_id': session['user_id']})
    Mensaje.query.filter((Mensaje.emisor_id == id) | (Mensaje.receptor_id == id)).delete()
    MensajeGrupal.query.filter_by(usuario_id=id).delete()
    Fichaje.query.filter_by(usuario_id=id).delete()
    TareaPersonal.query.filter_by(usuario_id=id).delete()
    NotaPersonal.query.filter_by(usuario_id=id).delete()
    HistoriaClinica.query.filter_by(usuario_id=id).update({'usuario_id': session['user_id']})
    Comentario.query.filter_by(usuario_id=id).delete()
    
    db.session.delete(medico)
    db.session.commit()
    
    flash(f'✅ Médico {medico.nombre_completo} eliminado')
    return redirect('/admin/medicos')

# ========== ADMIN: PANEL HORARIO ==========

@app.route('/admin/panel-horario')
@login_required
def admin_panel_horario():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    hoy = datetime.now().strftime('%d/%m/%Y')
    medicos = Usuario.query.filter_by(rol='medico').all()
    
    medicos_html = ""
    for m in medicos:
        fichajes_hoy = Fichaje.query.filter_by(
            usuario_id=m.id, 
            fecha=hoy
        ).order_by(Fichaje.fecha_hora).all()
        
        # Determinar estado
        estado = '🔴 Ausente'
        estado_class = 'ausente'
        if fichajes_hoy:
            ultimo = fichajes_hoy[-1]
            if ultimo.tipo == 'entrada' or ultimo.tipo == 'pausa_fin':
                estado = '🟢 En consulta'
                estado_class = 'trabajando'
            elif ultimo.tipo == 'pausa_inicio':
                estado = '☕ En pausa'
                estado_class = 'pausa'
            elif ultimo.tipo == 'salida':
                estado = '⚪ Finalizado'
                estado_class = 'finalizado'
        
        # Calcular tiempo trabajado
        tiempo = timedelta()
        entrada = None
        for f in fichajes_hoy:
            if f.tipo == 'entrada':
                entrada = datetime.strptime(f.fecha_hora, '%H:%M:%S')
            elif f.tipo == 'salida' and entrada:
                salida = datetime.strptime(f.fecha_hora, '%H:%M:%S')
                tiempo += (salida - entrada)
                entrada = None
            elif f.tipo == 'pausa_inicio' and entrada:
                pausa = datetime.strptime(f.fecha_hora, '%H:%M:%S')
                tiempo += (pausa - entrada)
                entrada = None
            elif f.tipo == 'pausa_fin':
                entrada = datetime.strptime(f.fecha_hora, '%H:%M:%S')
        
        if entrada:
            tiempo += (datetime.now() - entrada)
        
        horas = tiempo.seconds // 3600
        minutos = (tiempo.seconds % 3600) // 60
        
        hora_entrada = next((f.fecha_hora[:5] for f in fichajes_hoy if f.tipo == 'entrada'), '--:--')
        
        # Citas de hoy
        citas_hoy = Cita.query.filter_by(usuario_id=m.id, fecha_cita=datetime.now().strftime('%Y-%m-%d')).count()
        
        medicos_html += f"""
        <div style="background: white; border-radius: 12px; padding: 20px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-left: 4px solid {'#27ae60' if estado_class == 'trabajando' else '#f39c12' if estado_class == 'pausa' else '#95a5a6'};">
            <div style="display: flex; align-items: center; gap: 10px;">
                <div style="width: 40px; height: 40px; background: #1a5276; color: white; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-weight: bold;">
                    👨‍⚕️
                </div>
                <div style="flex:1;">
                    <strong>Dr/a. {m.nombre_completo}</strong>
                    <div style="font-size: 12px; color: #666;">{m.departamento}</div>
                </div>
            </div>
            <div style="margin-top: 15px; text-align: center;">
                <div style="font-size: 24px; font-weight: bold;">{horas:02d}:{minutos:02d}</div>
                <div style="font-size: 12px; color: #666;">Horas hoy</div>
            </div>
            <div style="margin-top: 10px; display: flex; justify-content: space-between;">
                <span>🚪 Entrada: {hora_entrada}</span>
                <span>{estado}</span>
            </div>
            <div style="margin-top: 10px; font-size: 13px; color: #666;">
                📅 Citas hoy: {citas_hoy}
            </div>
            <div style="margin-top: 10px;">
                <a href="/control-horario/historial?empleado_id={m.id}" class="btn btn-primary btn-sm">📊 Ver historial</a>
            </div>
        </div>
        """
    
    content = f"""
        <h2>👑 Panel de Control Horario</h2>
        <p style="color:#666; margin-bottom:20px;">📅 {hoy} - Visión general del equipo médico</p>
        
        <div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 20px;">
            {medicos_html if medicos_html else '<p>No hay médicos registrados</p>'}
        </div>
    """
    return base_html(content, "Panel Horario")

# ========== INFORME MENSUAL PDF ==========

@app.route('/admin/informe-mensual')
@login_required
def informe_mensual():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    hoy = datetime.now()
    nombre_mes = hoy.strftime('%B %Y').capitalize()
    inicio_mes = hoy.replace(day=1).strftime('%Y-%m-%d')
    
    # Citas del mes
    citas_mes = Cita.query.filter(Cita.fecha_cita >= inicio_mes).all()
    
    # Agrupar por especialidad
    deptos_data = {}
    for c in citas_mes:
        depto = c.medico.departamento if c.medico else 'General'
        if depto not in deptos_data:
            deptos_data[depto] = {'total': 0, 'realizadas': 0, 'pendientes': 0, 'alta': 0, 'media': 0, 'baja': 0}
        deptos_data[depto]['total'] += 1
        if c.asistida:
            deptos_data[depto]['realizadas'] += 1
        else:
            deptos_data[depto]['pendientes'] += 1
        deptos_data[depto][c.prioridad] += 1
    
    # Estadísticas generales
    total_medicos = Usuario.query.filter_by(rol='medico').count()
    total_pacientes = Paciente.query.filter(Paciente.fecha_creacion >= inicio_mes).count()
    total_historias = HistoriaClinica.query.filter(HistoriaClinica.fecha_creacion >= inicio_mes).count()
    
    # Crear PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Portada
    title_style = ParagraphStyle(
        'MainTitle',
        parent=styles['Title'],
        fontSize=28,
        textColor=colors.HexColor('#1a5276'),
        alignment=1,
        spaceAfter=40
    )
    elements.append(Paragraph("CLÍNICA MÉDICA", title_style))
    
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=18,
        textColor=colors.HexColor('#2c3e50'),
        alignment=1,
        spaceAfter=20
    )
    elements.append(Paragraph("Informe Mensual de Gestión", subtitle_style))
    
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("<hr width='50%' color='#1a5276' size='2'>", styles['Normal']))
    elements.append(Spacer(1, 1*cm))
    
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=14,
        alignment=1,
        spaceAfter=10
    )
    elements.append(Paragraph(f"<b>Período:</b> {nombre_mes}", info_style))
    elements.append(Paragraph(f"<b>Fecha de emisión:</b> {hoy.strftime('%d de %B de %Y')}", info_style))
    elements.append(Spacer(1, 2*cm))
    
    # Resumen ejecutivo
    elements.append(Paragraph("<b>Resumen Ejecutivo</b>", styles['Heading2']))
    elements.append(Spacer(1, 0.5*cm))
    
    productividad_general = int((sum(d['realizadas'] for d in deptos_data.values()) / len(citas_mes) * 100)) if citas_mes else 0
    
    resumen_data = [
        ['Indicador', 'Valor'],
        ['Total Médicos', str(total_medicos)],
        ['Citas del Mes', str(len(citas_mes))],
        ['Pacientes Nuevos', str(total_pacientes)],
        ['Historias Clínicas Creadas', str(total_historias)],
        ['Productividad General', f"{productividad_general}%"]
    ]
    
    resumen_table = Table(resumen_data, colWidths=[8*cm, 4*cm])
    resumen_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(resumen_table)
    
    elements.append(PageBreak())
    
    # Detalle por Especialidad
    elements.append(Paragraph("Análisis por Especialidad", styles['Heading1']))
    elements.append(Spacer(1, 0.5*cm))
    
    data = [['Especialidad', 'Total', 'Realizadas', 'Pendientes', 'Alta', 'Media', 'Baja', 'Productividad']]
    
    for depto, stats in sorted(deptos_data.items()):
        prod = int((stats['realizadas'] / stats['total'] * 100)) if stats['total'] > 0 else 0
        data.append([
            depto,
            str(stats['total']),
            str(stats['realizadas']),
            str(stats['pendientes']),
            str(stats['alta']),
            str(stats['media']),
            str(stats['baja']),
            f"{prod}%"
        ])
    
    table = Table(data, repeatRows=1, colWidths=[3*cm, 1.5*cm, 2*cm, 2*cm, 1.5*cm, 1.5*cm, 1.5*cm, 2*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    elements.append(table)
    
    # Pie de página
    elements.append(Spacer(1, 2*cm))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=1
    )
    elements.append(Paragraph("<hr width='100%' color='#bdc3c7' size='1'>", styles['Normal']))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(Paragraph(f"Clínica Médica - Informe generado el {hoy.strftime('%d/%m/%Y a las %H:%M')}", footer_style))
    elements.append(Paragraph("Documento confidencial - Solo para uso interno", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Clinica_Medica_Informe_{hoy.strftime("%Y%m")}.pdf'
    )


# ========== GESTIÓN DE RECETAS ==========

@app.route('/receta/nueva', methods=['GET', 'POST'])
@login_required
def nueva_receta():
    user = db.session.get(Usuario, session.get('user_id'))
    paciente_id = request.args.get('paciente_id', '')
    
    if request.method == 'POST':
        # Procesar medicamentos
        medicamentos = []
        nombres = request.form.getlist('medicamento_nombre[]')
        dosis = request.form.getlist('medicamento_dosis[]')
        frecuencias = request.form.getlist('medicamento_frecuencia[]')
        duraciones = request.form.getlist('medicamento_duracion[]')
        
        for i in range(len(nombres)):
            if nombres[i].strip():
                medicamentos.append({
                    'nombre': nombres[i],
                    'dosis': dosis[i],
                    'frecuencia': frecuencias[i],
                    'duracion': duraciones[i]
                })
        
        receta = Receta(
            paciente_id=request.form['paciente_id'],
            medico_id=user.id,
            diagnostico=request.form.get('diagnostico', ''),
            medicamentos=json.dumps(medicamentos),
            indicaciones=request.form.get('indicaciones', ''),
            proxima_cita=request.form.get('proxima_cita', '')
        )
        db.session.add(receta)
        db.session.commit()
        flash('✅ Receta emitida correctamente')
        return redirect(f'/receta/{receta.id}')
    
    # Obtener pacientes
    if session.get('rol') == 'admin':
        pacientes = Paciente.query.order_by(Paciente.nombre).all()
    else:
        pacientes = Paciente.query.filter_by(usuario_id=user.id).order_by(Paciente.nombre).all()
    
    pacientes_options = "".join([f'<option value="{p.id}" {"selected" if str(p.id) == paciente_id else ""}>{p.nombre}</option>' for p in pacientes])
    
    content = f"""
        <h2>📋 Nueva Receta Médica</h2>
        
        <form method="POST" id="recetaForm">
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 20px;">
                <div class="card">
                    <h3>👤 Datos del Paciente</h3>
                    <div class="form-group">
                        <label>Paciente *</label>
                        <select name="paciente_id" class="form-control" required>
                            <option value="">Seleccionar paciente...</option>
                            {pacientes_options}
                        </select>
                    </div>
                    
                    <div class="form-group">
                        <label>🔬 Diagnóstico</label>
                        <textarea name="diagnostico" class="form-control" rows="4" placeholder="Diagnóstico médico..."></textarea>
                    </div>
                    
                    <div class="form-group">
                        <label>📝 Indicaciones generales</label>
                        <textarea name="indicaciones" class="form-control" rows="3" placeholder="Reposo, cuidados especiales..."></textarea>
                    </div>
                    
                    <div class="form-group">
                        <label>📅 Próxima cita (opcional)</label>
                        <input type="date" name="proxima_cita" class="form-control">
                    </div>
                </div>
                
                <div class="card">
                    <h3>💊 Medicamentos</h3>
                    <div id="medicamentos-container">
                        <div class="medicamento-item" style="background: #f8f9fa; padding: 15px; border-radius: 12px; margin-bottom: 15px;">
                            <div class="form-group">
                                <label>Nombre del medicamento</label>
                                <input type="text" name="medicamento_nombre[]" class="form-control" placeholder="Ej: Paracetamol 500mg">
                            </div>
                            <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px;">
                                <div class="form-group">
                                    <label>Dosis</label>
                                    <input type="text" name="medicamento_dosis[]" class="form-control" placeholder="Ej: 1 comprimido">
                                </div>
                                <div class="form-group">
                                    <label>Frecuencia</label>
                                    <input type="text" name="medicamento_frecuencia[]" class="form-control" placeholder="Ej: Cada 8 horas">
                                </div>
                                <div class="form-group">
                                    <label>Duración</label>
                                    <input type="text" name="medicamento_duracion[]" class="form-control" placeholder="Ej: 7 días">
                                </div>
                            </div>
                        </div>
                    </div>
                    
                    <button type="button" class="btn btn-primary" onclick="agregarMedicamento()" style="margin-bottom: 20px;">
                        ➕ Añadir otro medicamento
                    </button>
                </div>
            </div>
            
            <div style="display: flex; gap: 10px; margin-top: 20px;">
                <button type="submit" class="btn btn-success" style="padding: 14px 32px; font-size: 16px;">
                    💾 Guardar Receta
                </button>
                <a href="/recetas" class="btn" style="background:#95a5a6;color:white; padding: 14px 32px;">Cancelar</a>
            </div>
        </form>
        
        <script>
            function agregarMedicamento() {{
                const container = document.getElementById('medicamentos-container');
                const nuevo = document.createElement('div');
                nuevo.className = 'medicamento-item';
                nuevo.style.cssText = 'background: #f8f9fa; padding: 15px; border-radius: 12px; margin-bottom: 15px;';
                nuevo.innerHTML = `
                    <div class="form-group">
                        <label>Nombre del medicamento</label>
                        <input type="text" name="medicamento_nombre[]" class="form-control" placeholder="Ej: Paracetamol 500mg">
                    </div>
                    <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px;">
                        <div class="form-group">
                            <label>Dosis</label>
                            <input type="text" name="medicamento_dosis[]" class="form-control" placeholder="Ej: 1 comprimido">
                        </div>
                        <div class="form-group">
                            <label>Frecuencia</label>
                            <input type="text" name="medicamento_frecuencia[]" class="form-control" placeholder="Ej: Cada 8 horas">
                        </div>
                        <div class="form-group">
                            <label>Duración</label>
                            <input type="text" name="medicamento_duracion[]" class="form-control" placeholder="Ej: 7 días">
                        </div>
                    </div>
                    <button type="button" class="btn btn-danger btn-sm" onclick="this.parentElement.remove()" style="margin-top: 10px;">🗑️ Eliminar</button>
                `;
                container.appendChild(nuevo);
            }}
        </script>
    """
    return base_html(content, "Nueva Receta")

@app.route('/receta/<int:id>')
@login_required
def ver_receta(id):
    receta = Receta.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    # Verificar permisos
    if session.get('rol') != 'admin' and receta.medico_id != user.id:
        flash('❌ No tienes permiso para ver esta receta')
        return redirect('/recetas')
    
    medicamentos = json.loads(receta.medicamentos) if receta.medicamentos else []
    
    medicamentos_html = ""
    for m in medicamentos:
        medicamentos_html += f"""
        <tr>
            <td>{m.get('nombre', '')}</td>
            <td>{m.get('dosis', '')}</td>
            <td>{m.get('frecuencia', '')}</td>
            <td>{m.get('duracion', '')}</td>
        </tr>
        """
    
    # Versión imprimible
    imprimible = request.args.get('imprimir', False)
    
    if imprimible:
        content = f"""
        <div style="max-width: 800px; margin: 40px auto; background: white; padding: 40px; border-radius: 20px; box-shadow: 0 10px 40px rgba(0,0,0,0.1);">
            <div style="display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid #1a5276; padding-bottom: 20px; margin-bottom: 30px;">
                <div>
                    <h1 style="color: #1a5276; margin: 0; font-size: 32px;">🏥 CLÍNICA MÉDICA</h1>
                    <p style="color: #666; margin: 5px 0 0;">Receta Médica Oficial</p>
                </div>
                <div style="text-align: right;">
                    <p style="margin: 0; font-weight: bold;">Nº {receta.id:06d}</p>
                    <p style="margin: 5px 0 0;">{receta.fecha_emision}</p>
                </div>
            </div>
            
            <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 30px; margin-bottom: 30px;">
                <div>
                    <h3 style="color: #1a5276; border-bottom: 1px solid #ddd; padding-bottom: 5px;">👤 Paciente</h3>
                    <p><strong>Nombre:</strong> {receta.paciente.nombre}</p>
                    <p><strong>Edad:</strong> {receta.paciente.fecha_nacimiento or 'No registrada'}</p>
                    <p><strong>Obra Social:</strong> {receta.paciente.obra_social or 'Particular'}</p>
                </div>
                <div>
                    <h3 style="color: #1a5276; border-bottom: 1px solid #ddd; padding-bottom: 5px;">👨‍⚕️ Médico</h3>
                    <p><strong>Dr/a:</strong> {receta.medico.nombre_completo}</p>
                    <p><strong>Especialidad:</strong> {receta.medico.departamento}</p>
                    <p><strong>Matrícula:</strong> MN-{receta.medico.id:05d}</p>
                </div>
            </div>
            
            <div style="margin-bottom: 30px;">
                <h3 style="color: #1a5276; border-bottom: 1px solid #ddd; padding-bottom: 5px;">🔬 Diagnóstico</h3>
                <p style="background: #f8f9fa; padding: 15px; border-radius: 8px;">{receta.diagnostico or 'No especificado'}</p>
            </div>
            
            <div style="margin-bottom: 30px;">
                <h3 style="color: #1a5276; border-bottom: 1px solid #ddd; padding-bottom: 5px;">💊 Medicamentos Recetados</h3>
                <table style="width: 100%; border-collapse: collapse;">
                    <thead>
                        <tr style="background: #1a5276; color: white;">
                            <th style="padding: 10px; text-align: left;">Medicamento</th>
                            <th style="padding: 10px; text-align: left;">Dosis</th>
                            <th style="padding: 10px; text-align: left;">Frecuencia</th>
                            <th style="padding: 10px; text-align: left;">Duración</th>
                        </tr>
                    </thead>
                    <tbody>
                        {medicamentos_html if medicamentos else '<tr><td colspan="4" style="padding: 20px; text-align: center; color: #999;">No hay medicamentos recetados</td></tr>'}
                    </tbody>
                </table>
            </div>
            
            <div style="margin-bottom: 30px;">
                <h3 style="color: #1a5276; border-bottom: 1px solid #ddd; padding-bottom: 5px;">📝 Indicaciones</h3>
                <p style="background: #f8f9fa; padding: 15px; border-radius: 8px;">{receta.indicaciones or 'Sin indicaciones adicionales'}</p>
            </div>
            
            {f'''
            <div style="margin-bottom: 30px;">
                <h3 style="color: #1a5276; border-bottom: 1px solid #ddd; padding-bottom: 5px;">📅 Próxima Cita</h3>
                <p style="font-size: 18px; font-weight: bold; color: #27ae60;">{receta.proxima_cita}</p>
            </div>
            ''' if receta.proxima_cita else ''}
            
            <div style="display: flex; justify-content: space-between; margin-top: 50px; padding-top: 30px; border-top: 1px dashed #ccc;">
                <div style="text-align: center;">
                    <div style="width: 200px; border-bottom: 1px solid #333; margin-bottom: 5px;"></div>
                    <p>Firma del Médico</p>
                </div>
                <div style="text-align: center;">
                    <div style="width: 200px; border-bottom: 1px solid #333; margin-bottom: 5px;"></div>
                    <p>Sello de la Clínica</p>
                </div>
            </div>
            
            <p style="text-align: center; margin-top: 40px; color: #999; font-size: 12px;">
                Este documento es una receta médica oficial. Válido por 30 días desde la fecha de emisión.
            </p>
        </div>
        
        <div style="text-align: center; margin: 30px;">
            <button onclick="window.print()" class="btn btn-primary" style="padding: 14px 32px; font-size: 16px;">🖨️ Imprimir Receta</button>
            <a href="/receta/{receta.id}/pdf" class="btn btn-success" style="padding: 14px 32px; font-size: 16px;">📄 Descargar PDF</a>
        </div>
        """
        return base_html(content, f"Receta #{receta.id:06d}")
    
    # Vista normal
    content = f"""
        <h2>📋 Receta Médica #{receta.id:06d}</h2>
        
        <div class="card">
            <div style="display: flex; gap: 20px; margin-bottom: 20px;">
                <a href="/receta/{receta.id}?imprimir=1" class="btn btn-primary" target="_blank">🖨️ Vista para Imprimir</a>
                <a href="/receta/{receta.id}/pdf" class="btn btn-success">📄 Descargar PDF</a>
                <a href="/receta/{receta.id}/enviar-email" class="btn btn-primary">📧 Enviar por Email</a>
                <a href="/recetas" class="btn" style="background:#95a5a6;color:white;">← Volver</a>
            </div>
            
            <iframe src="/receta/{receta.id}?imprimir=1" style="width: 100%; height: 800px; border: none; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.1);"></iframe>
        </div>
    """
    return base_html(content, f"Receta #{receta.id:06d}")   

@app.route('/receta/<int:id>/pdf')
@login_required
def descargar_receta_pdf(id):
    receta = Receta.query.get_or_404(id)
    
    # Crear PDF con ReportLab
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Cabecera
    elements.append(Paragraph("CLÍNICA MÉDICA", styles['Title']))
    elements.append(Paragraph(f"Receta Nº {receta.id:06d}", styles['Heading2']))
    elements.append(Spacer(1, 1*cm))
    
    # Datos paciente y médico
    data = [
        ['Paciente:', receta.paciente.nombre, 'Médico:', f"Dr/a. {receta.medico.nombre_completo}"],
        ['Fecha:', receta.fecha_emision, 'Especialidad:', receta.medico.departamento],
    ]
    table = Table(data, colWidths=[3*cm, 6*cm, 3*cm, 6*cm])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 12),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 1*cm))
    
    # Diagnóstico
    elements.append(Paragraph("Diagnóstico:", styles['Heading3']))
    elements.append(Paragraph(receta.diagnostico or 'No especificado', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Medicamentos
    elements.append(Paragraph("Medicamentos Recetados:", styles['Heading3']))
    medicamentos = json.loads(receta.medicamentos) if receta.medicamentos else []
    if medicamentos:
        med_data = [['Medicamento', 'Dosis', 'Frecuencia', 'Duración']]
        for m in medicamentos:
            med_data.append([m.get('nombre', ''), m.get('dosis', ''), m.get('frecuencia', ''), m.get('duracion', '')])
        med_table = Table(med_data, colWidths=[5*cm, 3*cm, 4*cm, 3*cm])
        med_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a5276')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(med_table)
    else:
        elements.append(Paragraph("No hay medicamentos recetados", styles['Normal']))
    
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("Indicaciones:", styles['Heading3']))
    elements.append(Paragraph(receta.indicaciones or 'Sin indicaciones', styles['Normal']))
    
    doc.build(elements)
    buffer.seek(0)
    
    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'Receta_{receta.id:06d}_{receta.paciente.nombre}.pdf'
    )


@app.route('/recetas')
@login_required
def recetas():
    user = db.session.get(Usuario, session.get('user_id'))
    
    if session.get('rol') == 'admin':
        recetas_list = Receta.query.order_by(Receta.id.desc()).all()
    else:
        recetas_list = Receta.query.filter_by(medico_id=user.id).order_by(Receta.id.desc()).all()
    
    recetas_html = ""
    for r in recetas_list[:30]:
        medicamentos_count = len(json.loads(r.medicamentos)) if r.medicamentos else 0
        
        recetas_html += f"""
        <tr>
            <td><strong>#{r.id:06d}</strong></td>
            <td>{r.paciente.nombre}</td>
            <td>{r.fecha_emision}</td>
            <td><span class="badge badge-success">{r.estado}</span></td>
            <td>{medicamentos_count} medicamentos</td>
            <td>
                <a href="/receta/{r.id}" class="btn btn-primary btn-sm">👁️ Ver</a>
                <a href="/receta/{r.id}?imprimir=1" class="btn btn-success btn-sm" target="_blank">🖨️</a>
            </td>
        </tr>
        """
    
    content = f"""
        <h2>📋 Mis Recetas Emitidas</h2>
        <p><a href="/receta/nueva" class="btn btn-success">➕ Nueva Receta</a></p>
        
        <table>
            <thead>
                <tr>
                    <th>Nº Receta</th>
                    <th>Paciente</th>
                    <th>Fecha</th>
                    <th>Estado</th>
                    <th>Medicamentos</th>
                    <th>Acciones</th>
                </tr>
            </thead>
            <tbody>
                {recetas_html if recetas_list else '<tr><td colspan="6" style="text-align:center; padding:30px;">No hay recetas emitidas</td></tr>'}
            </tbody>
        </table>
    """
    return base_html(content, "Recetas")

@app.route('/receta/<int:id>/enviar-email', methods=['GET', 'POST'])
@login_required
def enviar_receta_email_route(id):
    receta = Receta.query.get_or_404(id)
    user = db.session.get(Usuario, session.get('user_id'))
    
    # Verificar permisos
    if session.get('rol') != 'admin' and receta.medico_id != user.id:
        flash('❌ No tienes permiso')
        return redirect('/recetas')
    
    if not receta.paciente.email:
        flash('❌ El paciente no tiene email registrado')
        return redirect(f'/receta/{id}')
    
    if request.method == 'POST':
        email_destino = request.form.get('email', receta.paciente.email)
        mensaje_adicional = request.form.get('mensaje', '')
        
        # Crear mensaje HTML para el email
        mensaje_html = f"""
        <html>
        <head>
            <style>
                body {{ font-family: Arial, sans-serif; }}
                .header {{ background: #1a5276; color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 20px; }}
                .footer {{ background: #f5f5f5; padding: 15px; text-align: center; font-size: 12px; color: #666; }}
                .btn {{ background: #27ae60; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>🏥 CLÍNICA MÉDICA</h1>
                <p>Receta Médica Electrónica</p>
            </div>
            <div class="content">
                <h2>Hola {receta.paciente.nombre},</h2>
                <p>El Dr/a. <strong>{receta.medico.nombre_completo}</strong> te ha enviado una receta médica.</p>
                
                {f'<p><strong>Mensaje del médico:</strong><br>{mensaje_adicional}</p>' if mensaje_adicional else ''}
                
                <p>Adjunto a este correo encontrarás la receta en formato PDF con todos los detalles de tu tratamiento.</p>
                
                <p style="margin-top: 30px;">
                    <strong>Resumen de la receta:</strong><br>
                    📋 Nº Receta: {receta.id:06d}<br>
                    📅 Fecha: {receta.fecha_emision}<br>
                    🔬 Diagnóstico: {receta.diagnostico or 'No especificado'}
                </p>
                
                <p style="margin-top: 30px;">
                    Si tienes alguna duda, puedes contactarnos o solicitar una nueva cita.
                </p>
            </div>
            <div class="footer">
                <p>Clínica Médica | Tel: (011) 1234-5678 | Email: contacto@clinicamedica.com</p>
                <p>Este es un mensaje automático. Por favor, no respondas a este correo.</p>
            </div>
        </body>
        </html>
        """
        
        # Generar PDF
        pdf_buffer = generar_pdf_receta(receta)
        
        # Enviar email
        asunto = f"Receta Médica #{receta.id:06d} - Clínica Médica"
        nombre_archivo = f"Receta_{receta.id:06d}_{receta.paciente.nombre.replace(' ', '_')}.pdf"
        
        exito, mensaje = enviar_receta_email(
            email_destino,
            asunto,
            mensaje_html,
            pdf_buffer,
            nombre_archivo
        )
        
        if exito:
            flash(f'✅ Receta enviada correctamente a {email_destino}')
        else:
            flash(f'❌ Error al enviar: {mensaje}')
        
        return redirect(f'/receta/{id}')
    
    # GET - Mostrar formulario de envío
    content = f"""
        <h2>📧 Enviar Receta por Email</h2>
        
        <div class="card">
            <h3>📋 Receta #{receta.id:06d}</h3>
            <p><strong>Paciente:</strong> {receta.paciente.nombre}</p>
            <p><strong>Email registrado:</strong> {receta.paciente.email}</p>
            <p><strong>Médico:</strong> Dr/a. {receta.medico.nombre_completo}</p>
            
            <form method="POST" style="margin-top: 30px;">
                <div class="form-group">
                    <label>📧 Email de destino</label>
                    <input type="email" name="email" class="form-control" value="{receta.paciente.email}" required>
                    <small style="color: #666;">Puedes cambiarlo si el paciente prefiere otro email</small>
                </div>
                
                <div class="form-group">
                    <label>💬 Mensaje adicional (opcional)</label>
                    <textarea name="mensaje" class="form-control" rows="4" placeholder="Añade un mensaje personal para el paciente..."></textarea>
                </div>
                
                <div style="display: flex; gap: 10px; margin-top: 20px;">
                    <button type="submit" class="btn btn-success" style="padding: 12px 24px;">
                        📧 Enviar Receta por Email
                    </button>
                    <a href="/receta/{receta.id}" class="btn" style="background:#95a5a6;color:white; padding: 12px 24px;">
                        Cancelar
                    </a>
                </div>
            </form>
            
            <div style="margin-top: 30px; padding: 15px; background: #e8f4f8; border-radius: 8px;">
                <p style="margin: 0; color: #1a5276;">
                    <strong>ℹ️ Vista previa:</strong> El paciente recibirá un email con el formato profesional de la clínica y la receta adjunta en PDF.
                </p>
            </div>
        </div>
    """
    return base_html(content, f"Enviar Receta #{receta.id:06d}")



    db.create_all()
    
    # Crear backup automático al iniciar (cada día)
    backup_dir = 'instance/backups'
    if not os.path.exists(backup_dir):
        os.makedirs(backup_dir)
    
    hoy = datetime.now().strftime('%Y%m%d')
    backups_hoy = [f for f in os.listdir(backup_dir) if hoy in f and 'auto' in f] if os.path.exists(backup_dir) else []
    
    if not backups_hoy and os.path.exists('instance/clinica.db'):
        crear_backup(tipo='auto_inicio')
        print("🤖 Backup automático de inicio creado")
    
    # Crear admin por defecto si no existe
    if Usuario.query.count() == 0:
        admin = Usuario(
            username='admin',
            password_hash=generate_password_hash('admin123', method='pbkdf2:sha256'),
            nombre_completo='Administrador del Sistema',
            departamento='Dirección Médica',
            cargo='Director',
            rol='admin'
        )
        db.session.add(admin)
        db.session.commit()
        print("✅ Admin creado: admin / admin123")
    
     # ========== CREAR TIPOS DE HISTORIAL POR DEFECTO ==========
    if TipoHistorial.query.count() == 0:
        tipos_default = [
            {'nombre': 'Consulta General', 'descripcion': 'Consulta médica general o primera visita', 'color': '#3498db'},
            {'nombre': 'Revisión', 'descripcion': 'Revisión periódica o control rutinario', 'color': '#27ae60'},
            {'nombre': 'Urgencia', 'descripcion': 'Atención de urgencias o emergencias', 'color': '#e74c3c'},
            {'nombre': 'Especialista', 'descripcion': 'Derivación a médico especialista', 'color': '#9b59b6'},
            {'nombre': 'Análisis Clínicos', 'descripcion': 'Resultados de análisis de laboratorio', 'color': '#f39c12'},
            {'nombre': 'Radiología', 'descripcion': 'Radiografías, ecografías, TAC', 'color': '#1abc9c'},
            {'nombre': 'Control Cardiológico', 'descripcion': 'Control del corazón y sistema circulatorio', 'color': '#e74c3c'},
            {'nombre': 'Pediatría', 'descripcion': 'Consulta para niños y adolescentes', 'color': '#ff6b6b'},
        ]
        for t in tipos_default:
            tipo = TipoHistoriaClinica(
                nombre=t['nombre'],
                descripcion=t['descripcion'],
                color=t['color'],
                admin_id=1
            )
            db.session.add(tipo)
        db.session.commit()
        print("✅ Tipos de historial clínico creados")

# ========== GESTIÓN DE BACKUPS ==========

@app.route('/admin/backups')
@login_required
def admin_backups():
    if session.get('rol') != 'admin':
        flash('❌ Acceso restringido')
        return redirect('/dashboard')
    
    backup_dir = 'instance/backups'
    backups = []
    
    if os.path.exists(backup_dir):
        for f in sorted(os.listdir(backup_dir), reverse=True):
            if f.endswith('.db'):
                path = os.path.join(backup_dir, f)
                size = os.path.getsize(path)
                
                if 'auto' in f:
                    tipo = '🤖 Automático'
                    color = '#3498db'
                elif 'manual' in f:
                    tipo = '👤 Manual'
                    color = '#27ae60'
                else:
                    tipo = '📦 Backup'
                    color = '#95a5a6'
                
                try:
                    partes = f.replace('clinica_backup_', '').replace('.db', '').split('_')
                    if len(partes) >= 2:
                        fecha_str = f"{partes[0][:4]}-{partes[0][4:6]}-{partes[0][6:8]} {partes[1][:2]}:{partes[1][2:4]}"
                    else:
                        fecha_str = f.replace('clinica_backup_', '').replace('.db', '')
                except:
                    fecha_str = 'Desconocida'
                
                backups.append({
                    'nombre': f,
                    'fecha': fecha_str,
                    'tipo': tipo,
                    'color': color,
                    'tamaño_str': f"{size/1024:.1f} KB" if size < 1024*1024 else f"{size/(1024*1024):.2f} MB"
                })
    
    # Estado actual
    db_existe = os.path.exists('instance/clinica.db')
    db_size = os.path.getsize('instance/clinica.db') / 1024 if db_existe else 0
    pacientes_count = Paciente.query.count()
    citas_count = Cita.query.count()
    recetas_count = Receta.query.count() if 'Receta' in globals() else 0
    total_backups = len(backups)
    
    backups_html = ""
    for b in backups[:20]:
        backups_html += f"""
        <tr>
            <td><span style="background: {b['color']}; color: white; padding: 4px 12px; border-radius: 20px; font-size: 12px;">{b['tipo']}</span></td>
            <td>{b['fecha']}</td>
            <td>{b['tamaño_str']}</td>
            <td>
                <a href="/admin/restaurar-backup/{b['nombre']}" class="btn btn-warning btn-sm" 
                   onclick="return confirm('⚠️ ¿Restaurar este backup? Los datos actuales serán REEMPLAZADOS.')">
                    🔄 Restaurar
                </a>
                <a href="/admin/descargar-backup/{b['nombre']}" class="btn btn-primary btn-sm">📥 Descargar</a>
                <a href="/admin/eliminar-backup/{b['nombre']}" class="btn btn-danger btn-sm" 
                   onclick="return confirm('¿Eliminar este backup permanentemente?')">🗑️</a>
            </td>
        </tr>
        """
    
    content = f"""
        <h2>💾 Gestión de Backups</h2>
        
        <div class="card">
            <h3>📊 Estado Actual del Sistema</h3>
            <div style="display: grid; grid-template-columns: repeat(4, 1fr); gap: 20px;">
                <div style="text-align: center;">
                    <div style="font-size: 32px; font-weight: bold; color: {'#27ae60' if db_existe else '#e74c3c'};">{'✅' if db_existe else '❌'}</div>
                    <div>Base de datos</div>
                    <small>{db_size:.1f} KB</small>
                </div>
                <div style="text-align: center;">
                    <div style="font-size: 32px; font-weight: bold; color: #1a5276;">{pacientes_count}</div>
                    <div>Pacientes</div>
                </div>
                <div style="text-align: center;">
                    <div style="font-size: 32px; font-weight: bold; color: #1a5276;">{citas_count}</div>
                    <div>Citas</div>
                </div>
                <div style="text-align: center;">
                    <div style="font-size: 32px; font-weight: bold; color: #1a5276;">{recetas_count}</div>
                    <div>Recetas</div>
                </div>
            </div>
        </div>
        
        <div class="card" style="margin-top: 20px;">
            <h3>⚡ Acciones Rápidas</h3>
            <div style="display: flex; gap: 15px; flex-wrap: wrap;">
                <a href="/admin/crear-backup" class="btn btn-success" style="padding: 14px 24px;">
                    💾 Crear Backup Manual Ahora
                </a>
                <a href="/admin/exportar-todo" class="btn btn-primary" style="padding: 14px 24px;">
                    📊 Exportar a Excel
                </a>
                <a href="/admin/crear-backup-auto" class="btn" style="background: #3498db; color: white; padding: 14px 24px;">
                    🤖 Crear Backup Automático
                </a>
            </div>
        </div>
        
        <div class="card" style="margin-top: 20px;">
            <h3>📂 Backups Disponibles ({total_backups})</h3>
            <table>
                <thead>
                    <tr>
                        <th>Tipo</th>
                        <th>Fecha</th>
                        <th>Tamaño</th>
                        <th>Acciones</th>
                    </tr>
                </thead>
                <tbody>
                    {backups_html if backups else '<tr><td colspan="4" style="text-align:center; padding:30px;">📭 No hay backups disponibles. ¡Crea uno!</td></tr>'}
                </tbody>
            </table>
        </div>
        
        <div class="card" style="margin-top: 20px; background: #e8f4f8;">
            <h3>📖 Guía de Backups</h3>
            <ul style="line-height: 1.8;">
                <li><strong>💾 Backup Manual:</strong> Haz clic antes de cambios importantes.</li>
                <li><strong>🤖 Backup Automático:</strong> Se crea al iniciar la aplicación cada día.</li>
                <li><strong>🔄 Restaurar:</strong> Vuelve a un estado anterior (los datos actuales se reemplazan).</li>
                <li><strong>📥 Descargar:</strong> Guarda una copia en tu ordenador.</li>
                <li><strong>🗑️ Eliminar:</strong> Borra backups antiguos para liberar espacio.</li>
            </ul>
        </div>
    """
    return base_html(content, "Gestión de Backups")


@app.route('/admin/crear-backup')
@login_required
def crear_backup_manual():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    path, mensaje = crear_backup(tipo='manual')
    if path:
        flash(f'✅ Backup manual creado: {os.path.basename(path)}')
    else:
        flash(f'❌ {mensaje}')
    return redirect('/admin/backups')


@app.route('/admin/crear-backup-auto')
@login_required
def crear_backup_auto():
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    path, mensaje = crear_backup(tipo='auto')
    if path:
        flash(f'🤖 Backup automático creado: {os.path.basename(path)}')
    else:
        flash(f'❌ {mensaje}')
    return redirect('/admin/backups')


@app.route('/admin/descargar-backup/<filename>')
@login_required
def descargar_backup(filename):
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    from flask import send_from_directory
    backup_dir = 'instance/backups'
    return send_from_directory(backup_dir, filename, as_attachment=True)


@app.route('/admin/restaurar-backup/<filename>')
@login_required
def restaurar_backup(filename):
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    backup_path = os.path.join('instance/backups', filename)
    db_path = 'instance/clinica.db'
    
    if not os.path.exists(backup_path):
        flash('❌ Backup no encontrado')
        return redirect('/admin/backups')
    
    # Crear backup de emergencia antes de restaurar
    emergencia_path, _ = crear_backup(tipo='pre_restore')
    
    try:
        shutil.copy2(backup_path, db_path)
        flash(f'✅ Base de datos restaurada desde {filename}')
        if emergencia_path:
            flash(f'💾 Se creó backup de emergencia: {os.path.basename(emergencia_path)}')
    except Exception as e:
        flash(f'❌ Error al restaurar: {str(e)}')
    
    return redirect('/admin/backups')


@app.route('/admin/eliminar-backup/<filename>')
@login_required
def eliminar_backup(filename):
    if session.get('rol') != 'admin':
        return redirect('/dashboard')
    
    backup_path = os.path.join('instance/backups', filename)
    
    if os.path.exists(backup_path):
        os.remove(backup_path)
        flash(f'✅ Backup eliminado: {filename}')
    else:
        flash('❌ Backup no encontrado')
    
    return redirect('/admin/backups')

# ========== INICIALIZACIÓN ==========

with app.app_context():
    db.create_all()
    
    # Crear admin
    if Usuario.query.count() == 0:
        admin = Usuario(
            username='admin',
            password_hash=generate_password_hash('admin123', method='pbkdf2:sha256'),
            nombre_completo='Administrador del Sistema',
            departamento='Dirección Médica',
            cargo='Director',
            rol='admin'
        )
        db.session.add(admin)
        db.session.commit()
        print("🔐 Admin creado: admin / admin123")
    
    # Crear tipos de historia clínica
    if TipoHistoriaClinica.query.count() == 0:
        tipos_default = [
            {'nombre': 'Historia Clínica General', 'descripcion': 'Historia clínica general del paciente', 'color': '#3498db'},
            {'nombre': 'Control Cardiológico', 'descripcion': 'Seguimiento cardiológico', 'color': '#e74c3c'},
            {'nombre': 'Control Pediátrico', 'descripcion': 'Seguimiento pediátrico', 'color': '#27ae60'},
            {'nombre': 'Informe Radiológico', 'descripcion': 'Estudios de imagen', 'color': '#9b59b6'},
        ]
        for t in tipos_default:
            tipo = TipoHistoriaClinica(
                nombre=t['nombre'],
                descripcion=t['descripcion'],
                color=t['color'],
                admin_id=1
            )
            db.session.add(tipo)
        db.session.commit()
        print("✅ Tipos de historia clínica creados")
    
    print("🏥 ¡Clínica Médica lista para usar!")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)